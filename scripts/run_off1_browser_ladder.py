#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import statistics
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any


def utc_ts() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def parse_float(value: str, default: float) -> float:
    try:
        return float(value)
    except Exception:
        return default


def parse_int(value: str, default: int) -> int:
    try:
        return int(value)
    except Exception:
        return default


@dataclass
class CmdResult:
    args: list[str]
    returncode: int
    stdout: str
    stderr: str


def run_cmd(args: list[str], timeout_s: float = 60.0) -> CmdResult:
    try:
        cp = subprocess.run(args, capture_output=True, text=True, timeout=timeout_s)
        return CmdResult(args=args, returncode=cp.returncode, stdout=cp.stdout or "", stderr=cp.stderr or "")
    except subprocess.TimeoutExpired as exc:
        return CmdResult(args=args, returncode=124, stdout=exc.stdout or "", stderr=(exc.stderr or "") + "\n[timeout]")
    except FileNotFoundError:
        return CmdResult(args=args, returncode=127, stdout="", stderr=f"{args[0]}: command not found")


def adb_cmd(adb_bin: str, serial: str, args: list[str], timeout_s: float = 60.0) -> CmdResult:
    cmd = [adb_bin]
    if serial.strip():
        cmd.extend(["-s", serial.strip()])
    cmd.extend(args)
    return run_cmd(cmd, timeout_s=timeout_s)


def extract_status_text(res: CmdResult) -> str:
    text = (res.stdout + "\n" + res.stderr).strip()
    if not text:
        return ""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for ln in lines:
        if "error" in ln.lower() or "status:" in ln.lower() or "starting:" in ln.lower():
            return ln
    return lines[-1]


def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    sorted_vals = sorted(values)
    rank = (len(sorted_vals) - 1) * p
    lo = math.floor(rank)
    hi = math.ceil(rank)
    if lo == hi:
        return sorted_vals[int(rank)]
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (rank - lo)


def detect_authorized_devices(devices_text: str) -> int:
    count = 0
    for ln in devices_text.splitlines():
        ln = ln.strip()
        if not ln or ln.lower().startswith("list of devices"):
            continue
        parts = ln.split()
        if len(parts) >= 2 and parts[1] == "device":
            count += 1
    return count


def get_top_activity(adb_bin: str, serial: str) -> str:
    cmd = "dumpsys activity top | grep -iE 'mResumedActivity|topResumedActivity' | head -n 1"
    res = adb_cmd(adb_bin, serial, ["shell", cmd], timeout_s=20.0)
    return (res.stdout or "").strip() or (res.stderr or "").strip()


def build_launch_url(base_url: str, launch_counter: int) -> str:
    base = (base_url or "").strip()
    if base.startswith("file://"):
        return f"{base}#off1-tab-{launch_counter}"
    sep = "&" if "?" in base else "?"
    return f"{base}{sep}off1_tab={launch_counter}"


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def append_master_csv(master_path: Path, rows: list[dict[str, Any]]) -> None:
    master_path.parent.mkdir(parents=True, exist_ok=True)
    exists = master_path.exists()
    fields = [
        "run_id",
        "generated_utc",
        "url",
        "browser_package",
        "rung",
        "attempts",
        "successes",
        "success_rate_percent",
        "total_launches_after_rung",
        "rung_total_ms",
        "rung_total_s",
        "avg_ms",
        "median_ms",
        "p95_ms",
        "top_activity",
    ]
    with master_path.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def maybe_write_xlsx(output_dir: Path, run_id: str, detail_rows: list[dict[str, Any]], rung_rows: list[dict[str, Any]]) -> Path | None:
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except Exception:
        return None

    run_xlsx = output_dir / "reports" / "off1_browser_ladder.xlsx"
    run_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "OFF1_Detail"
    if detail_rows:
        headers = list(detail_rows[0].keys())
        ws_d.append(headers)
        for r in detail_rows:
            ws_d.append([r.get(h, "") for h in headers])

    ws_s = wb.create_sheet("OFF1_Summary")
    if rung_rows:
        headers = list(rung_rows[0].keys())
        ws_s.append(headers)
        for r in rung_rows:
            ws_s.append([r.get(h, "") for h in headers])
    wb.save(run_xlsx)

    master = output_dir / "offensive_test_master.xlsx"
    if master.exists():
        mwb = load_workbook(master)
    else:
        mwb = Workbook()
        mwb.active.title = "OFF1"
    if "OFF1" not in mwb.sheetnames:
        mwb.create_sheet("OFF1")
    ws = mwb["OFF1"]
    headers = [
        "run_id",
        "generated_utc",
        "url",
        "browser_package",
        "rung",
        "attempts",
        "successes",
        "success_rate_percent",
        "total_launches_after_rung",
        "rung_total_ms",
        "rung_total_s",
        "avg_ms",
        "median_ms",
        "p95_ms",
        "top_activity",
    ]
    if ws.max_row < 1 or ws.cell(1, 1).value is None:
        ws.append(headers)
    for r in rung_rows:
        ws.append([r.get(h, "") for h in headers])
    mwb.save(master)
    return run_xlsx


def main() -> int:
    ap = argparse.ArgumentParser(description="OFF1 Browser launch ladder via ADB (1..N launches)")
    ap.add_argument("--adb-bin", default="adb")
    ap.add_argument("--device-serial", default="")
    ap.add_argument("--url", required=True)
    ap.add_argument("--browser-package", default="")
    ap.add_argument("--rungs", type=int, default=5)
    ap.add_argument("--attempt-delay-s", type=float, default=0.6)
    ap.add_argument("--cold-start-per-rung", action="store_true")
    ap.add_argument("--cumulative-ladder", action="store_true", help="Rung N adds N launches on top of previous rungs.")
    ap.add_argument("--new-document", action="store_true", help="Ask Android to open each launch as a new document/task.")
    ap.add_argument("--output-root", default="test_results/offensive_tests")
    ap.add_argument("--case-id", default="OFF1-BROWSER-LADDER")
    args = ap.parse_args()

    run_id = f"{utc_ts()}-{args.case_id}"
    out_root = Path(args.output_root).expanduser().resolve()
    run_dir = out_root / run_id
    evidence = run_dir / "evidence"
    reports = run_dir / "reports"
    evidence.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    print(f"[ByteBite] OFF1 run: {run_dir}")

    devices = adb_cmd(args.adb_bin, args.device_serial, ["devices"], timeout_s=20.0)
    (evidence / "adb_devices.txt").write_text((devices.stdout or "") + "\n" + (devices.stderr or ""), encoding="utf-8")
    if devices.returncode != 0:
        print("[ByteBite] adb devices failed")
        return 2

    authorized = detect_authorized_devices(devices.stdout)
    if authorized == 0:
        print("[ByteBite] No authorized device detected")
        return 3

    detail_rows: list[dict[str, Any]] = []
    rung_rows: list[dict[str, Any]] = []
    now_iso = dt.datetime.now(dt.timezone.utc).isoformat()
    launch_counter = 0

    for rung in range(1, max(1, args.rungs) + 1):
        if args.cold_start_per_rung and args.browser_package.strip():
            adb_cmd(args.adb_bin, args.device_serial, ["shell", "am", "force-stop", args.browser_package.strip()], timeout_s=20.0)
            time.sleep(1.0)

        rung_durations_ms: list[float] = []
        rung_successes = 0
        launches_this_rung = rung

        for launch_idx in range(1, launches_this_rung + 1):
            launch_counter += 1
            launch_url = build_launch_url(args.url, launch_counter)
            start_args = ["shell", "am", "start"]
            if args.new_document:
                start_args.append("--activity-new-document")
            start_args.extend(["-a", "android.intent.action.VIEW", "-d", launch_url])
            if args.browser_package.strip():
                start_args.append(args.browser_package.strip())

            t0 = time.perf_counter()
            res = adb_cmd(
                args.adb_bin,
                args.device_serial,
                start_args,
                timeout_s=30.0,
            )
            elapsed_ms = (time.perf_counter() - t0) * 1000.0
            status_txt = extract_status_text(res)
            ok = (res.returncode == 0) and ("error" not in (status_txt or "").lower())
            rung_successes += 1 if ok else 0
            rung_durations_ms.append(elapsed_ms)

            detail_rows.append(
                {
                    "run_id": run_id,
                    "generated_utc": now_iso,
                    "rung": rung,
                    "launch_index": launch_idx,
                    "url": launch_url,
                    "browser_package": args.browser_package,
                    "returncode": res.returncode,
                    "success": ok,
                    "duration_ms": round(elapsed_ms, 3),
                    "status_text": status_txt,
                }
            )
            time.sleep(max(0.0, args.attempt_delay_s))

        top = get_top_activity(args.adb_bin, args.device_serial)
        attempts = launches_this_rung
        success_rate = (100.0 * rung_successes / attempts) if attempts else 0.0
        rung_launch_total_ms = sum(rung_durations_ms)
        rung_row = {
            "run_id": run_id,
            "generated_utc": now_iso,
            "url": args.url,
            "browser_package": args.browser_package,
            "rung": rung,
            "attempts": attempts,
            "successes": rung_successes,
            "success_rate_percent": round(success_rate, 2),
            "total_launches_after_rung": launch_counter,
            "rung_total_ms": round(rung_launch_total_ms, 3),
            "rung_total_s": round(rung_launch_total_ms / 1000.0, 3),
            "avg_ms": round(sum(rung_durations_ms) / len(rung_durations_ms), 3) if rung_durations_ms else 0.0,
            "median_ms": round(statistics.median(rung_durations_ms), 3) if rung_durations_ms else 0.0,
            "p95_ms": round(percentile(rung_durations_ms, 0.95), 3) if rung_durations_ms else 0.0,
            "top_activity": top,
        }
        rung_rows.append(rung_row)
        print(
            f"[ByteBite] rung={rung} success={rung_successes}/{attempts} "
            f"total_launches={launch_counter} "
            f"rung_total={rung_row['rung_total_s']}s "
            f"avg={rung_row['avg_ms']}ms med={rung_row['median_ms']}ms p95={rung_row['p95_ms']}ms"
        )

    detail_csv = reports / "off1_launch_details.csv"
    summary_csv = reports / "off1_rung_summary.csv"

    if detail_rows:
        with detail_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
            w.writeheader()
            w.writerows(detail_rows)

    if rung_rows:
        with summary_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rung_rows[0].keys()))
            w.writeheader()
            w.writerows(rung_rows)

    append_master_csv(out_root / "offensive_test_master.csv", rung_rows)
    run_xlsx = maybe_write_xlsx(out_root, run_id, detail_rows, rung_rows)

    payload = {
        "run_id": run_id,
        "generated_utc": now_iso,
        "url": args.url,
        "browser_package": args.browser_package,
        "rungs": args.rungs,
        "authorized_devices": authorized,
        "run_dir": str(run_dir),
        "detail_csv": str(detail_csv),
        "summary_csv": str(summary_csv),
        "master_csv": str(out_root / "offensive_test_master.csv"),
        "run_xlsx": str(run_xlsx) if run_xlsx else None,
        "master_xlsx": str(out_root / "offensive_test_master.xlsx") if (out_root / "offensive_test_master.xlsx").exists() else None,
    }
    write_json(run_dir / "summary.json", payload)

    print(f"[ByteBite] Detail CSV: {detail_csv}")
    print(f"[ByteBite] Summary CSV: {summary_csv}")
    if run_xlsx:
        print(f"[ByteBite] Run XLSX: {run_xlsx}")
    if payload["master_xlsx"]:
        print(f"[ByteBite] Master XLSX: {payload['master_xlsx']}")
    print(f"[ByteBite] Summary JSON: {run_dir / 'summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
