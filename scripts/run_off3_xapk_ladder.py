#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import math
import shlex
import statistics
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any


def utc_ts() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


@dataclass
class CmdResult:
    args: list[str]
    returncode: int
    stdout: str
    stderr: str


def run_cmd(args: list[str], timeout_s: float = 120.0) -> CmdResult:
    try:
        cp = subprocess.run(args, capture_output=True, text=True, timeout=timeout_s)
        return CmdResult(args=args, returncode=cp.returncode, stdout=cp.stdout or "", stderr=cp.stderr or "")
    except subprocess.TimeoutExpired as exc:
        return CmdResult(args=args, returncode=124, stdout=exc.stdout or "", stderr=(exc.stderr or "") + "\n[timeout]")
    except FileNotFoundError:
        return CmdResult(args=args, returncode=127, stdout="", stderr=f"{args[0]}: command not found")


def adb_cmd(adb_bin: str, serial: str, args: list[str], timeout_s: float = 120.0) -> CmdResult:
    cmd = [adb_bin]
    if serial.strip():
        cmd.extend(["-s", serial.strip()])
    cmd.extend(args)
    return run_cmd(cmd, timeout_s=timeout_s)


def detect_authorized_devices(devices_text: str) -> int:
    count = 0
    for ln in devices_text.splitlines():
        ln = ln.strip()
        if not ln or ln.lower().startswith("list of devices"):
            continue
        parts = ln.split()
        if len(parts) >= 2 and parts[1] == "device":
            count += 1
    return count


def sha256_local(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def parse_hex64(text: str) -> str:
    for token in (text or "").split():
        if len(token) == 64 and all(ch in "0123456789abcdefABCDEF" for ch in token):
            return token.lower()
    return ""


def remote_sha256(adb_bin: str, serial: str, device_path: str) -> str:
    candidates = [
        ["shell", "sha256sum", device_path],
        ["shell", "toybox", "sha256sum", device_path],
    ]
    for cmd in candidates:
        res = adb_cmd(adb_bin, serial, cmd, timeout_s=180.0)
        digest = parse_hex64(res.stdout + "\n" + res.stderr)
        if res.returncode == 0 and digest:
            return digest
    return ""


def remote_size_bytes(adb_bin: str, serial: str, device_path: str) -> int:
    cmd = f"wc -c < {shlex.quote(device_path)}"
    res = adb_cmd(adb_bin, serial, ["shell", cmd], timeout_s=60.0)
    tokens = (res.stdout or "").strip().split()
    for token in tokens:
        if token.isdigit():
            return int(token)
    return -1


def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    sorted_vals = sorted(values)
    rank = (len(sorted_vals) - 1) * p
    lo = math.floor(rank)
    hi = math.ceil(rank)
    if lo == hi:
        return sorted_vals[int(rank)]
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (rank - lo)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    import json

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def maybe_write_xlsx(run_dir: Path, detail_rows: list[dict[str, Any]], summary_rows: list[dict[str, Any]], output_root: Path) -> Path | None:
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except Exception:
        return None

    run_xlsx = run_dir / "reports" / "off3_xapk_ladder.xlsx"
    run_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "OFF3_Ladder_Detail"
    if detail_rows:
        headers = list(detail_rows[0].keys())
        ws_d.append(headers)
        for r in detail_rows:
            ws_d.append([r.get(h, "") for h in headers])

    ws_s = wb.create_sheet("OFF3_Ladder_Summary")
    if summary_rows:
        headers = list(summary_rows[0].keys())
        ws_s.append(headers)
        for r in summary_rows:
            ws_s.append([r.get(h, "") for h in headers])
    wb.save(run_xlsx)

    master = output_root / "offensive_test_master.xlsx"
    if master.exists():
        mwb = load_workbook(master)
    else:
        mwb = Workbook()
        mwb.active.title = "OFF1"
    if "OFF3_LADDER" not in mwb.sheetnames:
        mwb.create_sheet("OFF3_LADDER")
    ws = mwb["OFF3_LADDER"]
    headers = list(summary_rows[0].keys()) if summary_rows else []
    if headers and (ws.max_row < 1 or ws.cell(1, 1).value is None):
        ws.append(headers)
    for r in summary_rows:
        ws.append([r.get(h, "") for h in headers])
    mwb.save(master)

    return run_xlsx


def main() -> int:
    ap = argparse.ArgumentParser(description="OFF3 ladder: rung 1..N pushes unique filenames of same XAPK.")
    ap.add_argument("--adb-bin", default="adb")
    ap.add_argument("--device-serial", default="")
    ap.add_argument("--source-file", required=True, help="Path to the source .xapk file")
    ap.add_argument("--rungs", type=int, default=5)
    ap.add_argument("--device-dir", default="/sdcard/Download/bytebite_off3_ladder")
    ap.add_argument("--between-s", type=float, default=0.5)
    ap.add_argument("--output-root", default="test_results/offensive_tests")
    ap.add_argument("--case-id", default="OFF3-XAPK-LADDER")
    args = ap.parse_args()

    src = Path(args.source_file).expanduser().resolve()
    if not src.exists() or not src.is_file():
        print(f"[ByteBite] Missing source file: {src}")
        return 2

    run_id = f"{utc_ts()}-{args.case_id}"
    output_root = Path(args.output_root).expanduser().resolve()
    run_dir = output_root / run_id
    evidence = run_dir / "evidence"
    reports = run_dir / "reports"
    evidence.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    print(f"[ByteBite] OFF3 ladder run: {run_dir}")

    devices = adb_cmd(args.adb_bin, args.device_serial, ["devices"], timeout_s=20.0)
    (evidence / "adb_devices.txt").write_text((devices.stdout or "") + "\n" + (devices.stderr or ""), encoding="utf-8")
    if devices.returncode != 0:
        print("[ByteBite] adb devices failed")
        return 3
    if detect_authorized_devices(devices.stdout) == 0:
        print("[ByteBite] No authorized device")
        return 4

    mk = adb_cmd(args.adb_bin, args.device_serial, ["shell", "mkdir", "-p", args.device_dir], timeout_s=30.0)
    if mk.returncode != 0:
        print("[ByteBite] Could not create destination directory on device")
        return 5

    src_size = src.stat().st_size
    src_sha = sha256_local(src)
    base_stem = src.stem
    ext = src.suffix.lower()
    now_iso = dt.datetime.now(dt.timezone.utc).isoformat()

    detail_rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []

    for rung in range(1, max(1, args.rungs) + 1):
        rung_push_ms: list[float] = []
        rung_success = 0
        rung_names: list[str] = []
        for idx in range(1, rung + 1):
            dst_name = f"{base_stem}_r{rung:02d}_f{idx:02d}{ext}"
            rung_names.append(dst_name)
            dst_path = f"{args.device_dir.rstrip('/')}/{dst_name}"

            t0 = time.perf_counter()
            push = adb_cmd(args.adb_bin, args.device_serial, ["push", str(src), dst_path], timeout_s=600.0)
            push_ms = (time.perf_counter() - t0) * 1000.0
            rung_push_ms.append(push_ms)

            r_size = remote_size_bytes(args.adb_bin, args.device_serial, dst_path) if push.returncode == 0 else -1
            r_sha = remote_sha256(args.adb_bin, args.device_serial, dst_path) if push.returncode == 0 else ""
            size_match = (r_size == src_size)
            hash_match = (r_sha == src_sha and r_sha != "")
            success = (push.returncode == 0 and size_match and hash_match)
            if success:
                rung_success += 1

            detail_rows.append(
                {
                    "run_id": run_id,
                    "generated_utc": now_iso,
                    "rung": rung,
                    "rung_file_index": idx,
                    "file_name": dst_name,
                    "source_file_name": src.name,
                    "file_ext": ext,
                    "source_size_bytes": src_size,
                    "source_sha256": src_sha,
                    "device_path": dst_path,
                    "push_returncode": push.returncode,
                    "push_ms": round(push_ms, 3),
                    "remote_size_bytes": r_size,
                    "remote_sha256": r_sha,
                    "size_match": size_match,
                    "hash_match": hash_match,
                    "success": success,
                    "push_stderr": (push.stderr or "").strip(),
                }
            )
            print(f"[ByteBite] rung={rung} file={idx}/{rung} name={dst_name} success={success} push={push_ms:.3f}ms")
            time.sleep(max(0.0, args.between_s))

        attempts = rung
        total_ms = sum(rung_push_ms)
        summary_rows.append(
            {
                "run_id": run_id,
                "generated_utc": now_iso,
                "rung": rung,
                "file_name_pattern": f"{base_stem}_r{rung:02d}_fXX{ext}",
                "file_names": ", ".join(rung_names),
                "attempts": attempts,
                "successes": rung_success,
                "success_rate_percent": round((100.0 * rung_success / max(1, attempts)), 2),
                "total_time_ms": round(total_ms, 3),
                "total_time_s": round(total_ms / 1000.0, 3),
                "average_time_ms": round(total_ms / max(1, attempts), 3),
                "median_time_ms": round(statistics.median(rung_push_ms), 3),
                "p95_time_ms": round(percentile(rung_push_ms, 0.95), 3),
            }
        )

    detail_csv = reports / "off3_xapk_ladder_details.csv"
    summary_csv = reports / "off3_xapk_ladder_summary.csv"
    with detail_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
        w.writeheader()
        w.writerows(detail_rows)
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
        w.writeheader()
        w.writerows(summary_rows)

    run_xlsx = maybe_write_xlsx(run_dir, detail_rows, summary_rows, output_root)
    payload = {
        "run_id": run_id,
        "generated_utc": now_iso,
        "run_dir": str(run_dir),
        "source_file": str(src),
        "rungs": args.rungs,
        "device_dir": args.device_dir,
        "detail_csv": str(detail_csv),
        "summary_csv": str(summary_csv),
        "run_xlsx": str(run_xlsx) if run_xlsx else None,
    }
    write_json(run_dir / "summary.json", payload)

    print(f"[ByteBite] Detail CSV: {detail_csv}")
    print(f"[ByteBite] Summary CSV: {summary_csv}")
    if run_xlsx:
        print(f"[ByteBite] Run XLSX: {run_xlsx}")
    print(f"[ByteBite] Summary JSON: {run_dir / 'summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
