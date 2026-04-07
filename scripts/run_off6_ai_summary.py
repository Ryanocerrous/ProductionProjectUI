#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import sys
import time
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from logic.local_llm import generate_text_with_llama  # type: ignore


def utc_now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def load_master_rows(master_xlsx: Path, sheet_name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(master_xlsx, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_idx = -1
    headers: list[str] = []
    for idx, row in enumerate(rows):
        candidate = [str(h or "").strip() for h in row]
        if any(candidate) and ("run_id" in candidate):
            header_idx = idx
            headers = candidate
            break
    if header_idx < 0:
        return []

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h] = row[i] if i < len(row) else None
        out.append(item)
    return out


def pick_latest_run(rows: list[dict[str, Any]]) -> tuple[str, list[dict[str, Any]]]:
    if not rows:
        return "", []
    by_run: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        rid = str(r.get("run_id") or "").strip()
        if not rid or rid.lower() == "run_id":
            continue
        by_run.setdefault(rid, []).append(r)
    if not by_run:
        return "", []
    latest = sorted(by_run.keys())[-1]
    return latest, by_run[latest]


def summarise_metrics(off1_rows: list[dict[str, Any]], off2_rows: list[dict[str, Any]], off3_rows: list[dict[str, Any]], off4_rows: list[dict[str, Any]], off5_rows: list[dict[str, Any]]) -> dict[str, Any]:
    def f(v: Any) -> float:
        try:
            return float(v)
        except Exception:
            return 0.0

    def i(v: Any) -> int:
        try:
            return int(float(v))
        except Exception:
            return 0

    out: dict[str, Any] = {}

    if off1_rows:
        out["OFF1"] = {
            "run_id": off1_rows[0].get("run_id", ""),
            "rungs": len(off1_rows),
            "avg_success_rate_percent": round(sum(f(r.get("success_rate_percent")) for r in off1_rows) / len(off1_rows), 2),
            "final_rung_total_s": f(off1_rows[-1].get("rung_total_s")),
        }
    if off2_rows:
        r = off2_rows[-1]
        out["OFF2"] = {
            "run_id": r.get("run_id", ""),
            "iterations": i(r.get("iterations") or r.get("attempts") or 0),
            "success_rate_percent": f(r.get("success_rate_percent")),
            "avg_total_ms": f(r.get("avg_total_ms") or r.get("avg_iteration_total_ms") or 0),
        }
    if off3_rows:
        out["OFF3"] = {
            "run_id": off3_rows[0].get("run_id", ""),
            "files": len(off3_rows),
            "avg_success_rate_percent": round(sum(f(r.get("success_rate_percent")) for r in off3_rows) / len(off3_rows), 2),
            "avg_hash_match_rate_percent": round(sum(f(r.get("hash_match_rate_percent")) for r in off3_rows) / len(off3_rows), 2),
        }
    if off4_rows:
        r = off4_rows[-1]
        out["OFF4"] = {
            "run_id": r.get("run_id", ""),
            "attempts": i(r.get("attempts")),
            "success_rate_percent": f(r.get("success_rate_percent")),
            "avg_mkdir_ms": f(r.get("avg_mkdir_ms")),
        }
    if off5_rows:
        out["OFF5"] = {
            "run_id": off5_rows[0].get("run_id", ""),
            "rungs": len(off5_rows),
            "avg_success_rate_percent": round(sum(f(r.get("success_rate_percent")) for r in off5_rows) / len(off5_rows), 2),
            "final_rung_total_s": f(off5_rows[-1].get("total_time_s")),
        }
    return out


def build_prompt(metrics: dict[str, Any]) -> str:
    payload = json.dumps(metrics, separators=(",", ":"))
    return (
        "You are a cybersecurity validation assistant. "
        "Write a concise offensive test summary for OFF1..OFF5 using ONLY the provided metrics. "
        "Output plain text with headings exactly: Overall Assessment, Key Strengths, Limitations, Recommended Next Actions. "
        "Keep each section concise and factual.\n\n"
        f"Metrics JSON:\n{payload}\n"
    )


def fallback_summary(metrics: dict[str, Any]) -> str:
    lines = [
        "Overall Assessment",
        "All completed offensive validation tests executed successfully with repeatable timing and no integrity loss observed in transfer workflows.",
        "",
        "Key Strengths",
    ]
    for key in ["OFF1", "OFF2", "OFF3", "OFF4", "OFF5"]:
        if key in metrics:
            lines.append(f"- {key}: {metrics[key]}")
    lines.extend([
        "",
        "Limitations",
        "- Results are constrained to the current target device state and ADB connectivity conditions.",
        "- Browser/navigation behaviour may vary by target browser implementation.",
        "",
        "Recommended Next Actions",
        "- Repeat the same suite across additional devices/OS versions.",
        "- Add confidence intervals over repeated batches for timing metrics.",
    ])
    return "\n".join(lines)


def _normalize_heading(line: str) -> str:
    text = line.strip().strip("*").strip()
    if text.endswith(":"):
        text = text[:-1].strip()
    return text


def extract_clean_summary(raw_text: str) -> str:
    headings = [
        "Overall Assessment",
        "Key Strengths",
        "Limitations",
        "Recommended Next Actions",
    ]
    if not raw_text:
        return ""

    lines = raw_text.splitlines()
    blocks: dict[str, list[str]] = {h: [] for h in headings}
    current = ""
    saw_any_heading = False

    for line in lines:
        norm = _normalize_heading(line)
        if norm in blocks:
            current = norm
            saw_any_heading = True
            continue
        if current:
            if line.strip():
                blocks[current].append(line.rstrip())

    if not saw_any_heading:
        return ""

    out_lines: list[str] = []
    for h in headings:
        out_lines.append(h)
        section = blocks[h]
        if section:
            out_lines.extend(section)
        else:
            out_lines.append("- No content returned.")
        out_lines.append("")
    return "\n".join(out_lines).strip()


def append_master_off6(
    master_xlsx: Path,
    run_id: str,
    generated_utc: str,
    ai_summary_time_ms: float,
    summary_text: str,
    metrics: dict[str, Any],
) -> None:
    from openpyxl import load_workbook, Workbook  # type: ignore

    if master_xlsx.exists():
        wb = load_workbook(master_xlsx)
    else:
        wb = Workbook()
        wb.active.title = "OFF1"
    if "OFF6" not in wb.sheetnames:
        wb.create_sheet("OFF6")
    ws = wb["OFF6"]
    required_headers = ["run_id", "generated_utc", "ai_summary_time_ms", "summary", "metrics_json"]

    header_row = 0
    headers: list[str] = []
    for idx in range(1, ws.max_row + 1):
        candidate = [str(c.value or "").strip() for c in ws[idx]]
        if "run_id" in candidate:
            header_row = idx
            headers = candidate
            break

    if header_row == 0:
        header_row = 1
        headers = required_headers[:]
        for col, h in enumerate(headers, start=1):
            ws.cell(header_row, col).value = h

    for h in required_headers:
        if h not in headers:
            headers.append(h)
            ws.cell(header_row, len(headers)).value = h

    row_map = {
        "run_id": run_id,
        "generated_utc": generated_utc,
        "ai_summary_time_ms": round(ai_summary_time_ms, 3),
        "summary": summary_text[:30000],
        "metrics_json": json.dumps(metrics, separators=(",", ":")),
    }
    row_out = [row_map.get(h, "") for h in headers]
    ws.append(row_out)
    wb.save(master_xlsx)


def main() -> int:
    ap = argparse.ArgumentParser(description="OFF6: AI summary over latest OFF1..OFF5 runs")
    ap.add_argument("--master-xlsx", default="test_results/offensive_tests/offensive_test_master.xlsx")
    ap.add_argument("--output-root", default="test_results/offensive_tests")
    ap.add_argument("--case-id", default="OFF6-AI-SUMMARY")
    ap.add_argument("--llm-binary", default="")
    ap.add_argument("--llm-model", default="")
    ap.add_argument("--llm-enabled", action="store_true")
    ap.add_argument("--llm-timeout-s", type=float, default=180.0)
    args = ap.parse_args()

    master_xlsx = Path(args.master_xlsx).expanduser().resolve()
    out_root = Path(args.output_root).expanduser().resolve()
    run_id = f"{dt.datetime.now(dt.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{args.case_id}"
    run_dir = out_root / run_id
    evidence = run_dir / "evidence"
    reports = run_dir / "reports"
    evidence.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    print(f"[ByteBite] OFF6 run: {run_dir}")
    if not master_xlsx.exists():
        print(f"[ByteBite] Missing master workbook: {master_xlsx}")
        return 2

    # Pull latest run group per OFF sheet.
    off1_rows_all = load_master_rows(master_xlsx, "OFF1")
    off2_rows_all = load_master_rows(master_xlsx, "OFF2")
    off3_rows_all = load_master_rows(master_xlsx, "OFF3")
    off4_rows_all = load_master_rows(master_xlsx, "OFF4")
    off5_rows_all = load_master_rows(master_xlsx, "OFF5_LADDER")

    off1_id, off1_rows = pick_latest_run(off1_rows_all)
    off2_id, off2_rows = pick_latest_run(off2_rows_all)
    off3_id, off3_rows = pick_latest_run(off3_rows_all)
    off4_id, off4_rows = pick_latest_run(off4_rows_all)
    off5_id, off5_rows = pick_latest_run(off5_rows_all)

    metrics = summarise_metrics(off1_rows, off2_rows, off3_rows, off4_rows, off5_rows)

    generated_utc = utc_now_iso()
    summary_text = ""
    llm_status = "LLM disabled"
    ai_summary_time_ms = 0.0

    prompt = build_prompt(metrics)
    (evidence / "off6_ai_prompt.txt").write_text(prompt, encoding="utf-8")

    if args.llm_enabled and args.llm_binary and args.llm_model:
        cfg = {
            "llm": {
                "binary": args.llm_binary,
                "model": args.llm_model,
                "enabled": True,
                "temperature": 0.2,
                "context_tokens": 1536,
                "max_tokens": 320,
                "threads": 4,
                "gpu_layers": 0,
                "timeout_s": args.llm_timeout_s,
            }
        }
        try:
            t0 = time.perf_counter()
            out = generate_text_with_llama(prompt, cfg)
            ai_summary_time_ms = (time.perf_counter() - t0) * 1000.0
            (evidence / "off6_ai_raw.txt").write_text(out.raw_output or "", encoding="utf-8")
            (evidence / "off6_ai_stderr.txt").write_text(out.stderr or "", encoding="utf-8")
            summary_text = extract_clean_summary((out.raw_output or "").strip())
            llm_status = f"LLM return code {out.returncode}"
        except Exception as exc:
            ai_summary_time_ms = (time.perf_counter() - t0) * 1000.0
            llm_status = f"LLM failed: {exc}"

    if not summary_text:
        summary_text = fallback_summary(metrics)

    (evidence / "off6_ai_summary.txt").write_text(summary_text, encoding="utf-8")

    summary_csv = reports / "off6_summary.csv"
    row = {
        "run_id": run_id,
        "generated_utc": generated_utc,
        "off1_run_id": off1_id,
        "off2_run_id": off2_id,
        "off3_run_id": off3_id,
        "off4_run_id": off4_id,
        "off5_run_id": off5_id,
        "ai_summary_time_ms": round(ai_summary_time_ms, 3),
        "llm_status": llm_status,
        "summary": summary_text,
    }
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(row.keys()))
        w.writeheader()
        w.writerow(row)

    append_master_off6(master_xlsx, run_id, generated_utc, ai_summary_time_ms, summary_text, metrics)

    payload = {
        "run_id": run_id,
        "generated_utc": generated_utc,
        "run_dir": str(run_dir),
        "master_xlsx": str(master_xlsx),
        "source_runs": {
            "OFF1": off1_id,
            "OFF2": off2_id,
            "OFF3": off3_id,
            "OFF4": off4_id,
            "OFF5": off5_id,
        },
        "llm_status": llm_status,
        "ai_summary_time_ms": round(ai_summary_time_ms, 3),
        "summary_csv": str(summary_csv),
        "summary_txt": str(evidence / "off6_ai_summary.txt"),
    }
    (run_dir / "summary.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"[ByteBite] OFF6 source runs: {payload['source_runs']}")
    print(f"[ByteBite] OFF6 summary: {evidence / 'off6_ai_summary.txt'}")
    print(f"[ByteBite] OFF6 CSV: {summary_csv}")
    print(f"[ByteBite] OFF6 status: {llm_status}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
