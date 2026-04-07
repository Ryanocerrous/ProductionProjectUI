from __future__ import annotations

import argparse
import re
import shlex
import subprocess
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def safe_name(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("_") or "item"


@dataclass(slots=True)
class CmdResult:
    args: list[str]
    returncode: int
    stdout: str
    stderr: str

    @property
    def ok(self) -> bool:
        return self.returncode == 0


def run_cmd(args: list[str], timeout_s: float = 600.0) -> CmdResult:
    try:
        proc = subprocess.run(args, capture_output=True, text=True, check=False, timeout=timeout_s)
        return CmdResult(args=args, returncode=proc.returncode, stdout=proc.stdout or "", stderr=proc.stderr or "")
    except subprocess.TimeoutExpired as exc:
        return CmdResult(
            args=args,
            returncode=124,
            stdout=(exc.stdout or "") if isinstance(exc.stdout, str) else "",
            stderr=(exc.stderr or "Command timed out") if isinstance(exc.stderr, str) else "Command timed out",
        )
    except FileNotFoundError:
        return CmdResult(args=args, returncode=127, stdout="", stderr=f"{args[0]}: command not found")


def adb_cmd(adb_bin: str, serial: str, args: list[str], timeout_s: float = 600.0) -> CmdResult:
    cmd = [adb_bin]
    if serial:
        cmd.extend(["-s", serial])
    cmd.extend(args)
    return run_cmd(cmd, timeout_s=timeout_s)


def list_remote_files(adb_bin: str, serial: str, remote_dir: str, pattern: str) -> list[str]:
    base = shlex.quote(remote_dir.rstrip("/"))
    # Keep wildcard unquoted so Android shell expands it.
    shell_cmd = f"ls -1 {base}/{pattern} 2>/dev/null"
    res = adb_cmd(adb_bin, serial, ["shell", shell_cmd], timeout_s=30.0)
    if not res.ok and not res.stdout.strip():
        return []
    files = []
    for line in res.stdout.splitlines():
        value = line.strip()
        if not value:
            continue
        if value.startswith("ls: "):
            continue
        files.append(value)
    return sorted(set(files))


def pull_file(adb_bin: str, serial: str, remote_file: str, local_file: Path) -> tuple[CmdResult, float]:
    local_file.parent.mkdir(parents=True, exist_ok=True)
    if local_file.exists():
        local_file.unlink()
    t0 = time.perf_counter()
    res = adb_cmd(adb_bin, serial, ["pull", remote_file, str(local_file)], timeout_s=3600.0)
    elapsed = round(time.perf_counter() - t0, 3)
    return res, elapsed


def load_or_create_workbook(path: Path) -> Any:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


def ensure_headers(ws: Any, headers: list[str]) -> None:
    if ws.max_row < 1:
        ws.append(headers)
        return
    existing = [str(ws.cell(1, i + 1).value or "").strip() for i in range(len(headers))]
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)


def update_master_xlsx(path: Path, rows: list[dict[str, Any]], session_id: str) -> None:
    wb = load_or_create_workbook(path)

    raw_headers = [
        "utc",
        "session_id",
        "remote_file",
        "file_name",
        "file_size_mb",
        "iteration",
        "elapsed_seconds",
        "throughput_mb_per_s",
        "status",
        "stderr",
    ]
    raw_ws = wb["RawResults"] if "RawResults" in wb.sheetnames else wb.create_sheet("RawResults")
    ensure_headers(raw_ws, raw_headers)

    for row in rows:
        raw_ws.append(
            [
                row["utc"],
                row["session_id"],
                row["remote_file"],
                row["file_name"],
                row["file_size_mb"],
                row["iteration"],
                row["elapsed_seconds"],
                row["throughput_mb_per_s"],
                row["status"],
                row["stderr"],
            ]
        )

    summary_headers = [
        "session_id",
        "remote_file",
        "file_name",
        "runs",
        "passes",
        "success_rate_pct",
        "avg_time_seconds",
        "min_time_seconds",
        "max_time_seconds",
        "avg_throughput_mb_per_s",
    ]
    summary_ws = wb["SessionSummary"] if "SessionSummary" in wb.sheetnames else wb.create_sheet("SessionSummary")
    ensure_headers(summary_ws, summary_headers)

    by_file: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_file.setdefault(row["remote_file"], []).append(row)

    for remote_file, vals in sorted(by_file.items()):
        runs = len(vals)
        passes = sum(1 for x in vals if x["status"] == "PASS")
        times = [float(x["elapsed_seconds"]) for x in vals]
        speeds = [float(x["throughput_mb_per_s"]) for x in vals if float(x["throughput_mb_per_s"]) > 0.0]
        summary_ws.append(
            [
                session_id,
                remote_file,
                vals[0]["file_name"],
                runs,
                passes,
                round((passes / runs) * 100.0, 2) if runs else 0.0,
                round(sum(times) / len(times), 3) if times else 0.0,
                round(min(times), 3) if times else 0.0,
                round(max(times), 3) if times else 0.0,
                round(sum(speeds) / len(speeds), 3) if speeds else 0.0,
            ]
        )

    overall_headers = [
        "updated_utc",
        "total_rows",
        "total_sessions",
        "pass_rows",
        "fail_rows",
        "overall_success_rate_pct",
        "avg_time_seconds",
        "avg_throughput_mb_per_s",
    ]
    overall_ws = wb["Overall"] if "Overall" in wb.sheetnames else wb.create_sheet("Overall")
    ensure_headers(overall_ws, overall_headers)

    all_rows: list[tuple[Any, ...]] = []
    for i in range(2, raw_ws.max_row + 1):
        all_rows.append(
            (
                raw_ws.cell(i, 1).value,
                raw_ws.cell(i, 2).value,
                raw_ws.cell(i, 7).value,
                raw_ws.cell(i, 8).value,
                raw_ws.cell(i, 9).value,
            )
        )
    total_rows = len(all_rows)
    sessions = len({str(x[1]) for x in all_rows if x[1]})
    pass_rows = sum(1 for x in all_rows if str(x[4]) == "PASS")
    fail_rows = total_rows - pass_rows
    times = [float(x[2]) for x in all_rows if x[2] is not None]
    speeds = [float(x[3]) for x in all_rows if x[3] is not None]
    overall_ws.append(
        [
            utc_now_iso(),
            total_rows,
            sessions,
            pass_rows,
            fail_rows,
            round((pass_rows / total_rows) * 100.0, 2) if total_rows else 0.0,
            round(sum(times) / len(times), 3) if times else 0.0,
            round(sum(speeds) / len(speeds), 3) if speeds else 0.0,
        ]
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Run repeated ADB extraction timing tests for ByteBite test files and update a master XLSX."
    )
    parser.add_argument("--adb-bin", default="adb")
    parser.add_argument("--device-serial", default="")
    parser.add_argument("--remote-dir", default="/sdcard/Download/ByteBiteTest")
    parser.add_argument("--pattern", default="keyword_story_*MB.txt")
    parser.add_argument("--runs-per-file", type=int, default=5)
    parser.add_argument("--output-dir", default="test_results")
    parser.add_argument("--master-xlsx", default="")
    args = parser.parse_args()

    if args.runs_per_file < 1:
        print("[ByteBite] runs-per-file must be >= 1")
        return 2

    output_dir = Path(args.output_dir).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    session_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    session_dir = output_dir / f"extract_benchmark_{session_id}"
    pulls_dir = session_dir / "pulls_tmp"
    logs_dir = session_dir / "logs"
    pulls_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)

    master_xlsx = Path(args.master_xlsx).expanduser() if args.master_xlsx else (output_dir / "extraction_master.xlsx")

    devices = adb_cmd(args.adb_bin, args.device_serial.strip(), ["devices"], timeout_s=30.0)
    (logs_dir / "adb_devices.txt").write_text((devices.stdout or "") + "\n" + (devices.stderr or ""), encoding="utf-8")
    if not devices.ok:
        print("[ByteBite] adb devices failed")
        print(devices.stderr.strip())
        return 3

    remote_files = list_remote_files(args.adb_bin, args.device_serial.strip(), args.remote_dir.strip(), args.pattern.strip())
    if not remote_files:
        print("[ByteBite] No matching remote files found.")
        print(f"[ByteBite] Path checked: {args.remote_dir.rstrip('/')}/{args.pattern}")
        return 4

    print(f"[ByteBite] Files found: {len(remote_files)}")
    print(f"[ByteBite] Runs per file: {args.runs_per_file}")
    print(f"[ByteBite] Session dir: {session_dir}")

    rows: list[dict[str, Any]] = []
    for remote_file in remote_files:
        file_name = Path(remote_file).name
        mb_match = re.search(r"(\d+)MB", file_name, re.IGNORECASE)
        file_size_mb = int(mb_match.group(1)) if mb_match else 0
        file_slug = safe_name(file_name)
        print(f"[ByteBite] Testing {file_name} ...")

        for iteration in range(1, args.runs_per_file + 1):
            local_file = pulls_dir / f"{file_slug}.run{iteration}.tmp"
            res, elapsed = pull_file(args.adb_bin, args.device_serial.strip(), remote_file, local_file)
            pulled_bytes = local_file.stat().st_size if local_file.exists() else 0
            throughput = round((pulled_bytes / (1024 * 1024)) / elapsed, 3) if elapsed > 0 and pulled_bytes > 0 else 0.0
            status = "PASS" if (res.ok and pulled_bytes > 0) else "FAIL"

            (logs_dir / f"{file_slug}.run{iteration}.stdout.txt").write_text(res.stdout or "", encoding="utf-8")
            (logs_dir / f"{file_slug}.run{iteration}.stderr.txt").write_text(res.stderr or "", encoding="utf-8")

            rows.append(
                {
                    "utc": utc_now_iso(),
                    "session_id": session_id,
                    "remote_file": remote_file,
                    "file_name": file_name,
                    "file_size_mb": file_size_mb,
                    "iteration": iteration,
                    "elapsed_seconds": elapsed,
                    "throughput_mb_per_s": throughput,
                    "status": status,
                    "stderr": (res.stderr or "").strip()[:500],
                }
            )

            print(
                f"[ByteBite] {file_name} run {iteration}/{args.runs_per_file}: "
                f"{status} in {elapsed:.3f}s ({throughput:.3f} MB/s)"
            )

            if local_file.exists():
                try:
                    local_file.unlink()
                except Exception:
                    pass

    raw_csv = session_dir / "raw_results.csv"
    raw_headers = [
        "utc",
        "session_id",
        "remote_file",
        "file_name",
        "file_size_mb",
        "iteration",
        "elapsed_seconds",
        "throughput_mb_per_s",
        "status",
        "stderr",
    ]
    lines = [",".join(raw_headers)]
    for row in rows:
        values = [str(row[h]).replace("\n", " ").replace(",", ";") for h in raw_headers]
        lines.append(",".join(values))
    raw_csv.write_text("\n".join(lines) + "\n", encoding="utf-8")

    if Workbook is None or load_workbook is None:
        print("[ByteBite] openpyxl not installed; skipped XLSX update.")
    else:
        update_master_xlsx(master_xlsx, rows, session_id)
        print(f"[ByteBite] Master XLSX updated: {master_xlsx}")

    print(f"[ByteBite] Session complete: {session_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
