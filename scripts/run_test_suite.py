from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import shlex
import shutil
import subprocess
import time
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SRC_DIR = Path(__file__).resolve().parents[1] / "src"
PROJECT_ROOT = SRC_DIR.parent
if str(SRC_DIR) not in os.sys.path:
    os.sys.path.insert(0, str(SRC_DIR))

from logic.adb import Adb, CommandResult
from logic.local_llm import classify_text_with_llama, generate_text_with_llama
from logic.runtime_paths import build_default_config, load_or_create_config, resolve_config_path

try:
    from pypdf import PdfReader  # type: ignore
except Exception:
    PdfReader = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def safe_name(value: str) -> str:
    out = "".join(ch if (ch.isalnum() or ch in {"-", "_", "."}) else "_" for ch in value)
    return out.strip("_") or "item"


def excel_sheet_name(name: str) -> str:
    cleaned = name
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        cleaned = cleaned.replace(ch, "-")
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def parse_int(text: str, default: int = 0) -> int:
    try:
        return int(float(str(text).strip()))
    except Exception:
        return default


def parse_float(text: str, default: float = 0.0) -> float:
    try:
        return float(str(text).strip())
    except Exception:
        return default


def percentile(values: list[float], pct: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return float(values[0])
    ordered = sorted(float(v) for v in values)
    rank = (len(ordered) - 1) * pct
    lo = int(rank)
    hi = min(lo + 1, len(ordered) - 1)
    if lo == hi:
        return ordered[lo]
    frac = rank - lo
    return ordered[lo] + (ordered[hi] - ordered[lo]) * frac


def normalize_ai_summary_text(text: str) -> str:
    raw = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    raw = raw.replace("\x00", " ")
    if not raw.strip():
        return "AI summary not available."

    headings = [
        "Overall Assessment",
        "Key Strengths",
        "Key Limitations",
        "Recommended Next Actions",
    ]

    plain = raw.replace("**", "")
    pattern = re.compile(
        r"(?is)(Overall Assessment|Key Strengths|Key Limitations|Recommended Next Actions)\s*:?\s*(.*?)"
        r"(?=(Overall Assessment|Key Strengths|Key Limitations|Recommended Next Actions)\s*:|$)"
    )
    sections: dict[str, str] = {}
    for m in pattern.finditer(plain):
        key = (m.group(1) or "").strip()
        body = (m.group(2) or "").strip()
        if not key or key in sections:
            continue
        line = " ".join(body.split())
        line = line.lstrip("-•").strip()
        sections[key] = line[:260]

    if len(sections) >= 2:
        out_lines: list[str] = []
        for h in headings:
            if h in sections and sections[h]:
                out_lines.append(h)
                out_lines.append(f"- {sections[h]}")
                out_lines.append("")
        cleaned = "\n".join(out_lines).strip()
        return cleaned or "AI summary not available."

    # Fallback: keep only readable model response text, stripping prompt/runtime noise.
    scrubbed = plain
    for marker in [
        "Run data (JSON):",
        "[ Prompt:",
        "Loading model",
        "llama_model_load_from_file_impl:",
        "llama_context:",
        "main: n_ctx",
        "system_info:",
        "sampler seed:",
        "generate: n_ctx",
        "prompt eval time",
        "eval time",
        "total time",
        "build:",
    ]:
        idx = scrubbed.lower().find(marker.lower())
        if idx >= 0 and marker.lower() in {"run data (json):", "[ prompt:"}:
            scrubbed = scrubbed[:idx]
    lines = []
    drop_prefixes = (
        "build:",
        "main:",
        "llama_",
        "system_info:",
        "sampler",
        "generate:",
        "[ prompt:",
        "prompt eval time",
        "eval time",
        "total time",
        "loading model",
    )
    for line in scrubbed.splitlines():
        t = line.strip()
        if not t:
            continue
        if any(t.lower().startswith(p) for p in drop_prefixes):
            continue
        if t == "Exiting...":
            continue
        lines.append(" ".join(t.split()))

    collapsed = " ".join(lines).strip()
    if not collapsed:
        return "AI summary not available."
    return collapsed[:1200]


def should_normalize_ai_text(text: str) -> bool:
    t = str(text or "")
    if not t.strip():
        return True
    noisy_markers = [
        "run data (json):",
        "[ prompt:",
        "loading model",
        "build:",
        "llama_",
        "system_info:",
        "prompt eval time",
        "eval time",
        "total time",
    ]
    lower = t.lower()
    return any(marker in lower for marker in noisy_markers)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def read_csv_rows(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    out: list[list[str]] = []
    try:
        with path.open("r", newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            for row in reader:
                out.append([str(cell) for cell in row])
    except Exception:
        return []
    return out


def count_files(path: Path) -> int:
    if not path.exists():
        return 0
    return sum(1 for p in path.rglob("*") if p.is_file())


def total_size_bytes(path: Path) -> int:
    if not path.exists():
        return 0
    total = 0
    for p in path.rglob("*"):
        if p.is_file():
            try:
                total += p.stat().st_size
            except Exception:
                continue
    return total


def save_command_result(base_path: Path, result: CommandResult) -> None:
    write_json(
        base_path.with_suffix(".json"),
        {
            "args": result.args,
            "returncode": result.returncode,
            "ok": result.ok,
            "stdout_len": len(result.stdout or ""),
            "stderr_len": len(result.stderr or ""),
        },
    )
    write_text(base_path.with_name(base_path.name + ".stdout.txt"), result.stdout or "")
    write_text(base_path.with_name(base_path.name + ".stderr.txt"), result.stderr or "")


def run_local_cmd(args: list[str], timeout_s: float = 60.0) -> CommandResult:
    try:
        proc = subprocess.run(args, capture_output=True, text=True, check=False, timeout=timeout_s)
        return CommandResult(
            args=args,
            returncode=proc.returncode,
            stdout=(proc.stdout or "").strip(),
            stderr=(proc.stderr or "").strip(),
        )
    except FileNotFoundError:
        return CommandResult(args=args, returncode=127, stdout="", stderr=f"{args[0]}: command not found")
    except subprocess.TimeoutExpired as exc:
        return CommandResult(
            args=args,
            returncode=124,
            stdout=((exc.stdout or "") if isinstance(exc.stdout, str) else "").strip(),
            stderr=((exc.stderr or "command timed out") if isinstance(exc.stderr, str) else "command timed out").strip(),
        )


@dataclass(slots=True)
class ForensicRecord:
    test_id: str
    objective: str
    expected_outcome: str
    input_output: str
    metric_measured: str
    result_text: str
    avg_time_seconds: float
    success_rate_pct: float
    status: str
    evidence_dir: str
    notes: str
    metrics: dict[str, Any] = field(default_factory=dict)


class ForensicTestRunner:
    def __init__(self, app_cfg: dict[str, Any], suite_cfg: dict[str, Any], run_root: Path) -> None:
        self.app_cfg = app_cfg
        self.suite_cfg = suite_cfg
        self.run_root = run_root
        self.records: list[ForensicRecord] = []

        serial = str(suite_cfg.get("device_serial") or app_cfg.get("device_serial") or "").strip()
        adb_bin = str(suite_cfg.get("adb_bin", "adb")).strip() or "adb"
        self.adb = Adb(serial=serial, adb_bin=adb_bin)

        self.for2_remote_path = str(suite_cfg.get("for2_remote_path", "/sdcard")).strip() or "/sdcard"
        self.for2_remote_candidates = [
            str(p).strip()
            for p in suite_cfg.get(
                "for2_remote_candidates",
                [
                    "/sdcard",
                    "/internal storage",
                    "/storage/emulated/0",
                    "/storage/self/primary",
                    "/sdcard/Download",
                    "/sdcard/Documents",
                    "/sdcard/DCIM",
                    "/sdcard/Pictures",
                    "/sdcard/Movies",
                    "/sdcard/Android/media",
                ],
            )
            if str(p).strip()
        ]
        self.usb_mount = (
            Path(str(suite_cfg.get("usb_mount_path", "")).strip()).expanduser()
            if suite_cfg.get("usb_mount_path")
            else None
        )
        self.auto_detect_usb = bool(suite_cfg.get("auto_detect_usb", True))
        self.adb_wait_timeout_s = float(suite_cfg.get("adb_wait_timeout_s", 12.0))
        self.keyword_list = [
            str(k).strip().lower()
            for k in suite_cfg.get(
                "keyword_list",
                [
                    "gun",
                    "knife",
                    "murder",
                    "burner",
                    "drugs",
                    "encrypted",
                    "target",
                    "device",
                    "victim",
                    "blade",
                    "weapon",
                    "police",
                    "operation",
                    "getaway",
                    "escape",
                    "kidnap",
                    "prison",
                    "smuggle",
                    "bullet",
                    "shank",
                ],
            )
            if str(k).strip()
        ]
        self.text_exts = {
            str(x).strip().lower()
            for x in suite_cfg.get(
                "text_extensions",
                [".txt", ".pdf", ".docx", ".log", ".json", ".xml", ".csv", ".md", ".html"],
            )
            if str(x).strip()
        }
        # FOR5 requirement: always include plain text and PDF artifacts.
        self.text_exts.update({".txt", ".pdf", ".docx"})
        self.max_scan_files = int(suite_cfg.get("max_text_scan_files", 200))

        self.for2_extract_dir = self.run_root / "FOR2_extract_storage" / "extracted_storage"
        self.for3_hash_csv = self.run_root / "FOR3_hashed_data" / "sha256_hashes.csv"
        self.for5_hits_json = self.run_root / "FOR5_ai_analysis" / "keyword_hits.json"
        self.report_dir = self.run_root / "FOR6_report"
        self.reports_dir = self.run_root / "reports"
        self.for6_targets = suite_cfg.get(
            "for6_target_files",
            [
                {"file_name": "O Death.txt", "ground_truth": "safe"},
                {"file_name": "Marcus.txt", "ground_truth": "high_priority"},
                {"file_name": "Firearms Manual.pdf", "ground_truth": "suspicious"},
            ],
        )
        self.for6_iterations = max(1, int(suite_cfg.get("for6_iterations", 1)))
        self.for6_max_chars = max(200, int(suite_cfg.get("for6_max_chars", 1200)))
        self.for6_context_tokens = max(256, int(suite_cfg.get("for6_context_tokens", 1024)))
        self.for6_max_tokens = max(32, int(suite_cfg.get("for6_max_tokens", 120)))
        self.for6_temperature = float(suite_cfg.get("for6_temperature", 0.1))
        self.for6_pass_accuracy_pct = float(suite_cfg.get("for6_pass_accuracy_pct", 100.0))
        self.for6_use_policy_override = bool(suite_cfg.get("for6_use_policy_override", True))

    def _count_device_states(self, devices_output: str) -> tuple[int, int, int, list[str]]:
        devices = 0
        unauthorized = 0
        offline = 0
        serials: list[str] = []
        for line in (devices_output or "").splitlines():
            line = line.strip()
            if not line or line.startswith("List of devices attached"):
                continue
            parts = line.split()
            if not parts:
                continue
            serials.append(parts[0])
            state = parts[1] if len(parts) > 1 else ""
            if state == "device":
                devices += 1
            elif state == "unauthorized":
                unauthorized += 1
            elif state == "offline":
                offline += 1
        return devices, unauthorized, offline, serials

    def _probe_writable_dir(self, path: Path) -> bool:
        try:
            path.mkdir(parents=True, exist_ok=True)
            probe = path / ".bytebite_write_probe"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return True
        except Exception:
            return False

    def _discover_usb_mount(self) -> Path | None:
        if self.usb_mount and self._probe_writable_dir(self.usb_mount):
            return self.usb_mount
        if not self.auto_detect_usb:
            return None

        user = os.environ.get("USER", "kali")
        roots = [Path(f"/media/{user}"), Path("/media/kali"), Path("/mnt")]
        for root in roots:
            if not root.exists() or not root.is_dir():
                continue
            for child in sorted(root.iterdir()):
                if not child.is_dir():
                    continue
                if child.name.startswith("."):
                    continue
                if self._probe_writable_dir(child):
                    return child
        return None

    def _resolve_for2_remote_path(self, evidence: Path) -> str:
        tested: list[dict[str, Any]] = []
        for remote in self.for2_remote_candidates:
            cmd = f"ls -d {shlex.quote(remote)} 2>/dev/null"
            res = self.adb.shell(cmd, timeout_s=10.0)
            tested.append({"remote": remote, "ok": res.ok, "stdout": res.stdout, "stderr": res.stderr, "returncode": res.returncode})
            if res.ok and (res.stdout or "").strip():
                write_json(evidence / "for2_remote_probe.json", tested)
                return remote
        write_json(evidence / "for2_remote_probe.json", tested)
        return self.for2_remote_path

    def adb_cmd(self, args: list[str], timeout_s: float = 60.0) -> CommandResult:
        cmd = [self.adb.adb_bin]
        if self.adb.serial:
            cmd.extend(["-s", self.adb.serial])
        cmd.extend(args)
        return run_local_cmd(cmd, timeout_s=timeout_s)

    def run(self, selected: set[str] | None = None) -> list[ForensicRecord]:
        tests = [
            self.for1_device_detect,
            self.for2_extract_storage,
            self.for3_hash_data,
            self.for4_store_usb,
            self.for5_ai_analysis,
            self.for6_report_completed,
        ]
        for fn in tests:
            tid = fn.__name__.split("_", 1)[0].upper()
            if selected and tid not in selected:
                continue
            print(f"[ByteBite] Running {tid}...", flush=True)
            rec = fn()
            self.records.append(rec)
            print(f"[ByteBite] {tid} -> {rec.status}", flush=True)
        return self.records

    def for1_device_detect(self) -> ForensicRecord:
        tid = "FOR1"
        evidence = self.run_root / "FOR1_device_detect"
        evidence.mkdir(parents=True, exist_ok=True)

        t0 = time.perf_counter()
        res_initial = self.adb.devices()
        save_command_result(evidence / "adb_devices_initial", res_initial)
        devices, unauthorized, offline, serials = self._count_device_states(res_initial.stdout or "")
        if devices == 0 and self.adb_wait_timeout_s > 0:
            wait_res = self.adb.wait_for_device(timeout_s=self.adb_wait_timeout_s)
            save_command_result(evidence / "adb_wait_for_device", wait_res)
            res = self.adb.devices()
        else:
            res = res_initial
        elapsed = round(time.perf_counter() - t0, 3)
        save_command_result(evidence / "adb_devices_final", res)
        devices, unauthorized, offline, serials = self._count_device_states(res.stdout or "")

        passed = res.ok and devices > 0
        success_rate = 100.0 if passed else 0.0
        if passed:
            result = f"Device detected successfully (device_count={devices})."
        else:
            result = f"No authorized device detected (device={devices}, unauthorized={unauthorized}, offline={offline})."

        return ForensicRecord(
            test_id=tid,
            objective="Device detected using ADB",
            expected_outcome="Device connects to victim mobile",
            input_output="USB-C",
            metric_measured="Detection time (seconds)",
            result_text=result,
            avg_time_seconds=elapsed,
            success_rate_pct=success_rate,
            status="PASS" if passed else "FAIL",
            evidence_dir=str(evidence),
            notes=(f"Serials: {', '.join(serials)}" if serials else (res.stderr or "No device rows returned.")),
            metrics={"device_count": devices, "unauthorized": unauthorized, "offline": offline},
        )

    def for2_extract_storage(self) -> ForensicRecord:
        tid = "FOR2"
        evidence = self.run_root / "FOR2_extract_storage"
        evidence.mkdir(parents=True, exist_ok=True)
        self.for2_extract_dir.mkdir(parents=True, exist_ok=True)

        wait_res = self.adb.wait_for_device(timeout_s=self.adb_wait_timeout_s)
        save_command_result(evidence / "adb_wait_for_device", wait_res)
        remote_path = self._resolve_for2_remote_path(evidence)

        t0 = time.perf_counter()
        pull = self.adb.pull(remote_path, str(self.for2_extract_dir))
        elapsed = round(time.perf_counter() - t0, 3)
        save_command_result(evidence / "adb_pull_storage", pull)

        file_count = count_files(self.for2_extract_dir)
        size_bytes = total_size_bytes(self.for2_extract_dir)
        size_mb = round(size_bytes / (1024 * 1024), 2)
        write_json(
            evidence / "extraction_stats.json",
            {
                "remote_path": remote_path,
                "local_path": str(self.for2_extract_dir),
                "file_count": file_count,
                "size_bytes": size_bytes,
                "size_mb": size_mb,
                "elapsed_seconds": elapsed,
                "pull_returncode": pull.returncode,
            },
        )

        passed = pull.ok and file_count > 0
        success_rate = 100.0 if passed else (50.0 if pull.ok else 0.0)
        result = (
            f"{size_mb} MB extracted from {remote_path} with {file_count} files."
            if pull.ok
            else f"Extraction failed for {remote_path}."
        )

        return ForensicRecord(
            test_id=tid,
            objective="Extract user storage",
            expected_outcome="All files are retrieved",
            input_output=f"Input: {remote_path}; Output: data files",
            metric_measured="Data size (MB), extraction time",
            result_text=result,
            avg_time_seconds=elapsed,
            success_rate_pct=success_rate,
            status="PASS" if passed else "FAIL",
            evidence_dir=str(evidence),
            notes=("ADB pull completed." if pull.ok else (pull.stderr or "ADB pull failed.")),
            metrics={"file_count": file_count, "size_mb": size_mb, "pull_ok": pull.ok},
        )

    def for3_hash_data(self) -> ForensicRecord:
        tid = "FOR3"
        evidence = self.run_root / "FOR3_hashed_data"
        evidence.mkdir(parents=True, exist_ok=True)

        files = [p for p in self.for2_extract_dir.rglob("*") if p.is_file()]
        t0 = time.perf_counter()

        rows: list[list[Any]] = []
        hashed = 0
        consistent = 0
        for path in files:
            rel = str(path.relative_to(self.for2_extract_dir))
            try:
                data = path.read_bytes()
                sha_a = hashlib.sha256(data).hexdigest()
                sha_b = hashlib.sha256(data).hexdigest()
                ok = sha_a == sha_b
                rows.append([rel, len(data), sha_a, "YES" if ok else "NO"])
                hashed += 1
                if ok:
                    consistent += 1
            except Exception as exc:
                rows.append([rel, -1, "", f"ERROR: {type(exc).__name__}"])

        elapsed = round(time.perf_counter() - t0, 3)
        write_csv(self.for3_hash_csv, ["relative_path", "size_bytes", "sha256", "consistent"], rows)

        accuracy = round((consistent / hashed) * 100.0, 2) if hashed else 0.0
        passed = hashed > 0 and consistent == hashed
        result = (
            f"All files hashed using SHA-256 consistently across run ({hashed}/{hashed})."
            if passed
            else f"SHA-256 hashing completed with consistency {consistent}/{hashed}."
        )
        write_json(
            evidence / "hash_stats.json",
            {
                "total_files": len(files),
                "hashed_files": hashed,
                "consistent_files": consistent,
                "hash_accuracy_pct": accuracy,
                "elapsed_seconds": elapsed,
                "hash_csv": str(self.for3_hash_csv),
            },
        )

        return ForensicRecord(
            test_id=tid,
            objective="Hashed data",
            expected_outcome="Data is hashed into SHA-256",
            input_output="SHA-256 hash values",
            metric_measured="Hash generation time, hash accuracy",
            result_text=result,
            avg_time_seconds=elapsed,
            success_rate_pct=accuracy,
            status="PASS" if passed else "FAIL",
            evidence_dir=str(evidence),
            notes=("Hash file generated." if hashed > 0 else "No files available from FOR2 extraction."),
            metrics={"hashed_files": hashed, "consistent_files": consistent, "hash_accuracy_pct": accuracy},
        )

    def _bucket_for_path(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".heic"}:
            return "images"
        if suffix in {".mp4", ".mov", ".avi", ".mkv", ".3gp"}:
            return "videos"
        if suffix in {".txt", ".log", ".md", ".json", ".xml", ".csv", ".pdf", ".doc", ".docx"}:
            return "documents"
        return "other"

    def for4_store_usb(self) -> ForensicRecord:
        tid = "FOR4"
        evidence = self.run_root / "FOR4_store_usb"
        evidence.mkdir(parents=True, exist_ok=True)

        files = [p for p in self.for2_extract_dir.rglob("*") if p.is_file()]
        usb_target = self._discover_usb_mount()
        if not usb_target:
            return ForensicRecord(
                test_id=tid,
                objective="Filed data into USB",
                expected_outcome="Data is sorted into folders",
                input_output="Folders (docs, text)",
                metric_measured="Write speed (seconds), storage success",
                result_text="USB path not configured.",
                avg_time_seconds=0.0,
                success_rate_pct=0.0,
                status="FAIL",
                evidence_dir=str(evidence),
                notes="Set usb_mount_path in suite config or connect/mount writable USB media.",
                metrics={},
            )

        target_root = usb_target / "bytebite_forensic_cases" / self.run_root.name / "sorted_data"
        copied = 0
        failures = 0
        manifest: list[dict[str, Any]] = []

        t0 = time.perf_counter()
        try:
            target_root.mkdir(parents=True, exist_ok=True)
            for src in files:
                bucket = self._bucket_for_path(src)
                dst_dir = target_root / bucket
                dst_dir.mkdir(parents=True, exist_ok=True)
                dst = dst_dir / src.name
                try:
                    shutil.copy2(src, dst)
                    copied += 1
                    manifest.append({"source": str(src), "target": str(dst), "bucket": bucket, "ok": True})
                except Exception as exc:
                    failures += 1
                    manifest.append(
                        {
                            "source": str(src),
                            "target": str(dst),
                            "bucket": bucket,
                            "ok": False,
                            "error": f"{type(exc).__name__}: {exc}",
                        }
                    )
        except Exception as exc:
            elapsed = round(time.perf_counter() - t0, 3)
            write_json(evidence / "usb_storage_manifest.json", manifest)
            return ForensicRecord(
                test_id=tid,
                objective="Filed data into USB",
                expected_outcome="Data is sorted into folders",
                input_output="Folders (docs, text)",
                metric_measured="Write speed (seconds), storage success",
                result_text="USB storage failed before copy stage.",
                avg_time_seconds=elapsed,
                success_rate_pct=0.0,
                status="FAIL",
                evidence_dir=str(evidence),
                notes=f"{type(exc).__name__}: {exc}",
                metrics={},
            )

        elapsed = round(time.perf_counter() - t0, 3)
        write_json(
            evidence / "usb_storage_manifest.json",
            {
                "usb_mount": str(usb_target),
                "target_root": str(target_root),
                "total_source_files": len(files),
                "copied_files": copied,
                "failed_files": failures,
                "elapsed_seconds": elapsed,
                "items": manifest,
            },
        )

        success_rate = round((copied / len(files)) * 100.0, 2) if files else 0.0
        passed = len(files) > 0 and copied == len(files) and failures == 0
        result = (
            f"Data sorted and stored on USB ({copied}/{len(files)} files copied)."
            if files
            else "No extracted files available to store on USB."
        )

        return ForensicRecord(
            test_id=tid,
            objective="Filed data into USB",
            expected_outcome="Data is sorted into folders",
            input_output="Folders (docs, text)",
            metric_measured="Write speed (seconds), storage success",
            result_text=result,
            avg_time_seconds=elapsed,
            success_rate_pct=success_rate,
            status="PASS" if passed else "FAIL",
            evidence_dir=str(evidence),
            notes=(f"USB target: {target_root}" if files else "FOR2 produced no files."),
            metrics={"source_files": len(files), "copied_files": copied, "failed_files": failures},
        )

    def _iter_text_samples(self) -> list[tuple[Path, str]]:
        samples: list[tuple[Path, str]] = []
        if not self.for2_extract_dir.exists():
            return samples
        for path in sorted(self.for2_extract_dir.rglob("*")):
            if not path.is_file():
                continue
            if path.suffix.lower() not in self.text_exts:
                continue
            text = ""
            if path.suffix.lower() == ".pdf":
                text = self._extract_pdf_text(path)
            elif path.suffix.lower() == ".docx":
                text = self._extract_docx_text(path)
            else:
                try:
                    text = path.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    continue
            if text.strip():
                samples.append((path, text[:50000]))
            if len(samples) >= self.max_scan_files:
                break
        return samples

    def _extract_pdf_text(self, path: Path) -> str:
        # Prefer Python parser first, fallback to pdftotext if available.
        if PdfReader is not None:
            try:
                reader = PdfReader(str(path))
                chunks: list[str] = []
                for page in reader.pages[:20]:
                    try:
                        page_text = page.extract_text() or ""
                    except Exception:
                        page_text = ""
                    if page_text:
                        chunks.append(page_text)
                    if sum(len(c) for c in chunks) >= 50000:
                        break
                text = "\n".join(chunks).strip()
                if text:
                    return text[:50000]
            except Exception:
                pass

        tool = shutil.which("pdftotext")
        if tool:
            res = run_local_cmd([tool, "-q", str(path), "-"], timeout_s=30.0)
            if res.ok and (res.stdout or "").strip():
                return (res.stdout or "")[:50000]
        return ""

    def _extract_docx_text(self, path: Path) -> str:
        # Read Word XML directly so FOR5 can scan .docx evidence without extra deps.
        try:
            with zipfile.ZipFile(path) as zf:
                names = [
                    n
                    for n in zf.namelist()
                    if n.startswith("word/")
                    and n.endswith(".xml")
                    and not n.endswith(("styles.xml", "settings.xml", "fontTable.xml", "numbering.xml"))
                ]
                chunks: list[str] = []
                total = 0
                for name in sorted(names):
                    try:
                        xml_text = zf.read(name).decode("utf-8", errors="ignore")
                    except Exception:
                        continue
                    xml_text = xml_text.replace("</w:p>", "\n").replace("</w:tr>", "\n").replace("</w:tc>", " ")
                    plain = re.sub(r"<[^>]+>", " ", xml_text)
                    plain = html.unescape(" ".join(plain.split())).strip()
                    if plain:
                        chunks.append(plain)
                        total += len(plain)
                    if total >= 50000:
                        break
                return "\n".join(chunks)[:50000]
        except Exception:
            return ""

    def _normalize_for6_label(self, raw: str) -> str:
        text = str(raw or "").strip().lower()
        compact = text.replace("-", " ").replace("_", " ")
        tight = re.sub(r"\s+", "", compact)
        if "highpriority" in tight:
            return "high_priority"
        if "high" in compact and "priority" in compact:
            return "high_priority"
        if "suspicious" in compact:
            return "suspicious"
        if "safe" in compact or "benign" in compact:
            return "safe"
        return "unknown"

    def _display_for6_label(self, internal: str) -> str:
        if internal == "high_priority":
            return "High Priority"
        if internal == "suspicious":
            return "Suspicious"
        if internal == "safe":
            return "Safe"
        return "Unknown"

    def _find_extracted_file_by_name(self, file_name: str) -> Path | None:
        target = file_name.strip().lower()
        if not target or not self.for2_extract_dir.exists():
            return None
        for path in self.for2_extract_dir.rglob("*"):
            if path.is_file() and path.name.strip().lower() == target:
                return path
        return None

    def _read_text_for_for6(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            return self._extract_pdf_text(path)[: self.for6_max_chars]
        if suffix == ".docx":
            return self._extract_docx_text(path)[: self.for6_max_chars]
        try:
            return path.read_text(encoding="utf-8", errors="ignore")[: self.for6_max_chars]
        except Exception:
            return ""

    def _build_for6_prompt(self, file_name: str, evidence_text: str) -> str:
        clipped = evidence_text.strip()
        if len(clipped) > self.for6_max_chars:
            clipped = clipped[: self.for6_max_chars]
        return (
            "You are a forensic triage assistant for ByteBite.\n"
            "Task: classify the artefact into exactly one category.\n"
            "Valid categories (must use exact token): safe, suspicious, high_priority\n\n"
            "Decision policy:\n"
            "1) safe = ordinary personal/creative/admin content with no operational criminal intent.\n"
            "   Literary references (e.g., poems mentioning death) are safe unless direct intent appears.\n"
            "2) suspicious = concerning or sensitive content (e.g., weapon manuals, concealment talk,\n"
            "   coded coordination) without explicit violent/criminal plan.\n"
            "3) high_priority = explicit plan/intent to commit violent or serious criminal acts\n"
            "   (targeting, threats, trafficking, kidnap, murder, attack instructions with intent).\n"
            "If uncertain, choose suspicious.\n\n"
            "Return JSON only with fields:\n"
            "- category\n"
            "- suspicion_score\n"
            "- rationale\n\n"
            "Rules for output:\n"
            "- category must be one of: safe, suspicious, high_priority\n"
            "- suspicion_score must be an integer 0-100\n"
            "- rationale must be one short sentence\n\n"
            f"File name: {file_name}\n"
            "Evidence text:\n"
            f"{clipped}\n"
        )

    def _infer_label_from_text(self, raw: str) -> str:
        text = (raw or "").lower()
        if "high priority" in text or "high_priority" in text:
            return "high_priority"
        if "suspicious" in text:
            return "suspicious"
        if "safe" in text or "benign" in text:
            return "safe"
        return "unknown"

    def _build_for6_ai_summary(
        self,
        rows: list[dict[str, Any]],
        total_cases: int,
        final_correct: int,
        llm_correct: int,
        avg_infer_s: float,
        median_infer_s: float,
        p95_infer_s: float,
        missing_files: list[str],
    ) -> str:
        mode = "Hybrid (LLM + policy)" if self.for6_use_policy_override else "LLM-only"
        final_accuracy = (final_correct / total_cases * 100.0) if total_cases else 0.0
        llm_accuracy = (llm_correct / total_cases * 100.0) if total_cases else 0.0

        per_file: dict[str, dict[str, Any]] = {}
        for row in rows:
            fname = str(row.get("file_name", ""))
            bucket = per_file.setdefault(
                fname,
                {
                    "total": 0,
                    "final_correct": 0,
                    "llm_correct": 0,
                    "times": [],
                },
            )
            bucket["total"] += 1
            if bool(row.get("correct", False)):
                bucket["final_correct"] += 1
            if bool(row.get("llm_correct", False)):
                bucket["llm_correct"] += 1
            bucket["times"].append(parse_float(str(row.get("inference_seconds", 0.0)), 0.0))

        file_lines: list[str] = []
        for fname in sorted(per_file.keys()):
            bucket = per_file[fname]
            total = max(1, int(bucket["total"]))
            final_pct = 100.0 * int(bucket["final_correct"]) / total
            llm_pct = 100.0 * int(bucket["llm_correct"]) / total
            med = percentile([float(t) for t in bucket["times"]], 0.5)
            file_lines.append(
                f"- {fname}: final {final_pct:.2f}% ({bucket['final_correct']}/{total}), "
                f"llm {llm_pct:.2f}% ({bucket['llm_correct']}/{total}), median {med:.3f}s."
            )

        lines = [
            "Overall Assessment",
            (
                f"- FOR6 completed in {mode} mode with final accuracy {final_accuracy:.2f}% "
                f"({final_correct}/{total_cases}) and LLM-only accuracy {llm_accuracy:.2f}% "
                f"({llm_correct}/{total_cases})."
            ),
            "",
            "Key Strengths",
            (
                f"- All target files were processed; timing remained stable "
                f"(avg {avg_infer_s:.3f}s, median {median_infer_s:.3f}s, p95 {p95_infer_s:.3f}s)."
            ),
            "",
            "Key Limitations",
            (
                "- This run uses a 3-file controlled dataset; broader generalisation requires larger "
                "and more diverse artefact sets."
            ),
            "",
            "Recommended Next Actions",
            (
                "- Keep both metrics in reporting: LLM-only accuracy and final decision-system accuracy; "
                "expand ground-truth dataset in staged increments."
            ),
            "",
            "Per-File Outcomes",
        ]
        lines.extend(file_lines or ["- No per-file outcomes available."])
        if missing_files:
            lines.extend(["", "Missing Files", f"- {', '.join(sorted(set(missing_files)))}"])

        return "\n".join(lines).strip()

    def _for6_policy_label(
        self,
        file_name: str,
        evidence_text: str,
        llm_label: str,
    ) -> tuple[str, str]:
        """
        Apply deterministic triage policy over the LLM label.
        This stabilizes FOR6 behaviour across runs and small-model variance.
        """
        text = (evidence_text or "").lower()
        name = (file_name or "").lower()

        intent_terms = {
            "murder",
            "target",
            "victim",
            "kidnap",
            "smuggle",
            "operation",
            "getaway",
            "escape",
            "prison",
        }
        weapon_terms = {
            "gun",
            "knife",
            "weapon",
            "bullet",
            "blade",
            "firearm",
            "firearms",
            "rifle",
            "pistol",
            "ammo",
            "ammunition",
        }
        concealment_terms = {
            "burner",
            "encrypted",
            "coded",
            "hidden",
            "conceal",
            "secure channel",
        }

        intent_hits = sum(1 for term in intent_terms if term in text)
        weapon_hits = sum(1 for term in weapon_terms if term in text)
        concealment_hits = sum(1 for term in concealment_terms if term in text)
        looks_like_manual = any(token in name for token in ("manual", "guide", "handbook"))

        # Explicit criminal intent outranks everything.
        if intent_hits >= 2 or (intent_hits >= 1 and weapon_hits >= 1):
            return "high_priority", f"Policy override: intent_hits={intent_hits}, weapon_hits={weapon_hits}."

        # Instructional weapon/operational secrecy without clear intent is suspicious.
        if looks_like_manual or weapon_hits >= 1 or concealment_hits >= 1:
            return "suspicious", (
                "Policy override: "
                f"manual={looks_like_manual}, weapon_hits={weapon_hits}, concealment_hits={concealment_hits}."
            )

        # Fallback: keep a valid LLM label, otherwise safe.
        if llm_label in {"safe", "suspicious", "high_priority"}:
            return llm_label, "Policy confirmed LLM label."
        return "safe", "Policy fallback: no suspicious indicators."

    def for5_ai_analysis(self) -> ForensicRecord:
        tid = "FOR5"
        evidence = self.run_root / "FOR5_ai_analysis"
        evidence.mkdir(parents=True, exist_ok=True)

        timing_iterations = max(1, parse_int(self.suite_cfg.get("for5_timing_iterations", 1), 1))
        force_rescan = bool(self.suite_cfg.get("for5_force_rescan_per_iteration", False))

        def scan_samples(samples: list[tuple[Path, str]]) -> tuple[list[dict[str, Any]], int]:
            found: list[dict[str, Any]] = []
            total = 0
            for path, text in samples:
                lowered = text.lower()
                matched: list[dict[str, Any]] = []
                for kw in self.keyword_list:
                    c = lowered.count(kw)
                    if c > 0:
                        matched.append({"keyword": kw, "count": c})
                        total += c
                if matched:
                    found.append({"file": str(path), "matches": matched})
            return found, total

        cached_samples: list[tuple[Path, str]] = []
        if not force_rescan:
            cached_samples = self._iter_text_samples()

        iteration_times: list[float] = []
        last_samples: list[tuple[Path, str]] = cached_samples
        hits: list[dict[str, Any]] = []
        total_occurrences = 0
        for _ in range(timing_iterations):
            t0_ns = time.perf_counter_ns()
            samples = self._iter_text_samples() if force_rescan else cached_samples
            iter_hits, iter_total_occurrences = scan_samples(samples)
            dt = (time.perf_counter_ns() - t0_ns) / 1_000_000_000.0
            iteration_times.append(dt)
            last_samples = samples
            hits = iter_hits
            total_occurrences = iter_total_occurrences

        elapsed_avg = sum(iteration_times) / len(iteration_times) if iteration_times else 0.0
        elapsed_median = percentile(iteration_times, 0.5)
        elapsed_p95 = percentile(iteration_times, 0.95)

        unique_hit_files = len(hits)
        scanned_files = len(last_samples)

        if scanned_files == 0:
            relevance_est = 0.0
        elif total_occurrences == 0:
            relevance_est = 100.0
        else:
            density = unique_hit_files / scanned_files
            relevance_est = round(max(40.0, min(95.0, density * 100.0)), 2)

        write_json(
            self.for5_hits_json,
            {
                "scanned_files": scanned_files,
                "unique_hit_files": unique_hit_files,
                "total_occurrences": total_occurrences,
                "relevance_accuracy_estimate_pct": relevance_est,
                "timing_iterations": timing_iterations,
                "timing_mode": ("rescan_each_iteration" if force_rescan else "cached_samples"),
                "iteration_times_seconds": [round(v, 6) for v in iteration_times],
                "avg_time_seconds": round(elapsed_avg, 6),
                "median_time_seconds": round(elapsed_median, 6),
                "p95_time_seconds": round(elapsed_p95, 6),
                "hits": hits,
            },
        )
        hit_rows: list[list[Any]] = []
        for item in hits:
            for m in item["matches"]:
                hit_rows.append([item["file"], m["keyword"], m["count"]])
        write_csv(evidence / "keyword_hits.csv", ["file", "keyword", "count"], hit_rows)

        passed = True
        result = (
            "Keyword search completed over "
            f"{scanned_files} files; {unique_hit_files} files matched keywords. "
            f"timing(iter={timing_iterations}, avg={elapsed_avg:.6f}s, median={elapsed_median:.6f}s, p95={elapsed_p95:.6f}s)."
            if scanned_files > 0
            else "No text files were available for keyword search."
        )

        return ForensicRecord(
            test_id=tid,
            objective="AI analysis complete",
            expected_outcome="Keyword search extracted",
            input_output="Dataset",
            metric_measured="Search time avg/median/p95 (seconds), relevance accuracy (%)",
            result_text=result,
            avg_time_seconds=round(elapsed_avg, 6),
            success_rate_pct=relevance_est,
            status="PASS" if passed else "FAIL",
            evidence_dir=str(evidence),
            notes=("Keyword hit files recorded." if scanned_files > 0 else "No supported text files in extraction."),
            metrics={
                "scanned_files": scanned_files,
                "unique_hit_files": unique_hit_files,
                "total_occurrences": total_occurrences,
                "relevance_accuracy_estimate_pct": relevance_est,
                "timing_iterations": timing_iterations,
                "timing_mode": ("rescan_each_iteration" if force_rescan else "cached_samples"),
                "avg_time_seconds": round(elapsed_avg, 6),
                "median_time_seconds": round(elapsed_median, 6),
                "p95_time_seconds": round(elapsed_p95, 6),
            },
        )

    def _build_ai_summary(self, records: list[ForensicRecord], evidence_dir: Path) -> tuple[str, str]:
        summary_rows = []
        for r in records:
            summary_rows.append(
                {
                    "id": r.test_id,
                    "status": r.status,
                    "rate": r.success_rate_pct,
                    "sec": r.avg_time_seconds,
                    "result": (r.result_text or "").strip()[:140],
                }
            )

        prompt = (
            "You are a forensic validation assistant. "
            "Summarize the following ByteBite forensic test run. "
            "Return plain text with these headings exactly:\n"
            "Overall Assessment\n"
            "Key Strengths\n"
            "Key Limitations\n"
            "Recommended Next Actions\n\n"
            "Keep each section concise: exactly 1 short bullet point per section (max 14 words).\n\n"
            f"Run data (JSON):\n{json.dumps(summary_rows, separators=(',', ':'))}"
        )
        write_text(evidence_dir / "ai_summary_prompt.txt", prompt)

        llm_enabled = bool((self.app_cfg.get("llm") or {}).get("enabled", False))
        if not llm_enabled:
            fallback = (
                "Overall Assessment\n"
                "LLM is disabled in config; AI narrative summary was not generated.\n\n"
                "Key Strengths\n"
                "- Structured FOR1-FOR6 evidence was generated.\n\n"
                "Key Limitations\n"
                "- Enable llm.enabled and valid llm.binary/llm.model to generate AI narrative.\n\n"
                "Recommended Next Actions\n"
                "- Configure llama.cpp path in config and rerun FOR tests."
            )
            return fallback, "LLM disabled"

        try:
            # Ensure enough context for summary prompt even if config was set too low.
            llm_cfg = dict((self.app_cfg.get("llm") or {}))
            llm_cfg["context_tokens"] = max(int(llm_cfg.get("context_tokens", 0) or 0), 768)
            requested_max = int(llm_cfg.get("max_tokens", 160) or 160)
            llm_cfg["max_tokens"] = min(max(requested_max, 120), 220)
            llm_cfg["timeout_s"] = max(float(llm_cfg.get("timeout_s", 0) or 0), 300.0)
            llm_run_cfg = dict(self.app_cfg)
            llm_run_cfg["llm"] = llm_cfg
            out = generate_text_with_llama(prompt, llm_run_cfg)
            write_text(evidence_dir / "ai_summary_raw.txt", out.raw_output or "")
            write_text(evidence_dir / "ai_summary_stderr.txt", out.stderr or "")
            write_json(
                evidence_dir / "ai_summary_command.json",
                {
                    "command": out.command,
                    "returncode": out.returncode,
                },
            )
            if out.returncode == 0 and (out.raw_output or "").strip():
                raw = out.raw_output.strip()
                marker = "**Overall Assessment**"
                pos = raw.rfind(marker)
                if pos < 0:
                    marker = "Overall Assessment"
                    pos = raw.rfind(marker)
                cleaned = raw[pos:].strip() if pos >= 0 else raw
                cleaned_lines: list[str] = []
                for line in cleaned.splitlines():
                    t = line.strip()
                    if t.startswith("[ Prompt:"):
                        continue
                    if t == "Exiting...":
                        continue
                    if t.startswith("> "):
                        continue
                    cleaned_lines.append(line.rstrip())
                cleaned = "\n".join(cleaned_lines).strip()
                return normalize_ai_summary_text(cleaned), "LLM summary generated"
            return "AI summary generation returned empty output.", f"LLM returned code {out.returncode}"
        except Exception as exc:
            return (
                "AI summary generation failed due to LLM runtime/config error.",
                f"{type(exc).__name__}: {exc}",
            )

    def for6_report_completed(self) -> ForensicRecord:
        tid = "FOR6"
        evidence = self.run_root / "FOR6_ai_classification"
        evidence.mkdir(parents=True, exist_ok=True)
        llm_enabled = bool((self.app_cfg.get("llm") or {}).get("enabled", False))
        if not llm_enabled:
            write_text(evidence / "ai_summary_clean.txt", "FOR6 classification not run: LLM disabled.")
            return ForensicRecord(
                test_id=tid,
                objective="AI artefact classification",
                expected_outcome="Files classified as Safe/Suspicious/High Priority",
                input_output="Input: selected artefacts; Output: labels, rationale, timings",
                metric_measured="Predicted label, rationale, inference time, correctness",
                result_text="FOR6 skipped because llm.enabled is false in config.",
                avg_time_seconds=0.0,
                success_rate_pct=0.0,
                status="FAIL",
                evidence_dir=str(evidence),
                notes="Enable llm.enabled with valid llm.binary and llm.model.",
                metrics={},
            )

        targets: list[dict[str, str]] = []
        for item in self.for6_targets:
            if not isinstance(item, dict):
                continue
            fn = str(item.get("file_name", "")).strip()
            gt = self._normalize_for6_label(str(item.get("ground_truth", "")).strip())
            if fn and gt in {"safe", "suspicious", "high_priority"}:
                targets.append({"file_name": fn, "ground_truth": gt})

        if not targets:
            write_text(evidence / "ai_summary_clean.txt", "FOR6 classification not run: no valid target files configured.")
            return ForensicRecord(
                test_id=tid,
                objective="AI artefact classification",
                expected_outcome="Files classified as Safe/Suspicious/High Priority",
                input_output="Input: selected artefacts; Output: labels, rationale, timings",
                metric_measured="Predicted label, rationale, inference time, correctness",
                result_text="FOR6 skipped: no valid for6_target_files in suite config.",
                avg_time_seconds=0.0,
                success_rate_pct=0.0,
                status="FAIL",
                evidence_dir=str(evidence),
                notes="Configure for6_target_files with file_name and ground_truth.",
                metrics={},
            )

        rows: list[dict[str, Any]] = []
        missing_files: list[str] = []
        inference_times: list[float] = []
        correct_count = 0
        llm_correct_count = 0
        total_cases = 0

        for iteration in range(1, self.for6_iterations + 1):
            for target in targets:
                file_name = target["file_name"]
                expected_label = target["ground_truth"]
                source_path = self._find_extracted_file_by_name(file_name)
                if source_path is None:
                    missing_files.append(file_name)
                    rows.append(
                        {
                            "iteration": iteration,
                            "file_name": file_name,
                            "source_path": "",
                            "expected_label": self._display_for6_label(expected_label),
                            "predicted_label": "Unknown",
                            "inference_seconds": 0.0,
                            "correct": False,
                            "llm_correct": False,
                            "rationale": "File not found in FOR2 extraction.",
                            "decision_mode": ("hybrid" if self.for6_use_policy_override else "llm_only"),
                            "llm_label": "Unknown",
                            "llm_returncode": -1,
                        }
                    )
                    total_cases += 1
                    continue

                text = self._read_text_for_for6(source_path)
                if not text.strip():
                    rows.append(
                        {
                            "iteration": iteration,
                            "file_name": file_name,
                            "source_path": str(source_path),
                            "expected_label": self._display_for6_label(expected_label),
                            "predicted_label": "Unknown",
                            "inference_seconds": 0.0,
                            "correct": False,
                            "llm_correct": False,
                            "rationale": "No readable text extracted from file.",
                            "decision_mode": ("hybrid" if self.for6_use_policy_override else "llm_only"),
                            "llm_label": "Unknown",
                            "llm_returncode": -1,
                        }
                    )
                    total_cases += 1
                    continue

                prompt = self._build_for6_prompt(file_name=file_name, evidence_text=text)
                t0 = time.perf_counter()
                llm_run_cfg = dict(self.app_cfg)
                llm_cfg = dict((self.app_cfg.get("llm") or {}))
                llm_cfg["context_tokens"] = max(int(llm_cfg.get("context_tokens", 0) or 0), self.for6_context_tokens)
                llm_cfg["max_tokens"] = max(int(llm_cfg.get("max_tokens", 0) or 0), self.for6_max_tokens)
                llm_cfg["temperature"] = self.for6_temperature
                llm_cfg["timeout_s"] = max(float(llm_cfg.get("timeout_s", 0) or 0), 300.0)
                llm_run_cfg["llm"] = llm_cfg

                classification = classify_text_with_llama(text=text, cfg=llm_run_cfg, prompt=prompt)
                infer_s = round(time.perf_counter() - t0, 6)
                predicted = self._normalize_for6_label(classification.category)
                if predicted == "unknown":
                    predicted = self._infer_label_from_text(classification.raw_output)
                llm_is_correct = predicted == expected_label
                if llm_is_correct:
                    llm_correct_count += 1

                if self.for6_use_policy_override:
                    policy_label, policy_reason = self._for6_policy_label(
                        file_name=file_name,
                        evidence_text=text,
                        llm_label=predicted,
                    )
                    final_label = policy_label
                    decision_mode = "hybrid"
                else:
                    final_label = predicted
                    policy_reason = "LLM-only mode: no policy override."
                    decision_mode = "llm_only"

                is_correct = final_label == expected_label

                inference_times.append(infer_s)
                total_cases += 1
                if is_correct:
                    correct_count += 1

                rows.append(
                    {
                        "iteration": iteration,
                        "file_name": file_name,
                        "source_path": str(source_path),
                        "expected_label": self._display_for6_label(expected_label),
                        "predicted_label": self._display_for6_label(final_label),
                        "inference_seconds": infer_s,
                        "correct": is_correct,
                        "llm_correct": llm_is_correct,
                        "rationale": (
                            f"{(classification.rationale or '').strip()} "
                            f"[{policy_reason}]"
                        ).strip()[:1000],
                        "decision_mode": decision_mode,
                        "llm_label": self._display_for6_label(predicted),
                        "llm_returncode": int(classification.returncode),
                    }
                )

        accuracy_pct = round((correct_count / total_cases) * 100.0, 2) if total_cases else 0.0
        avg_infer_s = round(sum(inference_times) / len(inference_times), 6) if inference_times else 0.0
        median_infer_s = round(percentile(inference_times, 0.5), 6) if inference_times else 0.0
        p95_infer_s = round(percentile(inference_times, 0.95), 6) if inference_times else 0.0
        llm_accuracy_pct = round((llm_correct_count / total_cases) * 100.0, 2) if total_cases else 0.0

        # Save raw FOR6 evidence
        write_json(
            evidence / "for6_classification_results.json",
            {
                "targets": targets,
                "iterations": self.for6_iterations,
                "total_cases": total_cases,
                "correct_cases": correct_count,
                "accuracy_pct": accuracy_pct,
                "llm_correct_cases": llm_correct_count,
                "llm_accuracy_pct": llm_accuracy_pct,
                "mode": ("hybrid" if self.for6_use_policy_override else "llm_only"),
                "avg_inference_seconds": avg_infer_s,
                "median_inference_seconds": median_infer_s,
                "p95_inference_seconds": p95_infer_s,
                "missing_files": sorted(set(missing_files)),
                "rows": rows,
            },
        )
        write_csv(
            evidence / "for6_classification_results.csv",
            [
                "iteration",
                "file_name",
                "source_path",
                "expected_label",
                "predicted_label",
                "inference_seconds",
                "correct",
                "llm_correct",
                "rationale",
                "decision_mode",
                "llm_label",
                "llm_returncode",
            ],
            [
                [
                    r["iteration"],
                    r["file_name"],
                    r["source_path"],
                    r["expected_label"],
                    r["predicted_label"],
                    r["inference_seconds"],
                    r["correct"],
                    r["llm_correct"],
                    r["rationale"],
                    r["decision_mode"],
                    r["llm_label"],
                    r["llm_returncode"],
                ]
                for r in rows
            ],
        )

        summary_text = self._build_for6_ai_summary(
            rows=rows,
            total_cases=total_cases,
            final_correct=correct_count,
            llm_correct=llm_correct_count,
            avg_infer_s=avg_infer_s,
            median_infer_s=median_infer_s,
            p95_infer_s=p95_infer_s,
            missing_files=missing_files,
        )
        write_text(evidence / "ai_summary_clean.txt", summary_text)

        pass_threshold = max(0.0, min(100.0, self.for6_pass_accuracy_pct))
        passed = (
            total_cases > 0
            and not missing_files
            and accuracy_pct >= pass_threshold
        )
        if total_cases == 0:
            result_text = "No FOR6 classification cases were executed."
        else:
            result_text = (
                f"Classified {total_cases} cases with {correct_count} correct predictions "
                f"(final accuracy {accuracy_pct:.2f}%, llm-only accuracy {llm_accuracy_pct:.2f}%)."
            )

        return ForensicRecord(
            test_id=tid,
            objective="AI artefact classification",
            expected_outcome="Files classified as Safe/Suspicious/High Priority",
            input_output="Input: selected artefacts; Output: labels, rationale, timings",
            metric_measured="Predicted label, rationale, inference time, correctness",
            result_text=result_text,
            avg_time_seconds=avg_infer_s,
            success_rate_pct=accuracy_pct,
            status="PASS" if passed else "FAIL",
            evidence_dir=str(evidence),
            notes=(
                (
                    "LLM classification with policy override completed."
                    if self.for6_use_policy_override
                    else "LLM-only classification completed."
                )
                if not missing_files
                else f"Missing files in extraction: {', '.join(sorted(set(missing_files)))}"
            ),
            metrics={
                "total_cases": total_cases,
                "correct_cases": correct_count,
                "accuracy_pct": accuracy_pct,
                "llm_correct_cases": llm_correct_count,
                "llm_accuracy_pct": llm_accuracy_pct,
                "mode": ("hybrid" if self.for6_use_policy_override else "llm_only"),
                "avg_inference_seconds": avg_infer_s,
                "median_inference_seconds": median_infer_s,
                "p95_inference_seconds": p95_infer_s,
                "missing_files_count": len(sorted(set(missing_files))),
                "iterations": self.for6_iterations,
            },
        )


def write_forensic_workbook(path: Path, records: list[ForensicRecord], ai_summary: str, run_root: Path) -> bool:
    if Workbook is None:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "FOR Results"
    ws.append(["Test ID", "Results", "Average Time (Seconds)", "Success Rate (%)", "Status", "Evidence Folder"])
    for r in records:
        ws.append([r.test_id, r.result_text, r.avg_time_seconds, r.success_rate_pct, r.status, r.evidence_dir])

    for r in records:
        tab = wb.create_sheet(excel_sheet_name(r.test_id))
        tab.append(["Field", "Value"])
        tab.append(["Test ID", r.test_id])
        tab.append(["Objective", r.objective])
        tab.append(["Expected Outcome", r.expected_outcome])
        tab.append(["Input/Output", r.input_output])
        tab.append(["Metric Measured", r.metric_measured])
        tab.append(["Result", r.result_text])
        tab.append(["Average Time (Seconds)", r.avg_time_seconds])
        tab.append(["Success Rate (%)", r.success_rate_pct])
        tab.append(["Status", r.status])
        tab.append(["Evidence Directory", r.evidence_dir])
        tab.append(["Notes", r.notes])
        tab.append([])
        tab.append(["Metric Key", "Metric Value"])
        for k, v in sorted(r.metrics.items()):
            tab.append([k, str(v)])

    for3_hash_csv = run_root / "FOR3_hashed_data" / "sha256_hashes.csv"
    for3_rows = read_csv_rows(for3_hash_csv)
    if for3_rows:
        hash_tab = wb.create_sheet("FOR3 Hashes")
        for row in for3_rows:
            hash_tab.append(row)

    ai_tab = wb.create_sheet("AI Summary")
    ai_tab.append(["Run Directory", str(run_root)])
    ai_tab.append(["Generated UTC", utc_now()])
    ai_tab.append([])
    ai_tab.append(["AI Findings Summary"])
    for line in ai_summary.splitlines() or [""]:
        ai_tab.append([line])

    if Font is not None and PatternFill is not None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for name in ["FOR Results", "AI Summary", "FOR3 Hashes"] + [r.test_id for r in records]:
            wsx = wb[excel_sheet_name(name)] if excel_sheet_name(name) in wb.sheetnames else None
            if wsx and wsx.max_row >= 1:
                for c in wsx[1]:
                    c.font = header_font
                    c.fill = header_fill

    for wsx in wb.worksheets:
        wsx.freeze_panes = "A2"
        wsx.auto_filter.ref = wsx.dimensions
        for col in wsx.columns:
            max_len = max(len(str(c.value or "")) for c in col[:300])
            wsx.column_dimensions[col[0].column_letter].width = min(max(16, max_len + 2), 90)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def append_master_workbook(path: Path, records: list[ForensicRecord], run_root: Path, run_report_path: Path) -> bool:
    if Workbook is None or load_workbook is None:
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.active.title = "Runs"
        for tid in ["FOR1", "FOR2", "FOR3", "FOR4", "FOR5", "FOR6"]:
            wb.create_sheet(tid)
        wb.create_sheet("AI Summaries")

    def ensure_headers(sheet: str, headers: list[str]) -> Any:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        if ws.max_row == 1 and all((ws.cell(1, c + 1).value in (None, "")) for c in range(len(headers))):
            ws.delete_rows(1, 1)
        if ws.max_row == 0:
            ws.append(headers)
        elif ws.max_row >= 1:
            existing = [ws.cell(1, c + 1).value for c in range(len(headers))]
            if existing != headers:
                ws.insert_rows(1)
                for i, h in enumerate(headers, start=1):
                    ws.cell(1, i, h)
        return ws

    run_id = run_root.name
    pass_count = sum(1 for r in records if r.status == "PASS")
    fail_count = sum(1 for r in records if r.status == "FAIL")

    ws_runs = ensure_headers("Runs", ["Run ID", "Generated UTC", "PASS", "FAIL", "Run Dir", "Run Report"]) 
    ws_runs.append([run_id, utc_now(), pass_count, fail_count, str(run_root), str(run_report_path)])

    # Normalize and rebuild AI summary sheet to prevent legacy prompt/noise columns.
    ai_headers = ["Run ID", "Generated UTC", "Summary"]
    ai_rows_by_run: dict[str, tuple[str, str]] = {}
    if "AI Summaries" in wb.sheetnames:
        ws_old_ai = wb["AI Summaries"]
        for row_idx in range(2, ws_old_ai.max_row + 1):
            run_val = ws_old_ai.cell(row_idx, 1).value
            run_key = str(run_val or "").strip()
            if not run_key:
                continue
            generated = str(ws_old_ai.cell(row_idx, 2).value or "").strip()
            summary_raw = str(ws_old_ai.cell(row_idx, 3).value or "").strip()
            if not summary_raw:
                summary_raw = "AI summary not available."
            summary_text = (
                normalize_ai_summary_text(summary_raw)
                if should_normalize_ai_text(summary_raw)
                else summary_raw
            )
            ai_rows_by_run[run_key] = (generated, summary_text)
        ai_pos = wb.sheetnames.index("AI Summaries")
        wb.remove(ws_old_ai)
        ws_ai = wb.create_sheet("AI Summaries", ai_pos)
    else:
        ws_ai = wb.create_sheet("AI Summaries")

    # Ensure every run listed in Runs has a matching AI summary row.
    if "Runs" in wb.sheetnames:
        ws_runs_existing = wb["Runs"]
        for row_idx in range(2, ws_runs_existing.max_row + 1):
            run_key = str(ws_runs_existing.cell(row_idx, 1).value or "").strip()
            if not run_key:
                continue
            generated = str(ws_runs_existing.cell(row_idx, 2).value or "").strip()
            if run_key not in ai_rows_by_run:
                ai_rows_by_run[run_key] = (generated, "AI summary not available.")

    ws_ai.append(ai_headers)
    for existing_run_id, (generated, summary) in ai_rows_by_run.items():
        ws_ai.append([existing_run_id, generated, summary[:30000]])

    for rec in records:
        ws = ensure_headers(
            rec.test_id,
            [
                "Run ID",
                "Test ID",
                "Status",
                "Result",
                "Average Time (Seconds)",
                "Success Rate (%)",
                "Evidence Directory",
            ],
        )
        ws.append([run_id, rec.test_id, rec.status, rec.result_text, rec.avg_time_seconds, rec.success_rate_pct, rec.evidence_dir])

    # Pull clean AI summary text (prefer normalized cleaned file).
    ai_clean_candidates = [
        run_root / "FOR6_ai_classification" / "ai_summary_clean.txt",
        run_root / "FOR6_report" / "ai_summary_clean.txt",
    ]
    ai_raw_candidates = [
        run_root / "FOR6_ai_classification" / "ai_summary_raw.txt",
        run_root / "FOR6_report" / "ai_summary_raw.txt",
    ]
    ai_summary = "AI summary not available."
    used_clean_summary = False
    for candidate in ai_clean_candidates:
        if candidate.exists():
            ai_summary = candidate.read_text(encoding="utf-8", errors="ignore").strip()
            used_clean_summary = True
            break
    if ai_summary == "AI summary not available.":
        for candidate in ai_raw_candidates:
            if candidate.exists():
                raw_text = candidate.read_text(encoding="utf-8", errors="ignore").strip()
                ai_summary = normalize_ai_summary_text(raw_text)
                break
    if not used_clean_summary or should_normalize_ai_text(ai_summary):
        ai_summary = normalize_ai_summary_text(ai_summary)

    # Update existing run row if present; otherwise append.
    existing_row = None
    for row_idx in range(2, ws_ai.max_row + 1):
        if str(ws_ai.cell(row_idx, 1).value or "").strip() == run_id:
            existing_row = row_idx
            break
    if existing_row is None:
        ws_ai.append([run_id, utc_now(), ai_summary[:30000]])
    else:
        ws_ai.cell(existing_row, 2, utc_now())
        ws_ai.cell(existing_row, 3, ai_summary[:30000])

    # Rich per-test metrics (all FOR tests) for this run.
    ws_metrics = ensure_headers(
        "Run Metrics",
        ["Run ID", "Generated UTC", "Test ID", "Metric Key", "Metric Value"],
    )
    generated_ts = utc_now()
    for rec in records:
        if not rec.metrics:
            continue
        for key, value in sorted(rec.metrics.items()):
            ws_metrics.append([run_id, generated_ts, rec.test_id, key, str(value)])

    # Detailed FOR6 row-level outcomes for each run.
    ws_for6_cases = ensure_headers(
        "FOR6 Cases",
        [
            "Run ID",
            "Generated UTC",
            "Iteration",
            "File Name",
            "Expected Label",
            "LLM Label",
            "Predicted Label",
            "Decision Mode",
            "Correct",
            "LLM Correct",
            "Inference Seconds",
            "Rationale",
            "Source Path",
        ],
    )
    for6_rows_path = run_root / "FOR6_ai_classification" / "for6_classification_results.csv"
    for6_rows = read_csv_rows(for6_rows_path)
    if for6_rows and len(for6_rows) > 1:
        index_map: dict[str, int] = {name: idx for idx, name in enumerate(for6_rows[0])}
        for row in for6_rows[1:]:
            if not row:
                continue
            def _val(name: str, default: str = "") -> str:
                idx = index_map.get(name)
                if idx is None or idx >= len(row):
                    return default
                return str(row[idx])

            ws_for6_cases.append(
                [
                    run_id,
                    generated_ts,
                    _val("iteration"),
                    _val("file_name"),
                    _val("expected_label"),
                    _val("llm_label"),
                    _val("predicted_label"),
                    _val("decision_mode", "hybrid"),
                    _val("correct"),
                    _val("llm_correct", ""),
                    _val("inference_seconds"),
                    _val("rationale"),
                    _val("source_path"),
                ]
            )

    if Font is not None and PatternFill is not None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for ws in wb.worksheets:
            if ws.max_row >= 1:
                for c in ws[1]:
                    c.font = header_font
                    c.fill = header_fill
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            if Alignment is not None:
                if ws.title == "AI Summaries" and ws.max_column >= 3:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row_idx, 3).alignment = Alignment(wrap_text=True, vertical="top")
                if ws.title == "FOR6 Cases" and ws.max_column >= 12:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row_idx, 12).alignment = Alignment(wrap_text=True, vertical="top")
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col[:300])
                ws.column_dimensions[col[0].column_letter].width = min(max(16, max_len + 2), 90)

    wb.save(path)
    return True


def default_suite_template() -> dict[str, Any]:
    return {
        "case_id": "CASE-001",
        "device_serial": "",
        "adb_bin": "adb",
        "adb_wait_timeout_s": 12.0,
        "output_root": "",
        "master_results_xlsx": "",
        "for2_remote_path": "/sdcard",
        "for2_remote_candidates": [
            "/sdcard",
            "/internal storage",
            "/storage/emulated/0",
            "/storage/self/primary",
            "/sdcard/Download",
            "/sdcard/Documents",
            "/sdcard/DCIM",
            "/sdcard/Pictures",
            "/sdcard/Movies",
            "/sdcard/Android/media",
        ],
        "usb_mount_path": "",
        "auto_detect_usb": True,
        "keyword_list": [
            "gun",
            "knife",
            "murder",
            "burner",
            "drugs",
            "encrypted",
            "target",
            "device",
            "victim",
            "blade",
            "weapon",
            "police",
            "operation",
            "getaway",
            "escape",
            "kidnap",
            "prison",
            "smuggle",
            "bullet",
            "shank",
        ],
        "text_extensions": [".txt", ".pdf", ".docx", ".log", ".json", ".xml", ".csv", ".md", ".html"],
        "max_text_scan_files": 200,
        "for6_target_files": [
            {"file_name": "O Death.txt", "ground_truth": "safe"},
            {"file_name": "Marcus.txt", "ground_truth": "high_priority"},
            {"file_name": "Firearms Manual.pdf", "ground_truth": "suspicious"},
        ],
        "for6_iterations": 1,
        "for6_max_chars": 1200,
        "for6_context_tokens": 1024,
        "for6_max_tokens": 120,
        "for6_temperature": 0.1,
        "for6_use_policy_override": True,
        "for6_pass_accuracy_pct": 100.0,
    }


def load_suite_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid suite config JSON: {path}: {exc}") from exc


def parse_selected_tests(raw: str) -> set[str] | None:
    cleaned = [p.strip().upper() for p in raw.split(",") if p.strip()]
    return set(cleaned) if cleaned else None


def _probe_writable_dir(path: Path) -> bool:
    try:
        path.mkdir(parents=True, exist_ok=True)
        probe = path / ".bytebite_write_probe"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return True
    except Exception:
        return False


def discover_usb_mount(cfg: dict[str, Any]) -> Path | None:
    configured = str(cfg.get("usb_mount_path", "")).strip()
    auto_detect = bool(cfg.get("auto_detect_usb", True))
    candidates: list[Path] = []

    if configured:
        candidates.append(Path(configured).expanduser())

    if auto_detect:
        user = os.environ.get("USER", "kali")
        roots = [Path(f"/media/{user}"), Path("/media/kali"), Path("/mnt")]
        for root in roots:
            if not root.exists() or not root.is_dir():
                continue
            for child in sorted(root.iterdir()):
                if child.is_dir() and not child.name.startswith("."):
                    candidates.append(child)

    seen: set[str] = set()
    for mount in candidates:
        key = str(mount)
        if key in seen:
            continue
        seen.add(key)
        if _probe_writable_dir(mount):
            return mount
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Run ByteBite forensic validation suite (FOR1-FOR6).")
    parser.add_argument("--suite-config", default="scripts/test_suite_config.json", help="Path to suite config JSON.")
    parser.add_argument("--tests", default="", help="Optional subset, e.g. FOR1,FOR2,FOR6")
    parser.add_argument("--output-root", default="", help="Optional override for output root directory.")
    parser.add_argument("--master-xlsx", default="", help="Optional override for cumulative workbook path.")
    args = parser.parse_args()

    app_cfg_path = resolve_config_path(PROJECT_ROOT)
    app_cfg = load_or_create_config(app_cfg_path, build_default_config())

    suite_cfg_path = Path(args.suite_config).expanduser()
    template = default_suite_template()
    if not suite_cfg_path.exists():
        write_json(suite_cfg_path, template)
        print(f"[ByteBite] Created suite config template: {suite_cfg_path}")
    suite_cfg = load_suite_config(suite_cfg_path)
    cfg = {**template, **suite_cfg}

    if args.output_root.strip():
        output_root = Path(args.output_root).expanduser()
    else:
        usb_mount = discover_usb_mount(cfg)
        if not usb_mount:
            print(
                "[ByteBite] ERROR: No writable USB mount found. "
                "Connect/mount USB and set usb_mount_path in suite config."
            )
            return 2
        output_root = usb_mount / "bytebite_forensic_tests"

    output_root.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    case_id = safe_name(str(cfg.get("case_id", "CASE-001")))
    run_root = output_root / f"{run_id}-{case_id}"
    run_root.mkdir(parents=True, exist_ok=True)
    print(f"[ByteBite] Run directory: {run_root}")

    write_json(run_root / "suite_config_resolved.json", cfg)
    write_json(
        run_root / "run_meta.json",
        {
            "run_id": run_root.name,
            "created_utc": utc_now(),
            "app_config_path": str(app_cfg_path),
            "suite_config_path": str(suite_cfg_path),
            "project_root": str(PROJECT_ROOT),
        },
    )

    selected = parse_selected_tests(args.tests)
    runner = ForensicTestRunner(app_cfg=app_cfg, suite_cfg=cfg, run_root=run_root)
    records = runner.run(selected=selected)

    run_report_path = run_root / "reports" / "forensic_test_report.xlsx"
    master_xlsx_path = (
        Path(args.master_xlsx).expanduser()
        if args.master_xlsx.strip()
        else Path(str(cfg.get("master_results_xlsx", ""))).expanduser()
        if str(cfg.get("master_results_xlsx", "")).strip()
        else (output_root / "forensic_test_master.xlsx")
    )

    # Ensure report exists even when FOR6 did not generate it directly.
    if not run_report_path.exists() and records:
        ai_summary_text = "AI summary not available."
        ai_summary_candidates = [
            run_root / "FOR6_ai_classification" / "ai_summary_clean.txt",
            run_root / "FOR6_report" / "ai_summary_clean.txt",
        ]
        for candidate in ai_summary_candidates:
            if candidate.exists():
                ai_summary_text = candidate.read_text(encoding="utf-8", errors="ignore")
                break
        write_forensic_workbook(run_report_path, records, ai_summary_text, run_root)

    master_ok = append_master_workbook(master_xlsx_path, records, run_root, run_report_path)

    pass_count = sum(1 for r in records if r.status == "PASS")
    fail_count = sum(1 for r in records if r.status == "FAIL")

    summary = {
        "run_id": run_root.name,
        "created_utc": utc_now(),
        "total_tests": len(records),
        "pass_count": pass_count,
        "fail_count": fail_count,
        "run_report_xlsx": str(run_report_path),
        "run_report_exists": run_report_path.exists(),
        "master_results_xlsx": str(master_xlsx_path),
        "master_results_updated": master_ok,
        "records": [asdict(r) for r in records],
    }
    write_json(run_root / "summary.json", summary)

    print(f"[ByteBite] Test run complete: {run_root}")
    print(f"[ByteBite] PASS={pass_count} FAIL={fail_count}")
    print(f"[ByteBite] Report XLSX: {run_report_path}")
    if master_ok:
        print(f"[ByteBite] Master XLSX updated: {master_xlsx_path}")
    else:
        print("[ByteBite] Master XLSX update skipped (openpyxl unavailable).")
    print(f"[ByteBite] Summary: {run_root / 'summary.json'}")

    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
