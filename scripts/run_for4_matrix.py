from __future__ import annotations

import argparse
import base64
import json
import os
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


PNG_1X1_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9W9xJx8AAAAASUVORK5CYII="
)

MIN_PDF_BYTES = b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 300 144] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>
endobj
4 0 obj
<< /Length 55 >>
stream
BT
/F1 18 Tf
72 96 Td
(ByteBite FOR4 PDF test file) Tj
ET
endstream
endobj
5 0 obj
<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
endobj
xref
0 6
0000000000 65535 f
0000000010 00000 n
0000000062 00000 n
0000000121 00000 n
0000000252 00000 n
0000000358 00000 n
trailer
<< /Root 1 0 R /Size 6 >>
startxref
433
%%EOF
"""


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def run_cmd(args: list[str], cwd: Path | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(args, cwd=str(cwd) if cwd else None, check=False, text=True, capture_output=True)


def detect_usb_mount() -> Path | None:
    user = os.environ.get("USER", "kali")
    roots = [Path(f"/media/{user}"), Path("/media/kali"), Path("/mnt")]
    for root in roots:
        if not root.exists() or not root.is_dir():
            continue
        for child in sorted(root.iterdir()):
            if child.is_dir() and not child.name.startswith("."):
                probe = child / ".bytebite_for4_probe"
                try:
                    probe.write_text("ok", encoding="utf-8")
                    probe.unlink(missing_ok=True)
                    return child
                except Exception:
                    continue
    return None


def prepare_assets(project_root: Path) -> dict[str, Path]:
    src_assets = project_root / "test_assets"
    work = src_assets / "for4_matrix"
    work.mkdir(parents=True, exist_ok=True)

    docx_src = src_assets / "keyword_story.docx"
    txt_src = src_assets / "keyword_story_10MB.txt"
    if not docx_src.exists():
        raise FileNotFoundError(f"Missing {docx_src}")
    if not txt_src.exists():
        raise FileNotFoundError(f"Missing {txt_src}")

    docx = work / "for4_story.docx"
    txt = work / "for4_story.txt"
    image = work / "for4_photo.jpg"
    pdf = work / "for4_report.pdf"

    docx.write_bytes(docx_src.read_bytes())
    txt.write_bytes(txt_src.read_bytes())
    image.write_bytes(base64.b64decode(PNG_1X1_BASE64))
    pdf.write_bytes(MIN_PDF_BYTES)

    return {"docx": docx, "txt": txt, "image": image, "pdf": pdf}


def adb(adb_bin: str, serial: str, args: list[str]) -> subprocess.CompletedProcess[str]:
    cmd = [adb_bin]
    if serial:
        cmd.extend(["-s", serial])
    cmd.extend(args)
    return run_cmd(cmd)


def _load_or_create_wb(path: Path) -> Any:
    if Workbook is None or load_workbook is None:
        return None
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


def _ensure_headers(ws: Any, headers: list[str]) -> None:
    if ws.max_row < 1:
        ws.append(headers)
        return
    existing = [str(ws.cell(1, i + 1).value or "") for i in range(len(headers))]
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)


def append_master_rows(master_xlsx: Path, rows: list[dict[str, Any]]) -> None:
    wb = _load_or_create_wb(master_xlsx)
    if wb is None:
        raise RuntimeError("openpyxl unavailable")

    raw_headers = [
        "utc",
        "case_id",
        "combination_tested",
        "run_number",
        "result",
        "time_taken_seconds",
        "run_directory",
        "notes",
    ]
    raw_ws = wb["FOR4_Runs"] if "FOR4_Runs" in wb.sheetnames else wb.create_sheet("FOR4_Runs")
    _ensure_headers(raw_ws, raw_headers)

    for row in rows:
        raw_ws.append(
            [
                row["utc"],
                row["case_id"],
                row["combination_tested"],
                row["run_number"],
                row["result"],
                row["time_taken_seconds"],
                row["run_directory"],
                row["notes"],
            ]
        )

    summary_headers = [
        "case_id",
        "combination_tested",
        "runs",
        "pass_count",
        "fail_count",
        "success_rate_pct",
        "average_time_seconds",
        "min_time_seconds",
        "max_time_seconds",
        "last_updated_utc",
    ]
    summary_ws = wb["FOR4_Summary"] if "FOR4_Summary" in wb.sheetnames else wb.create_sheet("FOR4_Summary")
    _ensure_headers(summary_ws, summary_headers)
    if summary_ws.max_row > 1:
        summary_ws.delete_rows(2, summary_ws.max_row - 1)

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows_from_sheet(raw_ws):
        key = (row["case_id"], row["combination_tested"])
        grouped.setdefault(key, []).append(row)

    for (case_id, combo), vals in sorted(grouped.items()):
        runs = len(vals)
        pass_count = sum(1 for v in vals if v["result"] == "PASS")
        fail_count = runs - pass_count
        times = [float(v["time_taken_seconds"]) for v in vals if float(v["time_taken_seconds"]) > 0]
        summary_ws.append(
            [
                case_id,
                combo,
                runs,
                pass_count,
                fail_count,
                round((pass_count / runs) * 100.0, 2) if runs else 0.0,
                round(sum(times) / len(times), 3) if times else 0.0,
                round(min(times), 3) if times else 0.0,
                round(max(times), 3) if times else 0.0,
                utc_now(),
            ]
        )

    master_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(master_xlsx)


def rows_from_sheet(ws: Any) -> list[dict[str, Any]]:
    headers = [str(ws.cell(1, i).value or "").strip() for i in range(1, ws.max_column + 1)]
    out: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, c).value
        out.append(row)
    return out


def extract_for4_record(summary_json_path: Path) -> dict[str, Any] | None:
    if not summary_json_path.exists():
        return None
    try:
        payload = json.loads(summary_json_path.read_text(encoding="utf-8"))
    except Exception:
        return None
    for rec in payload.get("records", []):
        if str(rec.get("test_id", "")).upper() == "FOR4":
            return rec
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Run FOR4-only matrix (5 runs each) and maintain a FOR4-only master workbook.")
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--adb-bin", default="adb")
    parser.add_argument("--device-serial", default="")
    parser.add_argument("--remote-base", default="/sdcard/Download/ByteBite_FOR4_SRC")
    parser.add_argument("--output-root", default="test_results/for4_only_runs")
    parser.add_argument("--master-xlsx", default="test_results/for4_only_master.xlsx")
    parser.add_argument("--suite-template", default="scripts/test_suite_config.example.json")
    parser.add_argument("--runs-per-case", type=int, default=5)
    args = parser.parse_args()

    if args.runs_per_case < 1:
        print("[ByteBite] runs-per-case must be >= 1")
        return 2

    project_root = Path(args.project_root).expanduser().resolve()
    output_root = (project_root / args.output_root).resolve()
    master_xlsx = (project_root / args.master_xlsx).resolve()
    suite_template = (project_root / args.suite_template).resolve()
    logs_dir = (project_root / "test_results" / "for4_only_logs").resolve()

    output_root.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    master_xlsx.parent.mkdir(parents=True, exist_ok=True)

    if not suite_template.exists():
        raise FileNotFoundError(f"Missing suite template: {suite_template}")

    usb_mount = detect_usb_mount()
    if not usb_mount:
        print("[ByteBite] ERROR: no writable USB mount detected for FOR4.")
        return 3
    print(f"[ByteBite] USB mount: {usb_mount}")

    devices = adb(args.adb_bin, args.device_serial.strip(), ["devices"])
    if devices.returncode != 0:
        print("[ByteBite] adb devices failed")
        print(devices.stderr.strip())
        return 4
    print(devices.stdout.strip())
    if "\tdevice" not in (devices.stdout or ""):
        print("[ByteBite] ERROR: no authorized ADB device connected.")
        return 5

    assets = prepare_assets(project_root)
    cases: list[tuple[str, str, list[str]]] = [
        ("FOR4-C1", "1x .jpg only", ["image"]),
        ("FOR4-C2", "1x .docx only", ["docx"]),
        ("FOR4-C3", "1x .txt only", ["txt"]),
        ("FOR4-C4", "1x .pdf only", ["pdf"]),
        ("FOR4-C5", "1x .pdf + .docx + .txt", ["pdf", "docx", "txt"]),
        ("FOR4-C6", "1x .jpg + .txt", ["image", "txt"]),
        ("FOR4-C7", "1x .jpg + .docx", ["image", "docx"]),
        ("FOR4-C8", "1x .jpg + .pdf", ["image", "pdf"]),
        ("FOR4-C9", "1x .jpg + .pdf + .docx + .txt", ["image", "pdf", "docx", "txt"]),
    ]

    remote_base = args.remote_base.rstrip("/")
    prep = adb(args.adb_bin, args.device_serial.strip(), ["shell", f"rm -rf {remote_base} && mkdir -p {remote_base}"])
    if prep.returncode != 0:
        print("[ByteBite] ERROR: could not prepare remote base path.")
        print(prep.stderr.strip())
        return 6

    template_cfg: dict[str, Any] = json.loads(suite_template.read_text(encoding="utf-8"))
    py = sys.executable
    all_rows: list[dict[str, Any]] = []
    overall_ok = True

    for case_id, combo_label, keys in cases:
        remote_case = f"{remote_base}/{case_id}"
        print(f"[ByteBite] Case {case_id}: {combo_label}")

        rm_case = adb(args.adb_bin, args.device_serial.strip(), ["shell", f"rm -rf {remote_case} && mkdir -p {remote_case}"])
        if rm_case.returncode != 0:
            print(f"[ByteBite] {case_id} setup failed")
            overall_ok = False
            continue

        for key in keys:
            push = adb(args.adb_bin, args.device_serial.strip(), ["push", str(assets[key]), remote_case + "/"])
            if push.returncode != 0:
                print(f"[ByteBite] {case_id} push failed for {assets[key].name}")
                overall_ok = False

        for run_idx in range(1, args.runs_per_case + 1):
            case_run_id = f"{case_id}-R{run_idx}"
            cfg = dict(template_cfg)
            cfg["case_id"] = case_run_id
            cfg["for2_remote_path"] = remote_case
            cfg["for2_remote_candidates"] = [remote_case]
            cfg["usb_mount_path"] = str(usb_mount)
            cfg["auto_detect_usb"] = False

            cfg_path = logs_dir / f"{case_run_id}.suite.json"
            cfg_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")

            cmd = [
                py,
                "-u",
                "scripts/run_test_suite.py",
                "--suite-config",
                str(cfg_path),
                "--tests",
                "FOR2,FOR4",
                "--output-root",
                str(output_root),
            ]
            proc = run_cmd(cmd, cwd=project_root)
            (logs_dir / f"{case_run_id}.stdout.log").write_text(proc.stdout or "", encoding="utf-8")
            (logs_dir / f"{case_run_id}.stderr.log").write_text(proc.stderr or "", encoding="utf-8")

            run_dirs = sorted(output_root.glob(f"*-{case_run_id}"))
            run_dir = run_dirs[-1] if run_dirs else None
            summary_json = (run_dir / "summary.json") if run_dir else None
            for4 = extract_for4_record(summary_json) if summary_json else None

            row = {
                "utc": utc_now(),
                "case_id": case_id,
                "combination_tested": combo_label,
                "run_number": run_idx,
                "result": str(for4.get("status", "FAIL")) if for4 else "FAIL",
                "time_taken_seconds": float(for4.get("avg_time_seconds", 0.0)) if for4 else 0.0,
                "run_directory": str(run_dir) if run_dir else "",
                "notes": str(for4.get("notes", "")) if for4 else "FOR4 record missing.",
            }
            all_rows.append(row)
            print(
                f"[ByteBite] {case_run_id}: {row['result']} "
                f"time={row['time_taken_seconds']:.3f}s"
            )
            if row["result"] != "PASS":
                overall_ok = False

    if Workbook is None or load_workbook is None:
        print("[ByteBite] ERROR: openpyxl is required for master workbook output.")
        return 7

    append_master_rows(master_xlsx, all_rows)
    print(f"[ByteBite] FOR4-only master: {master_xlsx}")
    print(f"[ByteBite] FOR4-only runs: {output_root}")
    print(f"[ByteBite] FOR4-only logs: {logs_dir}")
    return 0 if overall_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())

