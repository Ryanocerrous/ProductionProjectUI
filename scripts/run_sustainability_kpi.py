#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:
    Workbook = None
    load_workbook = None


TEST_IDS = ["FOR1", "FOR2", "FOR3", "FOR4", "FOR5", "FOR6"]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_vcgencmd_value(out: str) -> float | None:
    # Examples:
    #   temp=51.2'C
    #   volt=0.7063V
    text = (out or "").strip()
    if "=" not in text:
        return None
    rhs = text.split("=", 1)[1]
    num = []
    for ch in rhs:
        if ch.isdigit() or ch in ".-":
            num.append(ch)
        else:
            break
    try:
        return float("".join(num))
    except Exception:
        return None


def read_cpu_times() -> tuple[int, int]:
    # Returns (idle, total)
    with open("/proc/stat", "r", encoding="utf-8") as f:
        line = f.readline().strip()
    parts = line.split()
    vals = [int(x) for x in parts[1:]]
    idle = vals[3] + vals[4] if len(vals) > 4 else vals[3]
    total = sum(vals)
    return idle, total


def read_mem_used_mb() -> float:
    # Prefer MemAvailable for real footprint.
    mem_total_kb = 0
    mem_avail_kb = 0
    with open("/proc/meminfo", "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("MemTotal:"):
                mem_total_kb = int(line.split()[1])
            elif line.startswith("MemAvailable:"):
                mem_avail_kb = int(line.split()[1])
    if mem_total_kb <= 0:
        return 0.0
    used_kb = max(0, mem_total_kb - mem_avail_kb)
    return used_kb / 1024.0


def read_disk_bytes() -> tuple[int, int]:
    # Aggregate major block devices: sd*, mmcblk*, nvme*
    # /proc/diskstats sectors are 512-byte units.
    read_bytes = 0
    write_bytes = 0
    with open("/proc/diskstats", "r", encoding="utf-8") as f:
        for line in f:
            parts = line.split()
            if len(parts) < 14:
                continue
            name = parts[2]
            if not (
                name.startswith("sd")
                or name.startswith("mmcblk")
                or name.startswith("nvme")
            ):
                continue
            if name[-1].isdigit() and "mmcblk" not in name:
                # Skip simple partition names like sda1, nvme0n1p1
                if "p" in name or name[-1].isdigit():
                    pass
            sectors_read = int(parts[5])
            sectors_written = int(parts[9])
            read_bytes += sectors_read * 512
            write_bytes += sectors_written * 512
    return read_bytes, write_bytes


@dataclass(slots=True)
class Sample:
    ts_utc: str
    cpu_percent: float
    ram_mb: float
    io_read_mb_s: float
    io_write_mb_s: float
    temp_c: float | None
    volts_v: float | None


class Sampler:
    def __init__(self, interval_s: float = 1.0):
        self.interval_s = max(0.2, float(interval_s))
        self.samples: list[Sample] = []
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None

    def start(self) -> None:
        self._stop.clear()
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=5.0)

    def _loop(self) -> None:
        prev_idle, prev_total = read_cpu_times()
        prev_read, prev_write = read_disk_bytes()
        prev_ts = time.time()
        while not self._stop.is_set():
            time.sleep(self.interval_s)
            now_ts = time.time()
            dt_s = max(1e-6, now_ts - prev_ts)

            idle, total = read_cpu_times()
            d_idle = max(0, idle - prev_idle)
            d_total = max(1, total - prev_total)
            cpu_percent = 100.0 * (1.0 - (d_idle / d_total))
            prev_idle, prev_total = idle, total

            read_b, write_b = read_disk_bytes()
            io_read_mb_s = max(0.0, (read_b - prev_read) / (1024.0 * 1024.0 * dt_s))
            io_write_mb_s = max(0.0, (write_b - prev_write) / (1024.0 * 1024.0 * dt_s))
            prev_read, prev_write = read_b, write_b

            ram_mb = read_mem_used_mb()

            temp_c: float | None = None
            volts_v: float | None = None
            try:
                t = subprocess.run(
                    ["vcgencmd", "measure_temp"],
                    capture_output=True,
                    text=True,
                    check=False,
                    timeout=1.5,
                )
                temp_c = parse_vcgencmd_value(t.stdout)
            except Exception:
                temp_c = None
            try:
                v = subprocess.run(
                    ["vcgencmd", "measure_volts"],
                    capture_output=True,
                    text=True,
                    check=False,
                    timeout=1.5,
                )
                volts_v = parse_vcgencmd_value(v.stdout)
            except Exception:
                volts_v = None

            self.samples.append(
                Sample(
                    ts_utc=utc_now_iso(),
                    cpu_percent=round(cpu_percent, 3),
                    ram_mb=round(ram_mb, 3),
                    io_read_mb_s=round(io_read_mb_s, 3),
                    io_write_mb_s=round(io_write_mb_s, 3),
                    temp_c=round(temp_c, 3) if temp_c is not None else None,
                    volts_v=round(volts_v, 4) if volts_v is not None else None,
                )
            )
            prev_ts = now_ts


def avg(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def peak(values: list[float]) -> float:
    return max(values) if values else 0.0


def summarize_samples(
    samples: list[Sample],
    duration_s: float,
    idle_power_w: float,
    load_power_w: float,
    carbon_g_per_kwh: float,
) -> dict[str, float]:
    cpu_vals = [s.cpu_percent for s in samples]
    ram_vals = [s.ram_mb for s in samples]
    io_read_vals = [s.io_read_mb_s for s in samples]
    io_write_vals = [s.io_write_mb_s for s in samples]
    temp_vals = [s.temp_c for s in samples if s.temp_c is not None]
    volt_vals = [s.volts_v for s in samples if s.volts_v is not None]

    cpu_avg = avg(cpu_vals)
    est_power_w = idle_power_w + (load_power_w - idle_power_w) * (cpu_avg / 100.0)
    energy_wh = est_power_w * (duration_s / 3600.0)
    co2_g = (energy_wh / 1000.0) * carbon_g_per_kwh

    return {
        "cpu_avg_percent": round(cpu_avg, 3),
        "cpu_peak_percent": round(peak(cpu_vals), 3),
        "ram_avg_mb": round(avg(ram_vals), 3),
        "ram_peak_mb": round(peak(ram_vals), 3),
        "io_read_avg_mb_s": round(avg(io_read_vals), 3),
        "io_write_avg_mb_s": round(avg(io_write_vals), 3),
        "temp_avg_c": round(avg(temp_vals), 3) if temp_vals else 0.0,
        "temp_peak_c": round(peak(temp_vals), 3) if temp_vals else 0.0,
        "volts_avg_v": round(avg(volt_vals), 4) if volt_vals else 0.0,
        "duration_s": round(duration_s, 3),
        "estimated_power_w": round(est_power_w, 3),
        "energy_wh": round(energy_wh, 6),
        "co2_g": round(co2_g, 6),
    }


def write_samples_csv(path: Path, samples: list[Sample]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "ts_utc",
        "cpu_percent",
        "ram_mb",
        "io_read_mb_s",
        "io_write_mb_s",
        "temp_c",
        "volts_v",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for s in samples:
            w.writerow(
                {
                    "ts_utc": s.ts_utc,
                    "cpu_percent": s.cpu_percent,
                    "ram_mb": s.ram_mb,
                    "io_read_mb_s": s.io_read_mb_s,
                    "io_write_mb_s": s.io_write_mb_s,
                    "temp_c": s.temp_c if s.temp_c is not None else "",
                    "volts_v": s.volts_v if s.volts_v is not None else "",
                }
            )


def ensure_ws_headers(ws: Any, headers: list[str]) -> None:
    existing = [str(ws.cell(1, i + 1).value or "").strip() for i in range(len(headers))]
    if ws.max_row < 1 or not any(existing):
        ws.append(headers)
        return
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)


def write_report_xlsx(path: Path, summary_rows: list[dict[str, Any]], run_dir: Path) -> bool:
    if Workbook is None:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Summary"
    headers = [
        "test_id",
        "status",
        "return_code",
        "cpu_avg_percent",
        "cpu_peak_percent",
        "ram_avg_mb",
        "ram_peak_mb",
        "io_read_avg_mb_s",
        "io_write_avg_mb_s",
        "temp_avg_c",
        "temp_peak_c",
        "volts_avg_v",
        "duration_s",
        "estimated_power_w",
        "energy_wh",
        "co2_g",
        "notes",
        "forensic_run_dir",
    ]
    ws.append(headers)
    for row in summary_rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"

    for tid in ["IDLE"] + TEST_IDS:
        sample_path = run_dir / "samples" / f"{tid.lower()}_samples.csv"
        if not sample_path.exists():
            continue
        sws = wb.create_sheet(tid)
        with sample_path.open("r", encoding="utf-8", newline="") as f:
            r = csv.reader(f)
            for line in r:
                sws.append(line)
        sws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def append_master_xlsx(master_xlsx: Path, summary_rows: list[dict[str, Any]], run_id: str) -> bool:
    if Workbook is None:
        return False
    if master_xlsx.exists():
        wb = load_workbook(master_xlsx)  # type: ignore[arg-type]
    else:
        wb = Workbook()
        wb.active.title = "KPI Runs"
    ws = wb["KPI Runs"] if "KPI Runs" in wb.sheetnames else wb.create_sheet("KPI Runs")
    headers = [
        "run_id",
        "generated_utc",
        "test_id",
        "status",
        "return_code",
        "cpu_avg_percent",
        "cpu_peak_percent",
        "ram_avg_mb",
        "ram_peak_mb",
        "io_read_avg_mb_s",
        "io_write_avg_mb_s",
        "temp_avg_c",
        "temp_peak_c",
        "volts_avg_v",
        "duration_s",
        "estimated_power_w",
        "energy_wh",
        "co2_g",
        "notes",
        "forensic_run_dir",
    ]
    ensure_ws_headers(ws, headers)
    for row in summary_rows:
        ws.append(
            [
                run_id,
                utc_now_iso(),
                row.get("test_id", ""),
                row.get("status", ""),
                row.get("return_code", ""),
                row.get("cpu_avg_percent", ""),
                row.get("cpu_peak_percent", ""),
                row.get("ram_avg_mb", ""),
                row.get("ram_peak_mb", ""),
                row.get("io_read_avg_mb_s", ""),
                row.get("io_write_avg_mb_s", ""),
                row.get("temp_avg_c", ""),
                row.get("temp_peak_c", ""),
                row.get("volts_avg_v", ""),
                row.get("duration_s", ""),
                row.get("estimated_power_w", ""),
                row.get("energy_wh", ""),
                row.get("co2_g", ""),
                row.get("notes", ""),
                row.get("forensic_run_dir", ""),
            ]
        )
    master_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(master_xlsx)
    return True


def run_for_test(
    tests_arg: str,
    python_bin: str,
    suite_cfg: Path,
    forensic_output_root: Path,
    env: dict[str, str] | None = None,
) -> tuple[int, str, str, Path]:
    forensic_output_root.mkdir(parents=True, exist_ok=True)
    cmd = [
        python_bin,
        "-u",
        "scripts/run_test_suite.py",
        "--suite-config",
        str(suite_cfg),
        "--tests",
        tests_arg,
        "--output-root",
        str(forensic_output_root),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, env=env, check=False)
    stdout = proc.stdout or ""
    stderr = proc.stderr or ""
    run_dir = Path("")
    for line in stdout.splitlines():
        if "Run directory:" in line:
            run_dir = Path(line.split("Run directory:", 1)[1].strip())
            break
    return proc.returncode, stdout, stderr, run_dir


def tests_arg_for_kpi(test_id: str) -> str:
    # Dependency-aware execution so standalone KPI slices still satisfy
    # prerequisite data flow within run_test_suite.
    mapping = {
        "FOR1": "FOR1",
        "FOR2": "FOR2",
        "FOR3": "FOR2,FOR3",
        "FOR4": "FOR2,FOR4",
        "FOR5": "FOR2,FOR5",
        "FOR6": "FOR2,FOR6",
    }
    return mapping.get(test_id, test_id)


def make_suite_cfg_for_kpi(base_cfg: Path, out_dir: Path, test_id: str) -> Path:
    data = json.loads(base_cfg.read_text(encoding="utf-8"))
    usb_dir = out_dir / "virtual_usb"
    usb_dir.mkdir(parents=True, exist_ok=True)
    data["usb_mount_path"] = str(usb_dir)
    data["auto_detect_usb"] = False
    data["strict_negative_tests"] = False
    cfg_path = out_dir / "configs" / f"{test_id.lower()}_suite_config.json"
    cfg_path.parent.mkdir(parents=True, exist_ok=True)
    cfg_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return cfg_path


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Run sustainability KPI evaluation over FOR1..FOR6 with CPU/RAM/I/O/energy/CO2 outputs."
    )
    ap.add_argument("--suite-config", default="scripts/test_suite_config.json")
    ap.add_argument("--output-root", default="test_results/sustainability_kpi")
    ap.add_argument("--master-xlsx", default="")
    ap.add_argument("--python-bin", default=sys.executable)
    ap.add_argument("--sample-interval-s", type=float, default=1.0)
    ap.add_argument("--idle-duration-s", type=float, default=20.0)
    ap.add_argument("--idle-power-w", type=float, default=3.0, help="Estimated Pi 5 idle power.")
    ap.add_argument("--load-power-w", type=float, default=8.0, help="Estimated Pi 5 loaded power.")
    ap.add_argument("--carbon-intensity-g-per-kwh", type=float, default=250.0)
    args = ap.parse_args()

    output_root = Path(args.output_root).expanduser().resolve()
    suite_cfg = Path(args.suite_config).expanduser().resolve()
    if not suite_cfg.exists():
        print(f"[ByteBite] Missing suite config: {suite_cfg}")
        return 2

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    run_dir = output_root / f"{run_id}-FOR1-6-KPI"
    samples_dir = run_dir / "samples"
    logs_dir = run_dir / "logs"
    reports_dir = run_dir / "reports"
    forensic_output_root = run_dir / "forensic_runs"
    samples_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)

    summary_rows: list[dict[str, Any]] = []
    print(f"[ByteBite] KPI run directory: {run_dir}")

    # Idle baseline
    print("[ByteBite] Sampling idle baseline...")
    idle_sampler = Sampler(interval_s=args.sample_interval_s)
    idle_t0 = time.perf_counter()
    idle_sampler.start()
    time.sleep(max(1.0, args.idle_duration_s))
    idle_sampler.stop()
    idle_duration = time.perf_counter() - idle_t0
    write_samples_csv(samples_dir / "idle_samples.csv", idle_sampler.samples)
    idle_summary = summarize_samples(
        idle_sampler.samples,
        idle_duration,
        args.idle_power_w,
        args.load_power_w,
        args.carbon_intensity_g_per_kwh,
    )
    summary_rows.append(
        {
            "test_id": "IDLE",
            "status": "PASS",
            "return_code": 0,
            **idle_summary,
            "notes": "Idle baseline (no forensic test running).",
            "forensic_run_dir": "",
        }
    )

    env = os.environ.copy()
    for tid in TEST_IDS:
        print(f"[ByteBite] Running {tid} with KPI sampling...")
        per_test_cfg = make_suite_cfg_for_kpi(suite_cfg, run_dir, tid)
        tests_arg = tests_arg_for_kpi(tid)
        sampler = Sampler(interval_s=args.sample_interval_s)
        t0 = time.perf_counter()
        sampler.start()
        rc, stdout, stderr, forensic_run_dir = run_for_test(
            tests_arg=tests_arg,
            python_bin=args.python_bin,
            suite_cfg=per_test_cfg,
            forensic_output_root=forensic_output_root,
            env=env,
        )
        sampler.stop()
        duration = time.perf_counter() - t0

        (logs_dir / f"{tid.lower()}_stdout.log").write_text(stdout, encoding="utf-8")
        (logs_dir / f"{tid.lower()}_stderr.log").write_text(stderr, encoding="utf-8")
        write_samples_csv(samples_dir / f"{tid.lower()}_samples.csv", sampler.samples)

        stats = summarize_samples(
            sampler.samples,
            duration,
            args.idle_power_w,
            args.load_power_w,
            args.carbon_intensity_g_per_kwh,
        )
        status = "PASS" if rc == 0 else "FAIL"
        notes = (
            f"Completed (tests={tests_arg})."
            if rc == 0
            else f"run_test_suite returned non-zero (tests={tests_arg})."
        )
        summary_rows.append(
            {
                "test_id": tid,
                "status": status,
                "return_code": rc,
                **stats,
                "notes": notes,
                "forensic_run_dir": str(forensic_run_dir) if str(forensic_run_dir) else "",
            }
        )
        print(f"[ByteBite] {tid} -> {status} (rc={rc})")

    summary_csv = reports_dir / "for1_6_sustainability_summary.csv"
    fields = [
        "test_id",
        "status",
        "return_code",
        "cpu_avg_percent",
        "cpu_peak_percent",
        "ram_avg_mb",
        "ram_peak_mb",
        "io_read_avg_mb_s",
        "io_write_avg_mb_s",
        "temp_avg_c",
        "temp_peak_c",
        "volts_avg_v",
        "duration_s",
        "estimated_power_w",
        "energy_wh",
        "co2_g",
        "notes",
        "forensic_run_dir",
    ]
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(summary_rows)

    report_xlsx = reports_dir / "for1_6_sustainability_report.xlsx"
    xlsx_ok = write_report_xlsx(report_xlsx, summary_rows, run_dir)

    master_xlsx = (
        Path(args.master_xlsx).expanduser().resolve()
        if args.master_xlsx.strip()
        else output_root / "forensic_sustainability_master.xlsx"
    )
    master_ok = append_master_xlsx(master_xlsx, summary_rows, run_id)

    summary = {
        "run_id": run_id,
        "created_utc": utc_now_iso(),
        "suite_config": str(suite_cfg),
        "run_dir": str(run_dir),
        "summary_csv": str(summary_csv),
        "report_xlsx": str(report_xlsx),
        "xlsx_written": xlsx_ok,
        "master_xlsx": str(master_xlsx),
        "master_updated": master_ok,
        "rows": summary_rows,
        "assumptions": {
            "idle_power_w": args.idle_power_w,
            "load_power_w": args.load_power_w,
            "carbon_intensity_g_per_kwh": args.carbon_intensity_g_per_kwh,
        },
    }
    (run_dir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    pass_count = sum(1 for r in summary_rows if r["status"] == "PASS")
    fail_count = sum(1 for r in summary_rows if r["status"] == "FAIL")
    print(f"[ByteBite] KPI summary CSV: {summary_csv}")
    if xlsx_ok:
        print(f"[ByteBite] KPI report XLSX: {report_xlsx}")
    else:
        print("[ByteBite] KPI report XLSX skipped (openpyxl unavailable).")
    if master_ok:
        print(f"[ByteBite] KPI master XLSX: {master_xlsx}")
    else:
        print("[ByteBite] KPI master XLSX not updated (openpyxl unavailable).")
    print(f"[ByteBite] PASS={pass_count} FAIL={fail_count}")
    print(f"[ByteBite] Run summary JSON: {run_dir / 'summary.json'}")
    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
