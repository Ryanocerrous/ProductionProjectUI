#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
import statistics
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any


def utc_ts() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


@dataclass
class CmdResult:
    args: list[str]
    returncode: int
    stdout: str
    stderr: str


def run_cmd(args: list[str], timeout_s: float = 60.0) -> CmdResult:
    try:
        cp = subprocess.run(args, capture_output=True, text=True, timeout=timeout_s)
        return CmdResult(args=args, returncode=cp.returncode, stdout=cp.stdout or "", stderr=cp.stderr or "")
    except subprocess.TimeoutExpired as exc:
        return CmdResult(args=args, returncode=124, stdout=exc.stdout or "", stderr=(exc.stderr or "") + "\n[timeout]")
    except FileNotFoundError:
        return CmdResult(args=args, returncode=127, stdout="", stderr=f"{args[0]}: command not found")


def adb_cmd(adb_bin: str, serial: str, args: list[str], timeout_s: float = 60.0) -> CmdResult:
    cmd = [adb_bin]
    if serial.strip():
        cmd.extend(["-s", serial.strip()])
    cmd.extend(args)
    return run_cmd(cmd, timeout_s=timeout_s)


def detect_authorized_devices(devices_text: str) -> int:
    count = 0
    for ln in devices_text.splitlines():
        ln = ln.strip()
        if not ln or ln.lower().startswith("list of devices"):
            continue
        parts = ln.split()
        if len(parts) >= 2 and parts[1] == "device":
            count += 1
    return count


def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    sorted_vals = sorted(values)
    rank = (len(sorted_vals) - 1) * p
    lo = math.floor(rank)
    hi = math.ceil(rank)
    if lo == hi:
        return sorted_vals[int(rank)]
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (rank - lo)


def get_top_activity(adb_bin: str, serial: str) -> str:
    cmd = "dumpsys activity top | grep -iE 'mResumedActivity|topResumedActivity' | head -n 1"
    res = adb_cmd(adb_bin, serial, ["shell", cmd], timeout_s=20.0)
    return (res.stdout or "").strip() or (res.stderr or "").strip()


def build_launch_url(base_url: str, iteration: int) -> str:
    sep = "&" if "?" in base_url else "?"
    return f"{base_url}{sep}off2_iter={iteration}&off2_ts={int(time.time() * 1000)}"


def get_screen_size(adb_bin: str, serial: str) -> tuple[int, int] | None:
    res = adb_cmd(adb_bin, serial, ["shell", "wm", "size"], timeout_s=20.0)
    text = (res.stdout or "") + "\n" + (res.stderr or "")
    m = re.search(r"(\d+)x(\d+)", text)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def maybe_write_xlsx(output_root: Path, detail_rows: list[dict[str, Any]], summary_rows: list[dict[str, Any]]) -> Path | None:
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except Exception:
        return None

    run_xlsx = output_root / "reports" / "off2_browser_click.xlsx"
    run_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "OFF2_Detail"
    if detail_rows:
        headers = list(detail_rows[0].keys())
        ws_d.append(headers)
        for r in detail_rows:
            ws_d.append([r.get(h, "") for h in headers])

    ws_s = wb.create_sheet("OFF2_Summary")
    if summary_rows:
        headers = list(summary_rows[0].keys())
        ws_s.append(headers)
        for r in summary_rows:
            ws_s.append([r.get(h, "") for h in headers])
    wb.save(run_xlsx)

    master = output_root / "offensive_test_master.xlsx"
    if master.exists():
        mwb = load_workbook(master)
    else:
        mwb = Workbook()
        mwb.active.title = "OFF1"
    if "OFF2" not in mwb.sheetnames:
        mwb.create_sheet("OFF2")
    ws = mwb["OFF2"]
    headers = list(summary_rows[0].keys()) if summary_rows else []
    if headers and (ws.max_row < 1 or ws.cell(1, 1).value is None):
        ws.append(headers)
    for r in summary_rows:
        ws.append([r.get(h, "") for h in headers])
    mwb.save(master)

    return run_xlsx


def main() -> int:
    ap = argparse.ArgumentParser(description="OFF2: Open browser URL and click link area via ADB tap")
    ap.add_argument("--adb-bin", default="adb")
    ap.add_argument("--device-serial", default="")
    ap.add_argument("--url", required=True)
    ap.add_argument("--browser-package", default="")
    ap.add_argument("--tap-x", type=int, default=0, help="Tap X coord. 0 = auto-center.")
    ap.add_argument("--tap-y", type=int, default=0, help="Tap Y coord. 0 = auto (link area).")
    ap.add_argument("--iterations", type=int, default=5)
    ap.add_argument("--pre-wait-s", type=float, default=3.0)
    ap.add_argument("--post-wait-s", type=float, default=3.0)
    ap.add_argument("--between-s", type=float, default=3.0)
    ap.add_argument("--output-root", default="test_results/offensive_tests")
    ap.add_argument("--case-id", default="OFF2-BROWSER-CLICK")
    args = ap.parse_args()

    run_id = f"{utc_ts()}-{args.case_id}"
    root = Path(args.output_root).expanduser().resolve()
    run_dir = root / run_id
    evidence = run_dir / "evidence"
    reports = run_dir / "reports"
    evidence.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    print(f"[ByteBite] OFF2 run: {run_dir}")

    devices = adb_cmd(args.adb_bin, args.device_serial, ["devices"], timeout_s=20.0)
    (evidence / "adb_devices.txt").write_text((devices.stdout or "") + "\n" + (devices.stderr or ""), encoding="utf-8")
    if devices.returncode != 0:
        print("[ByteBite] adb devices failed")
        return 2

    if detect_authorized_devices(devices.stdout) == 0:
        print("[ByteBite] No authorized device")
        return 3

    # Wake/unlock once at start.
    adb_cmd(args.adb_bin, args.device_serial, ["shell", "input", "keyevent", "KEYCODE_WAKEUP"], timeout_s=10.0)
    adb_cmd(args.adb_bin, args.device_serial, ["shell", "wm", "dismiss-keyguard"], timeout_s=10.0)
    adb_cmd(args.adb_bin, args.device_serial, ["shell", "input", "keyevent", "82"], timeout_s=10.0)

    tap_x = args.tap_x
    tap_y = args.tap_y
    if tap_x <= 0 or tap_y <= 0:
        size = get_screen_size(args.adb_bin, args.device_serial)
        if size:
            w, h = size
            # Lure link is centered and large; center-screen tap is most reliable.
            tap_x = w // 2
            tap_y = int(h * 0.50)
        else:
            tap_x = 540 if tap_x <= 0 else tap_x
            tap_y = 1200 if tap_y <= 0 else tap_y
    print(f"[ByteBite] Tap coordinates: x={tap_x}, y={tap_y}")

    detail_rows: list[dict[str, Any]] = []
    total_ms: list[float] = []
    open_ms: list[float] = []
    tap_ms: list[float] = []
    success_count = 0
    now_iso = dt.datetime.now(dt.timezone.utc).isoformat()

    for i in range(1, max(1, args.iterations) + 1):
        iter_t0 = time.perf_counter()
        launch_url = build_launch_url(args.url, i)

        t0 = time.perf_counter()
        start_args = ["shell", "am", "start", "-a", "android.intent.action.VIEW", "-d", launch_url]
        if args.browser_package.strip():
            start_args.append(args.browser_package.strip())
        open_res = adb_cmd(
            args.adb_bin,
            args.device_serial,
            start_args,
            timeout_s=30.0,
        )
        open_elapsed = (time.perf_counter() - t0) * 1000.0
        time.sleep(max(0.0, args.pre_wait_s))

        t1 = time.perf_counter()
        tap_res = adb_cmd(
            args.adb_bin,
            args.device_serial,
            ["shell", "input", "tap", str(tap_x), str(tap_y)],
            timeout_s=20.0,
        )
        tap_elapsed = (time.perf_counter() - t1) * 1000.0
        time.sleep(max(0.0, args.post_wait_s))

        top_activity = get_top_activity(args.adb_bin, args.device_serial)
        iter_elapsed = (time.perf_counter() - iter_t0) * 1000.0

        ok = (open_res.returncode == 0 and tap_res.returncode == 0)
        success_count += 1 if ok else 0
        total_ms.append(iter_elapsed)
        open_ms.append(open_elapsed)
        tap_ms.append(tap_elapsed)

        detail_rows.append(
            {
                "run_id": run_id,
                "generated_utc": now_iso,
                "iteration": i,
                "url": launch_url,
                "tap_x": tap_x,
                "tap_y": tap_y,
                "open_returncode": open_res.returncode,
                "tap_returncode": tap_res.returncode,
                "success": ok,
                "open_ms": round(open_elapsed, 3),
                "tap_ms": round(tap_elapsed, 3),
                "iteration_total_ms": round(iter_elapsed, 3),
                "top_activity": top_activity,
                "open_stdout": (open_res.stdout or "").strip(),
                "open_stderr": (open_res.stderr or "").strip(),
            }
        )

        print(
            f"[ByteBite] iter={i} success={ok} open={open_elapsed:.3f}ms "
            f"tap={tap_elapsed:.3f}ms total={iter_elapsed:.3f}ms"
        )

        time.sleep(max(0.0, args.between_s))

    summary_rows = [
        {
            "run_id": run_id,
            "generated_utc": now_iso,
            "url": args.url,
            "tap_x": tap_x,
            "tap_y": tap_y,
            "iterations": args.iterations,
            "successes": success_count,
            "success_rate_percent": round((100.0 * success_count / max(1, args.iterations)), 2),
            "avg_open_ms": round(sum(open_ms) / len(open_ms), 3),
            "avg_tap_ms": round(sum(tap_ms) / len(tap_ms), 3),
            "avg_total_ms": round(sum(total_ms) / len(total_ms), 3),
            "median_total_ms": round(statistics.median(total_ms), 3),
            "p95_total_ms": round(percentile(total_ms, 0.95), 3),
            "run_total_s": round(sum(total_ms) / 1000.0, 3),
        }
    ]

    detail_csv = reports / "off2_click_details.csv"
    summary_csv = reports / "off2_click_summary.csv"

    with detail_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
        w.writeheader()
        w.writerows(detail_rows)

    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
        w.writeheader()
        w.writerows(summary_rows)

    run_xlsx = maybe_write_xlsx(root, detail_rows, summary_rows)

    payload = {
        "run_id": run_id,
        "generated_utc": now_iso,
        "run_dir": str(run_dir),
        "url": args.url,
        "tap": {"x": tap_x, "y": tap_y},
        "iterations": args.iterations,
        "successes": success_count,
        "detail_csv": str(detail_csv),
        "summary_csv": str(summary_csv),
        "run_xlsx": str(run_xlsx) if run_xlsx else None,
        "master_xlsx": str(root / "offensive_test_master.xlsx") if (root / "offensive_test_master.xlsx").exists() else None,
    }
    write_json(run_dir / "summary.json", payload)

    print(f"[ByteBite] Detail CSV: {detail_csv}")
    print(f"[ByteBite] Summary CSV: {summary_csv}")
    if run_xlsx:
        print(f"[ByteBite] Run XLSX: {run_xlsx}")
    if payload["master_xlsx"]:
        print(f"[ByteBite] Master XLSX: {payload['master_xlsx']}")
    print(f"[ByteBite] Summary JSON: {run_dir / 'summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
