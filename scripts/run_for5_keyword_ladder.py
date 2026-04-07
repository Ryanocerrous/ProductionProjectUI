from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


KEYWORDS_20 = [
    "gun",
    "knife",
    "murder",
    "burner",
    "drugs",
    "encrypted",
    "target",
    "device",
    "victim",
    "blade",
    "weapon",
    "police",
    "operation",
    "getaway",
    "escape",
    "kidnap",
    "prison",
    "smuggle",
    "bullet",
    "shank",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def run_cmd(args: list[str], cwd: Path | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(args, cwd=str(cwd) if cwd else None, check=False, text=True, capture_output=True)


def adb(adb_bin: str, serial: str, args: list[str]) -> subprocess.CompletedProcess[str]:
    cmd = [adb_bin]
    if serial:
        cmd.extend(["-s", serial])
    cmd.extend(args)
    return run_cmd(cmd)


def _load_or_create_wb(path: Path) -> Any:
    if Workbook is None or load_workbook is None:
        return None
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


def _ensure_headers(ws: Any, headers: list[str]) -> None:
    if ws.max_row < 1:
        ws.append(headers)
        return
    existing = [str(ws.cell(1, i + 1).value or "") for i in range(len(headers))]
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)


def rows_from_sheet(ws: Any) -> list[dict[str, Any]]:
    headers = [str(ws.cell(1, i).value or "").strip() for i in range(1, ws.max_column + 1)]
    out: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, c).value
        out.append(row)
    return out


def _percentile(values: list[float], pct: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return float(values[0])
    ordered = sorted(float(v) for v in values)
    rank = (len(ordered) - 1) * pct
    lo = int(rank)
    hi = min(lo + 1, len(ordered) - 1)
    if lo == hi:
        return ordered[lo]
    frac = rank - lo
    return ordered[lo] + (ordered[hi] - ordered[lo]) * frac


def _load_keyword_count_map(run_dir: Path | None, for5_record: dict[str, Any] | None) -> dict[str, int]:
    if not run_dir:
        return {}
    candidates: list[Path] = [run_dir / "FOR5_ai_analysis" / "keyword_hits.json"]
    if for5_record:
        evidence_dir = str(for5_record.get("evidence_dir", "")).strip()
        if evidence_dir:
            candidates.append(Path(evidence_dir) / "keyword_hits.json")
    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            payload = json.loads(candidate.read_text(encoding="utf-8"))
        except Exception:
            continue
        counts: dict[str, int] = {}
        for hit in payload.get("hits", []):
            for match in hit.get("matches", []):
                kw = str(match.get("keyword", "")).strip().lower()
                if not kw:
                    continue
                counts[kw] = counts.get(kw, 0) + int(match.get("count", 0) or 0)
        return counts
    return {}


def append_master(master_xlsx: Path, rows: list[dict[str, Any]]) -> None:
    wb = _load_or_create_wb(master_xlsx)
    if wb is None:
        raise RuntimeError("openpyxl unavailable")

    raw_headers = [
        "utc",
        "run_id",
        "repeat_index",
        "keyword_count",
        "keywords_used",
        "expected_keyword_hits",
        "actual_keyword_hits",
        "false_positives",
        "false_negatives",
        "precision",
        "recall",
        "f1_score",
        "strict_status",
        "for5_status",
        "for5_time_seconds",
        "scanned_files",
        "unique_hit_files",
        "total_occurrences",
        "expected_total_occurrences",
        "actual_total_occurrences",
        "run_directory",
        "notes",
    ]
    raw_ws = wb["FOR5_Runs"] if "FOR5_Runs" in wb.sheetnames else wb.create_sheet("FOR5_Runs")
    _ensure_headers(raw_ws, raw_headers)
    for row in rows:
        raw_ws.append([row.get(k, "") for k in raw_headers])

    summary_headers = [
        "keyword_count",
        "keywords_used",
        "runs",
        "strict_pass_count",
        "strict_fail_count",
        "strict_success_rate_pct",
        "avg_time_seconds",
        "median_time_seconds",
        "p95_time_seconds",
        "avg_precision",
        "avg_recall",
        "avg_f1_score",
        "avg_false_positives",
        "avg_false_negatives",
        "avg_expected_keyword_hits",
        "avg_actual_keyword_hits",
        "avg_expected_total_occurrences",
        "avg_actual_total_occurrences",
        "avg_scanned_files",
        "last_updated_utc",
    ]
    summary_ws = wb["FOR5_Summary"] if "FOR5_Summary" in wb.sheetnames else wb.create_sheet("FOR5_Summary")
    _ensure_headers(summary_ws, summary_headers)
    if summary_ws.max_row > 1:
        summary_ws.delete_rows(2, summary_ws.max_row - 1)

    grouped: dict[int, list[dict[str, Any]]] = {}
    for row in rows_from_sheet(raw_ws):
        count = int(row.get("keyword_count") or 0)
        grouped.setdefault(count, []).append(row)

    for count in sorted(grouped.keys()):
        vals = grouped[count]
        runs = len(vals)
        strict_pass_count = sum(1 for v in vals if str(v.get("strict_status")) == "PASS")
        strict_fail_count = runs - strict_pass_count
        times = [float(v.get("for5_time_seconds") or 0.0) for v in vals]
        precisions = [float(v.get("precision") or 0.0) for v in vals]
        recalls = [float(v.get("recall") or 0.0) for v in vals]
        f1s = [float(v.get("f1_score") or 0.0) for v in vals]
        fps = [float(v.get("false_positives") or 0.0) for v in vals]
        fns = [float(v.get("false_negatives") or 0.0) for v in vals]
        exp_hits = [float(v.get("expected_keyword_hits") or 0.0) for v in vals]
        act_hits = [float(v.get("actual_keyword_hits") or 0.0) for v in vals]
        exp_occ = [float(v.get("expected_total_occurrences") or 0.0) for v in vals]
        act_occ = [float(v.get("actual_total_occurrences") or 0.0) for v in vals]
        scanned = [float(v.get("scanned_files") or 0.0) for v in vals]
        summary_ws.append(
            [
                count,
                str(vals[-1].get("keywords_used") or ""),
                runs,
                strict_pass_count,
                strict_fail_count,
                round((strict_pass_count / runs) * 100.0, 2) if runs else 0.0,
                round(sum(times) / len(times), 6) if times else 0.0,
                round(_percentile(times, 0.5), 6) if times else 0.0,
                round(_percentile(times, 0.95), 6) if times else 0.0,
                round(sum(precisions) / len(precisions), 4) if precisions else 0.0,
                round(sum(recalls) / len(recalls), 4) if recalls else 0.0,
                round(sum(f1s) / len(f1s), 4) if f1s else 0.0,
                round(sum(fps) / len(fps), 3) if fps else 0.0,
                round(sum(fns) / len(fns), 3) if fns else 0.0,
                round(sum(exp_hits) / len(exp_hits), 3) if exp_hits else 0.0,
                round(sum(act_hits) / len(act_hits), 3) if act_hits else 0.0,
                round(sum(exp_occ) / len(exp_occ), 3) if exp_occ else 0.0,
                round(sum(act_occ) / len(act_occ), 3) if act_occ else 0.0,
                round(sum(scanned) / len(scanned), 3) if scanned else 0.0,
                utc_now(),
            ]
        )

    master_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(master_xlsx)


def extract_for5(summary_json: Path) -> dict[str, Any] | None:
    if not summary_json.exists():
        return None
    try:
        payload = json.loads(summary_json.read_text(encoding="utf-8"))
    except Exception:
        return None
    for rec in payload.get("records", []):
        if str(rec.get("test_id", "")).upper() == "FOR5":
            return rec
    return None


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Run FOR5 ladder: run 1 uses first keyword, run 20 uses all 20 keywords."
    )
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--adb-bin", default="adb")
    parser.add_argument("--device-serial", default="")
    parser.add_argument("--docx-path", default="test_assets/The Operation.docx")
    parser.add_argument("--remote-dir", default="/sdcard/Download/ByteBiteTest")
    parser.add_argument("--output-root", default="test_results/for5_ladder_runs")
    parser.add_argument("--master-xlsx", default="test_results/for5_keyword_ladder_master.xlsx")
    parser.add_argument("--suite-template", default="scripts/test_suite_config.example.json")
    parser.add_argument("--max-keywords", type=int, default=20, help="How many ladder steps to run (1..20).")
    parser.add_argument("--repeats", type=int, default=1, help="How many full ladder passes to run.")
    parser.add_argument(
        "--for5-timing-iterations",
        type=int,
        default=25,
        help="Internal FOR5 timing loops per ladder step (higher gives clearer timing spread).",
    )
    parser.add_argument(
        "--for5-force-rescan",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Rescan extracted files each timing iteration for more realistic timing.",
    )
    parser.add_argument(
        "--expected-occurrences-per-keyword",
        type=int,
        default=1,
        help="Expected count per keyword in the controlled FOR5 test document.",
    )
    args = parser.parse_args()

    if args.max_keywords < 1 or args.max_keywords > len(KEYWORDS_20):
        print(f"[ByteBite] ERROR: max-keywords must be between 1 and {len(KEYWORDS_20)}")
        return 1
    if int(args.repeats) < 1:
        print("[ByteBite] ERROR: repeats must be >= 1")
        return 1

    project_root = Path(args.project_root).expanduser().resolve()
    docx_path = (project_root / args.docx_path).resolve()
    output_root = (project_root / args.output_root).resolve()
    master_xlsx = (project_root / args.master_xlsx).resolve()
    suite_template = (project_root / args.suite_template).resolve()
    logs_dir = (project_root / "test_results" / "for5_ladder_logs").resolve()

    if not docx_path.exists():
        raise FileNotFoundError(f"Missing docx: {docx_path}")
    if not suite_template.exists():
        raise FileNotFoundError(f"Missing suite template: {suite_template}")

    output_root.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    master_xlsx.parent.mkdir(parents=True, exist_ok=True)

    devices = adb(args.adb_bin, args.device_serial.strip(), ["devices"])
    if devices.returncode != 0 or "\tdevice" not in (devices.stdout or ""):
        print("[ByteBite] ERROR: no authorized ADB device connected.")
        print((devices.stdout or "").strip())
        print((devices.stderr or "").strip())
        return 2

    remote_dir = args.remote_dir.rstrip("/")
    prep = adb(args.adb_bin, args.device_serial.strip(), ["shell", f"mkdir -p {remote_dir}"])
    if prep.returncode != 0:
        print("[ByteBite] ERROR: failed to create remote directory.")
        print(prep.stderr.strip())
        return 3
    push = adb(args.adb_bin, args.device_serial.strip(), ["push", str(docx_path), remote_dir + "/"])
    if push.returncode != 0:
        print("[ByteBite] ERROR: failed to upload docx to victim device.")
        print(push.stderr.strip())
        return 4

    template_cfg: dict[str, Any] = json.loads(suite_template.read_text(encoding="utf-8"))
    py = sys.executable
    rows: list[dict[str, Any]] = []
    overall_ok = True

    for rep in range(1, int(args.repeats) + 1):
        for i in range(1, args.max_keywords + 1):
            keywords_used = KEYWORDS_20[:i]
            run_id = f"FOR5-R{rep:02d}-K{i:02d}"

            cfg = dict(template_cfg)
            cfg["case_id"] = run_id
            cfg["for2_remote_path"] = remote_dir
            cfg["for2_remote_candidates"] = [remote_dir]
            cfg["keyword_list"] = keywords_used
            cfg["text_extensions"] = [".txt", ".pdf", ".docx", ".log", ".json", ".xml", ".csv", ".md", ".html"]
            cfg["auto_detect_usb"] = False
            cfg["for5_timing_iterations"] = max(1, int(args.for5_timing_iterations))
            cfg["for5_force_rescan_per_iteration"] = bool(args.for5_force_rescan)

            cfg_path = logs_dir / f"{run_id}.suite.json"
            cfg_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")

            cmd = [
                py,
                "-u",
                "scripts/run_test_suite.py",
                "--suite-config",
                str(cfg_path),
                "--tests",
                "FOR2,FOR5",
                "--output-root",
                str(output_root),
            ]
            proc = run_cmd(cmd, cwd=project_root)
            (logs_dir / f"{run_id}.stdout.log").write_text(proc.stdout or "", encoding="utf-8")
            (logs_dir / f"{run_id}.stderr.log").write_text(proc.stderr or "", encoding="utf-8")

            run_dirs = sorted(output_root.glob(f"*-{run_id}"))
            run_dir = run_dirs[-1] if run_dirs else None
            summary_json = (run_dir / "summary.json") if run_dir else None
            for5 = extract_for5(summary_json) if summary_json else None

            kw_counts = _load_keyword_count_map(run_dir, for5)
            expected_set = {k.strip().lower() for k in keywords_used}
            actual_set = {k for k, v in kw_counts.items() if int(v) > 0}
            tp = len(expected_set & actual_set)
            fp = len(actual_set - expected_set)
            fn = len(expected_set - actual_set)
            precision = (tp / (tp + fp)) if (tp + fp) > 0 else 1.0
            recall = (tp / (tp + fn)) if (tp + fn) > 0 else 1.0
            f1 = (2.0 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0.0
            expected_occ_total = len(expected_set) * max(int(args.expected_occurrences_per_keyword), 0)
            actual_occ_total = sum(int(v or 0) for v in kw_counts.values())

            metrics = dict(for5.get("metrics", {})) if for5 else {}
            base_status = str(for5.get("status", "FAIL")) if for5 else "FAIL"
            strict_status = "PASS" if (base_status == "PASS" and fp == 0 and fn == 0) else "FAIL"
            row = {
                "utc": utc_now(),
                "run_id": run_id,
                "repeat_index": rep,
                "keyword_count": i,
                "keywords_used": ", ".join(keywords_used),
                "expected_keyword_hits": len(expected_set),
                "actual_keyword_hits": tp,
                "false_positives": fp,
                "false_negatives": fn,
                "precision": round(precision, 4),
                "recall": round(recall, 4),
                "f1_score": round(f1, 4),
                "strict_status": strict_status,
                "for5_status": base_status,
                "for5_time_seconds": float(for5.get("avg_time_seconds", 0.0)) if for5 else 0.0,
                "scanned_files": int(metrics.get("scanned_files", 0)),
                "unique_hit_files": int(metrics.get("unique_hit_files", 0)),
                "total_occurrences": int(metrics.get("total_occurrences", 0)),
                "expected_total_occurrences": expected_occ_total,
                "actual_total_occurrences": actual_occ_total,
                "run_directory": str(run_dir) if run_dir else "",
                "notes": str(for5.get("notes", "")) if for5 else "FOR5 record missing.",
            }
            rows.append(row)

            print(
                f"[ByteBite] {run_id}: strict={row['strict_status']} base={row['for5_status']} "
                f"time={row['for5_time_seconds']:.6f}s "
                f"TP={tp} FP={fp} FN={fn} F1={row['f1_score']:.4f}"
            )
            if row["strict_status"] != "PASS":
                overall_ok = False

    if Workbook is None or load_workbook is None:
        print("[ByteBite] ERROR: openpyxl is required for master workbook output.")
        return 5

    append_master(master_xlsx, rows)
    print(f"[ByteBite] FOR5 ladder master: {master_xlsx}")
    print(f"[ByteBite] FOR5 ladder runs: {output_root}")
    print(f"[ByteBite] FOR5 ladder logs: {logs_dir}")
    return 0 if overall_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
