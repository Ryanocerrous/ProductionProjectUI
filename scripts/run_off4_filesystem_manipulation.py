#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import math
import statistics
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any


def utc_ts() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


@dataclass
class CmdResult:
    args: list[str]
    returncode: int
    stdout: str
    stderr: str


def run_cmd(args: list[str], timeout_s: float = 60.0) -> CmdResult:
    try:
        cp = subprocess.run(args, capture_output=True, text=True, timeout=timeout_s)
        return CmdResult(args=args, returncode=cp.returncode, stdout=cp.stdout or "", stderr=cp.stderr or "")
    except subprocess.TimeoutExpired as exc:
        return CmdResult(args=args, returncode=124, stdout=exc.stdout or "", stderr=(exc.stderr or "") + "\n[timeout]")
    except FileNotFoundError:
        return CmdResult(args=args, returncode=127, stdout="", stderr=f"{args[0]}: command not found")


def adb_cmd(adb_bin: str, serial: str, args: list[str], timeout_s: float = 60.0) -> CmdResult:
    cmd = [adb_bin]
    if serial.strip():
        cmd.extend(["-s", serial.strip()])
    cmd.extend(args)
    return run_cmd(cmd, timeout_s=timeout_s)


def detect_authorized_devices(devices_text: str) -> int:
    count = 0
    for ln in devices_text.splitlines():
        ln = ln.strip()
        if not ln or ln.lower().startswith("list of devices"):
            continue
        parts = ln.split()
        if len(parts) >= 2 and parts[1] == "device":
            count += 1
    return count


def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    sorted_vals = sorted(values)
    rank = (len(sorted_vals) - 1) * p
    lo = math.floor(rank)
    hi = math.ceil(rank)
    if lo == hi:
        return sorted_vals[int(rank)]
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (rank - lo)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    import json

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def maybe_write_xlsx(run_dir: Path, detail_rows: list[dict[str, Any]], summary_rows: list[dict[str, Any]], output_root: Path) -> Path | None:
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
    except Exception:
        return None

    run_xlsx = run_dir / "reports" / "off4_filesystem_manipulation.xlsx"
    run_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "OFF4_Detail"
    if detail_rows:
        headers = list(detail_rows[0].keys())
        ws_d.append(headers)
        for r in detail_rows:
            ws_d.append([r.get(h, "") for h in headers])

    ws_s = wb.create_sheet("OFF4_Summary")
    if summary_rows:
        headers = list(summary_rows[0].keys())
        ws_s.append(headers)
        for r in summary_rows:
            ws_s.append([r.get(h, "") for h in headers])
    wb.save(run_xlsx)

    master = output_root / "offensive_test_master.xlsx"
    if master.exists():
        mwb = load_workbook(master)
    else:
        mwb = Workbook()
        mwb.active.title = "OFF1"
    if "OFF4" not in mwb.sheetnames:
        mwb.create_sheet("OFF4")
    ws = mwb["OFF4"]
    headers = list(summary_rows[0].keys()) if summary_rows else []
    if headers and (ws.max_row < 1 or ws.cell(1, 1).value is None):
        ws.append(headers)
    for r in summary_rows:
        ws.append([r.get(h, "") for h in headers])
    mwb.save(master)

    return run_xlsx


def main() -> int:
    ap = argparse.ArgumentParser(description="OFF4: Filesystem manipulation timing via ADB mkdir/rm")
    ap.add_argument("--adb-bin", default="adb")
    ap.add_argument("--device-serial", default="")
    ap.add_argument("--device-dir", default="/sdcard/Download/bytebite_off4_fs")
    ap.add_argument("--iterations", type=int, default=10)
    ap.add_argument("--between-s", type=float, default=0.3)
    ap.add_argument("--output-root", default="test_results/offensive_tests")
    ap.add_argument("--case-id", default="OFF4-FILESYSTEM-MANIPULATION")
    args = ap.parse_args()

    run_id = f"{utc_ts()}-{args.case_id}"
    output_root = Path(args.output_root).expanduser().resolve()
    run_dir = output_root / run_id
    evidence = run_dir / "evidence"
    reports = run_dir / "reports"
    evidence.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    print(f"[ByteBite] OFF4 run: {run_dir}")

    devices = adb_cmd(args.adb_bin, args.device_serial, ["devices"], timeout_s=20.0)
    (evidence / "adb_devices.txt").write_text((devices.stdout or "") + "\n" + (devices.stderr or ""), encoding="utf-8")
    if devices.returncode != 0:
        print("[ByteBite] adb devices failed")
        return 2
    if detect_authorized_devices(devices.stdout) == 0:
        print("[ByteBite] No authorized device")
        return 3

    now_iso = dt.datetime.now(dt.timezone.utc).isoformat()
    detail_rows: list[dict[str, Any]] = []
    mkdir_ms_list: list[float] = []
    success_count = 0

    for i in range(1, max(1, args.iterations) + 1):
        rm = adb_cmd(args.adb_bin, args.device_serial, ["shell", "rm", "-rf", args.device_dir], timeout_s=20.0)

        t0 = time.perf_counter()
        mk = adb_cmd(args.adb_bin, args.device_serial, ["shell", "mkdir", "-p", args.device_dir], timeout_s=20.0)
        mkdir_ms = (time.perf_counter() - t0) * 1000.0

        verify_cmd = f'if [ -d "{args.device_dir}" ]; then echo OK; else echo MISSING; fi'
        verify = adb_cmd(args.adb_bin, args.device_serial, ["shell", verify_cmd], timeout_s=20.0)
        exists_ok = "OK" in (verify.stdout or "")
        success = (rm.returncode == 0 and mk.returncode == 0 and exists_ok)
        if success:
            success_count += 1
        mkdir_ms_list.append(mkdir_ms)

        detail_rows.append(
            {
                "run_id": run_id,
                "generated_utc": now_iso,
                "iteration": i,
                "device_dir": args.device_dir,
                "rm_returncode": rm.returncode,
                "mkdir_returncode": mk.returncode,
                "dir_exists": exists_ok,
                "mkdir_ms": round(mkdir_ms, 3),
                "success": success,
                "rm_stderr": (rm.stderr or "").strip(),
                "mkdir_stderr": (mk.stderr or "").strip(),
            }
        )
        print(f"[ByteBite] iter={i} mkdir_ms={mkdir_ms:.3f} success={success}")
        time.sleep(max(0.0, args.between_s))

    attempts = len(detail_rows)
    summary_rows = [
        {
            "run_id": run_id,
            "generated_utc": now_iso,
            "device_dir": args.device_dir,
            "attempts": attempts,
            "successes": success_count,
            "success_rate_percent": round((100.0 * success_count / max(1, attempts)), 2),
            "total_mkdir_ms": round(sum(mkdir_ms_list), 3),
            "avg_mkdir_ms": round(sum(mkdir_ms_list) / max(1, attempts), 3),
            "median_mkdir_ms": round(statistics.median(mkdir_ms_list), 3),
            "p95_mkdir_ms": round(percentile(mkdir_ms_list, 0.95), 3),
        }
    ]

    detail_csv = reports / "off4_filesystem_details.csv"
    summary_csv = reports / "off4_filesystem_summary.csv"
    with detail_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
        w.writeheader()
        w.writerows(detail_rows)
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
        w.writeheader()
        w.writerows(summary_rows)

    run_xlsx = maybe_write_xlsx(run_dir, detail_rows, summary_rows, output_root)
    payload = {
        "run_id": run_id,
        "generated_utc": now_iso,
        "run_dir": str(run_dir),
        "device_dir": args.device_dir,
        "detail_csv": str(detail_csv),
        "summary_csv": str(summary_csv),
        "run_xlsx": str(run_xlsx) if run_xlsx else None,
    }
    write_json(run_dir / "summary.json", payload)

    print(f"[ByteBite] Detail CSV: {detail_csv}")
    print(f"[ByteBite] Summary CSV: {summary_csv}")
    if run_xlsx:
        print(f"[ByteBite] Run XLSX: {run_xlsx}")
    print(f"[ByteBite] Summary JSON: {run_dir / 'summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
