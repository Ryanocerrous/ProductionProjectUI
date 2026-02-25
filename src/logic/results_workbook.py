"""Excel workbook writer for cumulative ByteBite results."""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _ensure_sheet(wb, name: str, headers: list[str]):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row == 0:
        ws.append(headers)
    elif ws.max_row == 1 and all(c.value is None for c in ws[1]):
        ws.delete_rows(1)
        ws.append(headers)
    else:
        existing = [cell.value for cell in ws[1]]
        if existing != headers:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers)
    return ws


def _delete_existing_run(ws, run_key: str, key_col: int) -> None:
    for row in range(ws.max_row, 1, -1):
        value = ws.cell(row=row, column=key_col).value
        if str(value) == run_key:
            ws.delete_rows(row, 1)


def _friendly_mode(mode: str, phase: str) -> str:
    m = (mode or "").strip().lower()
    p = (phase or "").strip().lower()
    if m == "offensive":
        return "Offensive Test"
    if m == "forensic":
        return "Forensic Test"
    if m == "comparison_phase":
        return f"Comparison ({p or 'phase'})"
    return mode or "Unknown"


def _friendly_status(status: str) -> str:
    s = (status or "").strip().lower()
    if s == "success":
        return "Success"
    if s == "error":
        return "Error"
    if s == "cancelled":
        return "Cancelled"
    if s == "skipped":
        return "Skipped"
    return status or "Unknown"


def _main_issue(payload: dict[str, Any]) -> str:
    err = str(payload.get("error", "") or "").strip()
    if err:
        return err
    for step in payload.get("steps", []) or []:
        if not bool(step.get("ok")):
            step_name = str(step.get("name", "") or "unknown_step")
            step_err = str(step.get("error", "") or "").strip()
            return f"{step_name}: {step_err}" if step_err else f"{step_name} failed"
    return ""


def _style_sheet(ws) -> None:
    try:
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    body_font = Font(size=12, name="Calibri")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = body_font
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells[:500])
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(18, max_len + 4), 64)


def _rebuild_summary(runs_ws, steps_ws, summary_ws) -> None:
    run_rows = []
    run_ids: list[str] = []
    for row in runs_ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        run_id = str(row[2] or "").strip()
        if run_id:
            run_ids.append(run_id)
        run_rows.append(
            {
                "status": str(row[6] or ""),
                "elapsed_s": _to_float(row[7], 0.0),
            }
        )

    total_runs = len(run_rows)
    success_count = sum(1 for r in run_rows if r["status"] == "success")
    success_rate = (success_count / total_runs * 100.0) if total_runs else 0.0
    mean_elapsed = sum(r["elapsed_s"] for r in run_rows) / total_runs if total_runs else 0.0

    by_step: dict[str, list[float]] = defaultdict(list)
    for row in steps_ws.iter_rows(min_row=2, values_only=True):
        step_name = str(row[6] or "").strip()
        duration_ms = _to_float(row[8], 0.0)
        if step_name:
            by_step[step_name].append(duration_ms)
    bottlenecks = sorted(
        ((name, sum(vals) / len(vals), max(vals), len(vals)) for name, vals in by_step.items() if vals),
        key=lambda x: x[1],
        reverse=True,
    )[:10]

    run_last = max(2, runs_ws.max_row)
    run_ids_range = f"'Runs'!$C$2:$C${run_last}"
    run_lookup_range = f"'Runs'!$C$2:$C${run_last}"
    mode_range = f"'Runs'!$D$2:$D${run_last}"
    phase_range = f"'Runs'!$E$2:$E${run_last}"
    profile_range = f"'Runs'!$F$2:$F${run_last}"
    status_range = f"'Runs'!$G$2:$G${run_last}"
    elapsed_range = f"'Runs'!$H$2:$H${run_last}"
    step_count_range = f"'Runs'!$I$2:$I${run_last}"
    failed_steps_range = f"'Runs'!$J$2:$J${run_last}"
    error_range = f"'Runs'!$K$2:$K${run_last}"
    run_json_range = f"'Runs'!$L$2:$L${run_last}"

    summary_ws.delete_rows(1, summary_ws.max_row)
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Selected Run ID", ""])
    summary_ws.append(["Last updated UTC", datetime.now(timezone.utc).isoformat()])
    summary_ws.append(["Total runs", total_runs])
    summary_ws.append(["Successes", success_count])
    summary_ws.append(["Success rate (%)", round(success_rate, 2)])
    summary_ws.append(["Mean elapsed (s)", round(mean_elapsed, 3)])
    summary_ws.append([])
    summary_ws.append(["Selected Run Detail", "Value"])
    summary_ws.append(["Mode", f'=IFERROR(INDEX({mode_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Phase", f'=IFERROR(INDEX({phase_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Profile", f'=IFERROR(INDEX({profile_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Status", f'=IFERROR(INDEX({status_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Elapsed (s)", f'=IFERROR(INDEX({elapsed_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Step count", f'=IFERROR(INDEX({step_count_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Failed steps", f'=IFERROR(INDEX({failed_steps_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Error", f'=IFERROR(INDEX({error_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append(["Run JSON", f'=IFERROR(INDEX({run_json_range},MATCH($B$2,{run_lookup_range},0)),"")'])
    summary_ws.append([])
    summary_ws.append(["Step", "Mean ms", "Max ms", "Samples"])
    for name, mean_ms, max_ms, samples in bottlenecks:
        summary_ws.append([name, round(mean_ms, 2), round(max_ms, 2), samples])

    # Build an in-sheet helper list for robust dropdown behavior across Excel viewers.
    summary_ws["H1"] = "run_ids"
    for idx, rid in enumerate(run_ids, start=2):
        summary_ws.cell(row=idx, column=8, value=rid)
    for idx in range(len(run_ids) + 2, max(summary_ws.max_row + 1, 1000)):
        summary_ws.cell(row=idx, column=8, value=None)
    summary_ws.column_dimensions["H"].hidden = True
    run_list_end = max(2, len(run_ids) + 1)
    dropdown_formula = f"=$H$2:$H${run_list_end}"

    try:
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception:
        _style_sheet(summary_ws)
        return

    # Recreate dropdown validation each rebuild so stale ranges are removed.
    summary_ws.data_validations.dataValidation = []
    validation = DataValidation(type="list", formula1=dropdown_formula, allow_blank=True)
    validation.error = "Choose a Run ID from the dropdown."
    validation.errorTitle = "Invalid Run ID"
    summary_ws.add_data_validation(validation)
    validation.add("B2")

    if total_runs > 0 and not summary_ws["B2"].value:
        summary_ws["B2"] = run_ids[0] if run_ids else f"=INDEX({run_ids_range},1)"

    _style_sheet(summary_ws)


def append_run_to_workbook(workbook_path: Path, run_json_path: Path, payload: dict[str, Any]) -> bool:
    """Append/update one run in a cumulative .xlsx workbook.

    Returns False when openpyxl is unavailable; True on success.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        return False

    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    run_json_path = Path(run_json_path)
    run_key = run_json_path.resolve().as_posix()
    meta = payload.get("meta", {}) or {}

    if workbook_path.exists():
        wb = load_workbook(workbook_path)
    else:
        wb = Workbook()
        wb.active.title = "Runs"

    runs_ws = _ensure_sheet(
        wb,
        "Runs",
        [
            "logged_utc",
            "run_key",
            "run_id",
            "mode",
            "phase",
            "profile",
            "status",
            "elapsed_s",
            "step_count",
            "failed_step_count",
            "error",
            "run_json",
        ],
    )
    steps_ws = _ensure_sheet(
        wb,
        "Steps",
        [
            "logged_utc",
            "run_key",
            "run_id",
            "mode",
            "phase",
            "step_index",
            "step_name",
            "ok",
            "duration_ms",
            "error",
            "details_json",
        ],
    )
    summary_ws = _ensure_sheet(wb, "Summary", ["Metric", "Value"])
    easy_ws = _ensure_sheet(
        wb,
        "Easy Read",
        [
            "Logged (UTC)",
            "Run ID",
            "Test Type",
            "Result",
            "Duration (s)",
            "Steps Passed",
            "Steps Failed",
            "Main Issue",
            "Run File",
        ],
    )

    _delete_existing_run(runs_ws, run_key=run_key, key_col=2)
    _delete_existing_run(steps_ws, run_key=run_key, key_col=2)
    _delete_existing_run(easy_ws, run_key=run_key, key_col=9)

    steps = payload.get("steps", []) or []
    failed_steps = sum(1 for s in steps if not bool(s.get("ok")))
    passed_steps = len(steps) - failed_steps
    now = datetime.now(timezone.utc).isoformat()

    runs_ws.append(
        [
            now,
            run_key,
            str(meta.get("run_id", "")),
            str(meta.get("mode", "")),
            str(meta.get("phase", "")),
            str(meta.get("profile", "")),
            str(payload.get("status", "")),
            _to_float(payload.get("elapsed_s"), 0.0),
            len(steps),
            failed_steps,
            str(payload.get("error", "") or ""),
            run_json_path.as_posix(),
        ]
    )

    for idx, step in enumerate(steps, start=1):
        steps_ws.append(
            [
                now,
                run_key,
                str(meta.get("run_id", "")),
                str(meta.get("mode", "")),
                str(meta.get("phase", "")),
                idx,
                str(step.get("name", "")),
                bool(step.get("ok")),
                _to_float(step.get("duration_ms"), 0.0),
                str(step.get("error", "") or ""),
                str(step.get("details", "") or ""),
            ]
        )

    easy_ws.append(
        [
            now,
            str(meta.get("run_id", "")),
            _friendly_mode(str(meta.get("mode", "")), str(meta.get("phase", ""))),
            _friendly_status(str(payload.get("status", ""))),
            round(_to_float(payload.get("elapsed_s"), 0.0), 3),
            passed_steps,
            failed_steps,
            _main_issue(payload),
            run_key,
        ]
    )

    _rebuild_summary(runs_ws, steps_ws, summary_ws)
    _style_sheet(runs_ws)
    _style_sheet(steps_ws)
    _style_sheet(easy_ws)
    wb.save(workbook_path)
    return True
