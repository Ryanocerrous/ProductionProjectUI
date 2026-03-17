from __future__ import annotations

import csv
import json
import os
import re
import shutil
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from PIL import ExifTags, Image
except Exception:  # pragma: no cover - optional dependency
    ExifTags = None
    Image = None

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover - optional dependency
    Workbook = None

from logic.local_llm import classify_text_with_llama, generate_text_with_llama

TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".csv",
    ".log",
    ".json",
    ".xml",
    ".html",
    ".htm",
    ".rtf",
    ".eml",
    ".msg",
    ".yaml",
    ".yml",
    ".ini",
    ".conf",
    ".sql",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp", ".tif", ".tiff", ".heic"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".avi", ".mkv", ".wmv", ".3gp", ".webm", ".m4v"}
MESSAGE_HINTS = ("message", "messages", "sms", "mms", "chat", "whatsapp", "telegram", "signal", "imessage")

ISO_TS = re.compile(r"\b\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:?\d{2})?\b")
SIMPLE_TS = re.compile(r"\b\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}\b")
LOGCAT_TS = re.compile(r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}\b")
EPOCH_TS = re.compile(r"\b1\d{9,12}\b")

SEVERITY = {"safe": 1, "suspicious": 2, "high_priority": 3}


@dataclass(slots=True)
class TriageRow:
    source_path: str
    relative_path: str
    category: str
    size_bytes: int
    modified_utc: str
    label: str
    rule_label: str
    llm_label: str
    rule_score: int
    llm_score: int
    high_hits: str
    suspicious_hits: str
    rationale: str
    sorted_path: str
    error: str = ""


@dataclass(slots=True)
class TimelineEvent:
    timestamp_utc: str
    source_path: str
    event_type: str
    confidence: str
    summary: str


def run_post_extraction_analysis(
    *,
    source_dir: Path,
    cfg: dict[str, Any],
) -> dict[str, Any]:
    analysis_cfg = dict((cfg.get("forensic_analysis") or {}))
    if not bool(analysis_cfg.get("enabled", True)):
        return {"enabled": False}

    src_root = _resolve_source_dir(Path(source_dir), analysis_cfg)
    if not src_root.exists() or not src_root.is_dir():
        raise FileNotFoundError(f"forensic analysis source_dir not found: {src_root}")

    out_root = _resolve_output_dir(src_root, analysis_cfg)
    triage_root = out_root / "triage"
    timeline_root = out_root / "timeline"
    triage_root.mkdir(parents=True, exist_ok=True)
    timeline_root.mkdir(parents=True, exist_ok=True)

    rows = _run_triage(src_root=src_root, triage_root=triage_root, cfg=cfg, analysis_cfg=analysis_cfg)
    events = _build_timeline(src_root=src_root, analysis_cfg=analysis_cfg)

    triage_csv = triage_root / "triage_manifest.csv"
    triage_json = triage_root / "triage_manifest.json"
    _write_csv(triage_csv, [asdict(r) for r in rows])
    triage_json.write_text(json.dumps([asdict(r) for r in rows], indent=2), encoding="utf-8")
    triage_xlsx = triage_root / "triage_manifest.xlsx"
    triage_xlsx_created = _write_xlsx(triage_xlsx, [asdict(r) for r in rows], "Triage")

    timeline_csv = timeline_root / "timeline.csv"
    timeline_json = timeline_root / "timeline.json"
    _write_csv(timeline_csv, [asdict(e) for e in events])
    timeline_json.write_text(json.dumps([asdict(e) for e in events], indent=2), encoding="utf-8")
    timeline_xlsx = timeline_root / "timeline.xlsx"
    timeline_xlsx_created = _write_xlsx(timeline_xlsx, [asdict(e) for e in events], "Timeline")

    label_counts: dict[str, int] = {"safe": 0, "suspicious": 0, "high_priority": 0}
    for row in rows:
        label_counts[row.label] = label_counts.get(row.label, 0) + 1

    llm_summary_text = _build_llm_findings_summary(rows=rows, events=events, cfg=cfg, analysis_cfg=analysis_cfg)
    llm_summary_txt = out_root / "llm_findings_summary.txt"
    llm_summary_md = out_root / "llm_findings_summary.md"
    llm_summary_txt.write_text(llm_summary_text, encoding="utf-8")
    llm_summary_md.write_text(_to_markdown_summary(llm_summary_text), encoding="utf-8")

    investigator_report_xlsx = out_root / str(analysis_cfg.get("investigator_report_name", "investigator_report.xlsx"))
    investigator_report_created = _write_investigator_report(
        investigator_report_xlsx,
        rows=rows,
        events=events,
        llm_summary_text=llm_summary_text,
        source_dir=src_root,
        output_dir=out_root,
        label_counts=label_counts,
    )

    summary = {
        "enabled": True,
        "source_dir": str(src_root),
        "output_dir": str(out_root),
        "triage_count": len(rows),
        "timeline_event_count": len(events),
        "label_counts": label_counts,
        "triage_csv": str(triage_csv),
        "timeline_csv": str(timeline_csv),
        "triage_xlsx": str(triage_xlsx),
        "timeline_xlsx": str(timeline_xlsx),
        "triage_xlsx_created": triage_xlsx_created,
        "timeline_xlsx_created": timeline_xlsx_created,
        "llm_findings_summary_txt": str(llm_summary_txt),
        "llm_findings_summary_md": str(llm_summary_md),
        "investigator_report_xlsx": str(investigator_report_xlsx),
        "investigator_report_created": investigator_report_created,
    }
    (out_root / "analysis_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return summary


def _build_llm_findings_summary(
    *,
    rows: list[TriageRow],
    events: list[TimelineEvent],
    cfg: dict[str, Any],
    analysis_cfg: dict[str, Any],
) -> str:
    high = [r for r in rows if r.label == "high_priority"]
    suspicious = [r for r in rows if r.label == "suspicious"]
    safe = [r for r in rows if r.label == "safe"]

    max_items = max(3, int(analysis_cfg.get("summary_max_items", 12)))
    top_rows = sorted(rows, key=lambda r: (SEVERITY.get(r.label, 0), r.rule_score, r.llm_score), reverse=True)[:max_items]
    event_preview = events[: max_items * 2]
    bullets_rows = []
    for row in top_rows:
        bits = [row.relative_path, row.label, f"rule={row.rule_score}"]
        if row.high_hits:
            bits.append(f"high_hits={row.high_hits}")
        if row.suspicious_hits:
            bits.append(f"suspicious_hits={row.suspicious_hits}")
        bullets_rows.append(" | ".join(bits))
    bullets_events = [f"{e.timestamp_utc} | {e.source_path} | {e.event_type} | {e.summary}" for e in event_preview]

    fallback = _fallback_narrative(rows=rows, events=events)
    if not bool((cfg.get("llm") or {}).get("enabled", False)):
        return fallback + "\n\n[LLM status] LLM disabled in config, summary produced by rule engine."

    prompt = (
        "You are writing a forensic investigator briefing.\n"
        "Produce a concise, professional narrative with headings:\n"
        "1) Executive Summary\n"
        "2) High Priority Findings\n"
        "3) Suspicious Patterns\n"
        "4) Timeline Highlights\n"
        "5) Recommended Next Actions\n"
        "Do not use markdown tables. Keep it readable for non-technical investigators.\n\n"
        f"Counts: total={len(rows)}, high_priority={len(high)}, suspicious={len(suspicious)}, safe={len(safe)}\n"
        "Top triage entries:\n- "
        + "\n- ".join(bullets_rows if bullets_rows else ["None"])
        + "\n\nTimeline highlights:\n- "
        + "\n- ".join(bullets_events if bullets_events else ["None"])
    )
    try:
        generation = generate_text_with_llama(prompt, cfg)
        text = (generation.raw_output or "").strip()
        if text:
            return text
        if generation.stderr:
            return fallback + f"\n\n[LLM status] Empty output. stderr: {generation.stderr[:300]}"
    except Exception as exc:
        return fallback + f"\n\n[LLM status] LLM summary failed: {exc}"
    return fallback + "\n\n[LLM status] LLM returned no summary text."


def _fallback_narrative(*, rows: list[TriageRow], events: list[TimelineEvent]) -> str:
    high = [r for r in rows if r.label == "high_priority"]
    suspicious = [r for r in rows if r.label == "suspicious"]
    safe = [r for r in rows if r.label == "safe"]

    lines = [
        "Executive Summary",
        f"- Total artefacts triaged: {len(rows)}",
        f"- High priority: {len(high)} | Suspicious: {len(suspicious)} | Safe: {len(safe)}",
        "",
        "High Priority Findings",
    ]
    if high:
        for row in high[:8]:
            lines.append(f"- {row.relative_path} ({row.category}) hits={row.high_hits or 'n/a'}")
    else:
        lines.append("- No high-priority artefacts were flagged.")
    lines.extend(["", "Timeline Highlights"])
    for event in events[:12]:
        lines.append(f"- {event.timestamp_utc}: {event.source_path} ({event.event_type})")
    lines.extend(["", "Recommended Next Actions", "- Preserve high-priority files and verify context with source logs."])
    return "\n".join(lines).strip()


def _to_markdown_summary(text: str) -> str:
    title = "# ByteBite LLM Findings Summary\n\n"
    return title + text.strip() + "\n"


def _write_investigator_report(
    path: Path,
    *,
    rows: list[TriageRow],
    events: list[TimelineEvent],
    llm_summary_text: str,
    source_dir: Path,
    output_dir: Path,
    label_counts: dict[str, int],
) -> bool:
    if Workbook is None:
        return False

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Case Summary"
    ws_summary.append(["Field", "Value"])
    ws_summary.append(["Generated UTC", datetime.now(timezone.utc).isoformat()])
    ws_summary.append(["Source Directory", str(source_dir)])
    ws_summary.append(["Output Directory", str(output_dir)])
    ws_summary.append(["Total Triaged Artefacts", len(rows)])
    ws_summary.append(["High Priority", label_counts.get("high_priority", 0)])
    ws_summary.append(["Suspicious", label_counts.get("suspicious", 0)])
    ws_summary.append(["Safe", label_counts.get("safe", 0)])
    ws_summary.append(["Timeline Events", len(events)])
    ws_summary.freeze_panes = "A2"

    ws_llm = wb.create_sheet("LLM Narrative")
    ws_llm.append(["LLM Findings Summary"])
    for line in llm_summary_text.splitlines():
        ws_llm.append([line])
    ws_llm.column_dimensions["A"].width = 140

    ws_triage = wb.create_sheet("Triage")
    triage_rows = [asdict(r) for r in rows]
    if triage_rows:
        headers = list(triage_rows[0].keys())
        ws_triage.append(headers)
        for row in triage_rows:
            ws_triage.append([row.get(h, "") for h in headers])
        ws_triage.freeze_panes = "A2"

    ws_timeline = wb.create_sheet("Timeline")
    timeline_rows = [asdict(e) for e in events]
    if timeline_rows:
        headers = list(timeline_rows[0].keys())
        ws_timeline.append(headers)
        for row in timeline_rows:
            ws_timeline.append([row.get(h, "") for h in headers])
        ws_timeline.freeze_panes = "A2"

    wb.save(path)
    return True


def _resolve_source_dir(default_source: Path, analysis_cfg: dict[str, Any]) -> Path:
    override = str(analysis_cfg.get("source_dir", "")).strip()
    if not override:
        return default_source
    path = Path(override).expanduser()
    return path if path.is_absolute() else (default_source / path)


def _resolve_output_dir(src_root: Path, analysis_cfg: dict[str, Any]) -> Path:
    override = str(analysis_cfg.get("output_dir", "")).strip()
    if override:
        out = Path(override).expanduser()
        return out if out.is_absolute() else (src_root / out)
    name = str(analysis_cfg.get("output_dir_name", "analysis")).strip() or "analysis"
    return src_root / name


def _run_triage(*, src_root: Path, triage_root: Path, cfg: dict[str, Any], analysis_cfg: dict[str, Any]) -> list[TriageRow]:
    max_files = max(1, int(analysis_cfg.get("max_files", 500)))
    max_text_chars = max(300, int(analysis_cfg.get("max_text_chars", 2200)))
    max_text_bytes = max(2048, int(analysis_cfg.get("max_text_bytes", 1024 * 1024)))
    copy_mode = str(analysis_cfg.get("copy_mode", "link_or_copy")).strip().lower()
    ignore_dirs = {str(d).strip().lower() for d in (analysis_cfg.get("ignore_dirs") or []) if str(d).strip()}
    if not ignore_dirs:
        ignore_dirs = {"analysis", "triage", "timeline", "__pycache__"}

    high_kw = [str(x).strip().lower() for x in (analysis_cfg.get("high_priority_keywords") or []) if str(x).strip()]
    susp_kw = [str(x).strip().lower() for x in (analysis_cfg.get("suspicious_keywords") or []) if str(x).strip()]
    if not high_kw:
        high_kw = ["murder", "kill", "knife", "gun", "firearm", "bomb", "explosive", "kidnap", "ransom"]
    if not susp_kw:
        susp_kw = ["burner", "encrypted", "vault", "wipe", "delete", "crypto", "cash", "drugs", "weapon", "hide"]

    rows: list[TriageRow] = []
    for index, path in enumerate(_iter_files(src_root, ignore_dirs), start=1):
        if index > max_files:
            break
        relative = path.relative_to(src_root).as_posix()
        category = _categorize(path)
        st = path.stat()
        modified_utc = datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat()
        sample = _extract_text_sample(path, max_text_chars=max_text_chars, max_text_bytes=max_text_bytes)
        evidence = f"FILE: {relative}\nCATEGORY: {category}\nTEXT:\n{sample}".strip()
        lowered = evidence.lower()

        high_hits = _keyword_hits(lowered, high_kw)
        susp_hits = _keyword_hits(lowered, susp_kw)
        rule_score = min(100, len(high_hits) * 35 + len(susp_hits) * 15)
        rule_label = _rule_label(rule_score, high_hits)

        llm_label = ""
        llm_score = 0
        rationale = f"rule-based: high_hits={high_hits or []}, suspicious_hits={susp_hits or []}"
        error = ""
        final_label = rule_label
        if bool((cfg.get("llm") or {}).get("enabled", False)):
            try:
                prompt = (
                    "You are a forensic triage assistant.\n"
                    "Classify this artifact into exactly one label: safe, suspicious, high_priority.\n"
                    "Return JSON only with fields: category, suspicion_score, rationale.\n"
                    "Treat direct violence/weapon planning terms as high_priority.\n\n"
                    f"{evidence}\n"
                )
                resp = classify_text_with_llama(evidence, cfg, prompt=prompt)
                llm_label = _normalize_label(resp.category)
                llm_score = int(resp.suspicion_score)
                if resp.rationale:
                    rationale = resp.rationale
            except Exception as exc:
                error = f"llm_error: {exc}"

        final_label = _merge_labels(rule_label, llm_label)
        sorted_path = _place_file(path=path, src_root=src_root, triage_root=triage_root, label=final_label, copy_mode=copy_mode)

        rows.append(
            TriageRow(
                source_path=str(path),
                relative_path=relative,
                category=category,
                size_bytes=int(st.st_size),
                modified_utc=modified_utc,
                label=final_label,
                rule_label=rule_label,
                llm_label=llm_label,
                rule_score=rule_score,
                llm_score=llm_score,
                high_hits="; ".join(high_hits),
                suspicious_hits="; ".join(susp_hits),
                rationale=rationale,
                sorted_path=str(sorted_path),
                error=error,
            )
        )

    return rows


def _iter_files(src_root: Path, ignore_dirs: set[str]):
    for path in src_root.rglob("*"):
        if not path.is_file():
            continue
        rel_parts = {p.lower() for p in path.relative_to(src_root).parts[:-1]}
        if rel_parts.intersection(ignore_dirs):
            continue
        yield path


def _categorize(path: Path) -> str:
    suffix = path.suffix.lower()
    name = path.name.lower()
    if suffix in IMAGE_EXTENSIONS:
        return "photos"
    if suffix in VIDEO_EXTENSIONS:
        return "video"
    if suffix in TEXT_EXTENSIONS:
        if any(h in name for h in MESSAGE_HINTS):
            return "messages"
        return "documents"
    if any(h in name for h in MESSAGE_HINTS):
        return "messages"
    return "binary"


def _extract_text_sample(path: Path, *, max_text_chars: int, max_text_bytes: int) -> str:
    chunks: list[str] = [path.name]
    exif_line = _extract_exif_line(path)
    if exif_line:
        chunks.append(exif_line)

    if path.suffix.lower() in TEXT_EXTENSIONS:
        try:
            if path.stat().st_size <= max_text_bytes:
                raw = path.read_text(encoding="utf-8", errors="ignore")
                clean = re.sub(r"\s+", " ", raw).strip()
                if clean:
                    chunks.append(clean[:max_text_chars])
        except Exception:
            pass

    return " | ".join(chunks)[:max_text_chars]


def _extract_exif_line(path: Path) -> str:
    if Image is None or ExifTags is None or path.suffix.lower() not in IMAGE_EXTENSIONS:
        return ""
    try:
        with Image.open(path) as img:
            exif = img.getexif()
            if not exif:
                return ""
            tags = {ExifTags.TAGS.get(k, str(k)): v for k, v in exif.items()}
        dt = tags.get("DateTimeOriginal") or tags.get("DateTimeDigitized") or tags.get("DateTime")
        model = tags.get("Model", "")
        if dt or model:
            return f"EXIF date={dt or ''} model={model or ''}".strip()
    except Exception:
        return ""
    return ""


def _keyword_hits(text: str, keywords: list[str]) -> list[str]:
    hits: list[str] = []
    for kw in keywords:
        if kw and kw in text:
            hits.append(kw)
    return hits


def _rule_label(score: int, high_hits: list[str]) -> str:
    if high_hits or score >= 70:
        return "high_priority"
    if score >= 20:
        return "suspicious"
    return "safe"


def _normalize_label(label: str) -> str:
    token = (label or "").strip().lower().replace(" ", "_")
    if token in {"safe", "suspicious", "high_priority"}:
        return token
    if token in {"high", "critical", "urgent"}:
        return "high_priority"
    if token in {"warning", "medium"}:
        return "suspicious"
    return ""


def _merge_labels(rule_label: str, llm_label: str) -> str:
    if not llm_label:
        return rule_label
    return llm_label if SEVERITY.get(llm_label, 0) >= SEVERITY.get(rule_label, 0) else rule_label


def _place_file(*, path: Path, src_root: Path, triage_root: Path, label: str, copy_mode: str) -> Path:
    rel = path.relative_to(src_root)
    dest = triage_root / label / rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest = _unique_path(dest)

    if copy_mode == "link":
        try:
            os.link(path, dest)
            return dest
        except Exception:
            pointer = dest.with_suffix(dest.suffix + ".pointer.txt")
            pointer.write_text(str(path), encoding="utf-8")
            return pointer

    if copy_mode == "copy":
        try:
            shutil.copy2(path, dest)
            return dest
        except Exception:
            pointer = dest.with_suffix(dest.suffix + ".pointer.txt")
            pointer.write_text(str(path), encoding="utf-8")
            return pointer

    # default: link_or_copy
    try:
        os.link(path, dest)
        return dest
    except Exception:
        try:
            shutil.copy2(path, dest)
            return dest
        except Exception:
            pointer = dest.with_suffix(dest.suffix + ".pointer.txt")
            pointer.write_text(str(path), encoding="utf-8")
            return pointer


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    idx = 1
    while True:
        candidate = parent / f"{stem}__{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def _build_timeline(*, src_root: Path, analysis_cfg: dict[str, Any]) -> list[TimelineEvent]:
    ignore_dirs = {str(d).strip().lower() for d in (analysis_cfg.get("ignore_dirs") or []) if str(d).strip()}
    if not ignore_dirs:
        ignore_dirs = {"analysis", "triage", "timeline", "__pycache__"}

    max_files = max(1, int(analysis_cfg.get("max_files", 500)))
    max_events_per_file = max(1, int(analysis_cfg.get("timeline_max_events_per_file", 12)))
    max_scan_bytes = max(2048, int(analysis_cfg.get("timeline_max_text_bytes", 512 * 1024)))
    now_utc = datetime.now(timezone.utc)

    events: list[tuple[datetime, TimelineEvent]] = []
    for index, path in enumerate(_iter_files(src_root, ignore_dirs), start=1):
        if index > max_files:
            break
        st = path.stat()
        rel = path.relative_to(src_root).as_posix()

        mtime = datetime.fromtimestamp(st.st_mtime, tz=timezone.utc)
        events.append(
            (
                mtime,
                TimelineEvent(
                    timestamp_utc=mtime.isoformat(),
                    source_path=rel,
                    event_type="file_modified",
                    confidence="high",
                    summary=f"Filesystem modified time for {rel}",
                ),
            )
        )

        ctime = datetime.fromtimestamp(st.st_ctime, tz=timezone.utc)
        if abs((ctime - mtime).total_seconds()) > 1:
            events.append(
                (
                    ctime,
                    TimelineEvent(
                        timestamp_utc=ctime.isoformat(),
                        source_path=rel,
                        event_type="file_metadata_changed",
                        confidence="medium",
                        summary=f"Filesystem metadata change time for {rel}",
                    ),
                )
            )

        exif_dt = _extract_exif_datetime(path)
        if exif_dt is not None:
            events.append(
                (
                    exif_dt,
                    TimelineEvent(
                        timestamp_utc=exif_dt.isoformat(),
                        source_path=rel,
                        event_type="exif_timestamp",
                        confidence="high",
                        summary="Image EXIF capture timestamp",
                    ),
                )
            )

        if path.suffix.lower() in TEXT_EXTENSIONS:
            text_events = _extract_text_timestamps(path, rel, max_events_per_file, max_scan_bytes, now_utc)
            events.extend(text_events)

    events.sort(key=lambda item: item[0])
    return [item[1] for item in events]


def _extract_exif_datetime(path: Path) -> datetime | None:
    if Image is None or ExifTags is None or path.suffix.lower() not in IMAGE_EXTENSIONS:
        return None
    try:
        with Image.open(path) as img:
            exif = img.getexif()
            if not exif:
                return None
            tags = {ExifTags.TAGS.get(k, str(k)): v for k, v in exif.items()}
            raw = tags.get("DateTimeOriginal") or tags.get("DateTimeDigitized") or tags.get("DateTime")
            if not raw:
                return None
            # EXIF commonly uses YYYY:MM:DD HH:MM:SS in local timezone.
            parsed = datetime.strptime(str(raw), "%Y:%m:%d %H:%M:%S")
            return parsed.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _extract_text_timestamps(
    path: Path,
    rel: str,
    max_events_per_file: int,
    max_scan_bytes: int,
    now_utc: datetime,
) -> list[tuple[datetime, TimelineEvent]]:
    try:
        if path.stat().st_size > max_scan_bytes:
            return []
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return []

    out: list[tuple[datetime, TimelineEvent]] = []
    seen: set[str] = set()

    def append_event(ts: datetime, token: str, event_type: str, confidence: str) -> None:
        key = f"{ts.isoformat()}::{token}"
        if key in seen:
            return
        seen.add(key)
        snippet = _find_snippet(text, token)
        out.append(
            (
                ts,
                TimelineEvent(
                    timestamp_utc=ts.isoformat(),
                    source_path=rel,
                    event_type=event_type,
                    confidence=confidence,
                    summary=snippet,
                ),
            )
        )

    for match in ISO_TS.finditer(text):
        token = match.group(0)
        parsed = _parse_iso(token)
        if parsed is not None:
            append_event(parsed, token, "text_timestamp_iso", "high")
        if len(out) >= max_events_per_file:
            return out

    for match in SIMPLE_TS.finditer(text):
        token = match.group(0)
        parsed = _parse_simple(token)
        if parsed is not None:
            append_event(parsed, token, "text_timestamp_simple", "medium")
        if len(out) >= max_events_per_file:
            return out

    for match in LOGCAT_TS.finditer(text):
        token = match.group(0)
        parsed = _parse_logcat(token, year=now_utc.year)
        if parsed is not None:
            append_event(parsed, token, "text_timestamp_logcat", "medium")
        if len(out) >= max_events_per_file:
            return out

    for match in EPOCH_TS.finditer(text):
        token = match.group(0)
        parsed = _parse_epoch(token)
        if parsed is not None:
            append_event(parsed, token, "text_timestamp_epoch", "low")
        if len(out) >= max_events_per_file:
            return out

    return out


def _parse_iso(token: str) -> datetime | None:
    value = token.replace(" ", "T")
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def _parse_simple(token: str) -> datetime | None:
    value = token.replace("T", " ")
    try:
        dt = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        return dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _parse_logcat(token: str, year: int) -> datetime | None:
    try:
        dt = datetime.strptime(f"{year} {token}", "%Y %b %d %H:%M:%S")
        return dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _parse_epoch(token: str) -> datetime | None:
    try:
        raw = int(token)
        if raw > 1_000_000_000_000:
            raw = raw // 1000
        if raw < 946684800 or raw > 4_102_444_800:  # 2000-01-01 to 2100-01-01
            return None
        return datetime.fromtimestamp(raw, tz=timezone.utc)
    except Exception:
        return None


def _find_snippet(text: str, token: str, radius: int = 65) -> str:
    idx = text.find(token)
    if idx < 0:
        return token
    left = max(0, idx - radius)
    right = min(len(text), idx + len(token) + radius)
    snippet = re.sub(r"\s+", " ", text[left:right]).strip()
    prefix = "..." if left > 0 else ""
    suffix = "..." if right < len(text) else ""
    return f"{prefix}{snippet}{suffix}"


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, Any]], sheet_name: str) -> bool:
    if Workbook is None:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"
    wb.save(path)
    return True
